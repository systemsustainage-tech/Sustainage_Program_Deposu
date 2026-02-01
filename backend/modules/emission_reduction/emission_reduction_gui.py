import logging
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Emisyon Azaltma Projeleri GUI
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from utils.language_manager import LanguageManager
from .emission_reduction_manager import EmissionReductionManager
from config.database import DB_PATH


class EmissionReductionGUI:
    """Emisyon Azaltma Projeleri Modülü GUI"""

    def __init__(self, parent, company_id: int, db_path: str = DB_PATH) -> None:
        self.parent = parent
        self.company_id = company_id
        self.db_path = db_path
        self.lm = LanguageManager()
        self.manager = EmissionReductionManager(db_path)

        try:
            self.parent.winfo_toplevel().state('zoomed')
        except Exception as e:
            logging.error(f"Silent error caught: {str(e)}")

        self.setup_ui()
        self.load_data()

    def setup_ui(self) -> None:
        """Arayüzü oluştur"""
        # Ana frame
        main_frame = tk.Frame(self.parent, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)

        # Başlık
        header_frame = tk.Frame(main_frame, bg='#28a745', height=70)
        header_frame.pack(fill='x', pady=(0, 15))
        header_frame.pack_propagate(False)

        title_frame = tk.Frame(header_frame, bg='#28a745')
        title_frame.pack(side='left', padx=20, pady=15)

        title_label = tk.Label(title_frame, text=self.lm.tr("emission_reduction_projects", " Emisyon Azaltma Projeleri"),
                              font=('Segoe UI', 16, 'bold'), fg='white', bg='#28a745')
        title_label.pack(side='left')

        subtitle_label = tk.Label(title_frame, text=self.lm.tr("emission_reduction_subtitle", "Karbon ayak izini azaltan projelerin yönetimi"),
                                 font=('Segoe UI', 11), fg='#e8f5e8', bg='#28a745')
        subtitle_label.pack(side='left', padx=(10, 0))

        # Tab kontrolü
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)

        # Genel Bakış sekmesi
        self.overview_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.overview_frame, text=self.lm.tr("overview", "Genel Bakış"))
        self.create_overview_tab()

        # Projeler sekmesi
        self.projects_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.projects_frame, text=self.lm.tr("projects", "Projeler"))
        self.create_projects_tab()

        # Kategoriler sekmesi
        self.categories_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.categories_frame, text=self.lm.tr("categories", "Kategoriler"))
        self.create_categories_tab()

        # İlerleme Takibi sekmesi
        self.progress_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.progress_frame, text=self.lm.tr("progress_tracking", "İlerleme Takibi"))
        self.create_progress_tab()

        # Raporlar sekmesi
        self.reports_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.reports_frame, text=self.lm.tr("reports", "Raporlar"))
        self.create_reports_tab()

    def create_overview_tab(self) -> None:
        """Genel bakış sekmesini oluştur"""
        # İstatistik kartları
        stats_frame = tk.Frame(self.overview_frame, bg='white')
        stats_frame.pack(fill='x', padx=20, pady=20)

        # İstatistik verilerini al
        stats = self.manager.get_project_statistics(self.company_id)

        # İstatistik kartları
        stats_data = [
            (self.lm.tr("total_projects", "Toplam Proje"), stats['total_projects'], '#007bff'),
            (self.lm.tr("active_projects", "Aktif Proje"), stats['active_projects'], '#28a745'),
            (self.lm.tr("expected_reduction", "Beklenen Azaltma"), f"{stats['total_expected_reduction']:.1f} tCO2e", '#ffc107'),
            (self.lm.tr("actual_reduction", "Gerçekleşen Azaltma"), f"{stats['total_actual_reduction']:.1f} tCO2e", '#17a2b8')
        ]

        for i, (title, value, color) in enumerate(stats_data):
            card = tk.Frame(stats_frame, bg=color, relief='raised', bd=1)
            card.pack(side='left', fill='x', expand=True, padx=(0, 10) if i < len(stats_data)-1 else 0)

            card_content = tk.Frame(card, bg=color)
            card_content.pack(fill='both', expand=True, padx=15, pady=10)

            title_label = tk.Label(card_content, text=title, font=('Segoe UI', 9, 'bold'),
                                  fg='white', bg=color)
            title_label.pack()

            value_label = tk.Label(card_content, text=str(value), font=('Segoe UI', 16, 'bold'),
                                  fg='white', bg=color)
            value_label.pack()

        # Kategori dağılımı
        category_frame = tk.Frame(self.overview_frame, bg='white')
        category_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        tk.Label(category_frame, text=self.lm.tr("category_distribution", "Kategori Dağılımı"), font=('Segoe UI', 12, 'bold'),
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 10))

        # Kategori dağılımı tablosu
        category_tree = ttk.Treeview(category_frame, columns=('count',), show='headings', height=6)
        category_tree.heading('#0', text=self.lm.tr("category", 'Kategori'))
        category_tree.heading('count', text=self.lm.tr("project_count", 'Proje Sayısı'))
        category_tree.column('#0', width=200)
        category_tree.column('count', width=100)

        for category, count in stats['category_distribution'].items():
            category_tree.insert('', 'end', text=category, values=(count,))

        category_tree.pack(fill='both', expand=True)

    def create_projects_tab(self) -> None:
        """Projeler sekmesini oluştur"""
        # Butonlar
        button_frame = tk.Frame(self.projects_frame, bg='white')
        button_frame.pack(fill='x', padx=20, pady=20)

        tk.Button(button_frame, text=self.lm.tr("new_project", "+ Yeni Proje"), font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#28a745', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.show_new_project_form).pack(side='left', padx=(0, 10))

        tk.Button(button_frame, text=self.lm.tr("btn_edit", "️ Düzenle"), font=('Segoe UI', 11),
                 fg='white', bg='#007bff', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.edit_project).pack(side='left', padx=(0, 10))

        tk.Button(button_frame, text=self.lm.tr("update_progress", " İlerleme Güncelle"), font=('Segoe UI', 11),
                 fg='white', bg='#ffc107', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.show_progress_form).pack(side='left')

        # Projeler tablosu
        table_frame = tk.Frame(self.projects_frame, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        columns = ('project_name', 'category', 'status', 'expected_reduction', 'actual_reduction', 'progress', 'start_date', 'end_date')
        column_headers = {
            'project_name': self.lm.tr("project_name", 'Proje Adı'),
            'category': self.lm.tr("category", 'Kategori'),
            'status': self.lm.tr("status", 'Durum'),
            'expected_reduction': self.lm.tr("expected_reduction", 'Beklenen Azaltma'),
            'actual_reduction': self.lm.tr("actual_reduction", 'Gerçekleşen Azaltma'),
            'progress': self.lm.tr("progress", 'İlerleme'),
            'start_date': self.lm.tr("start_date", 'Başlangıç'),
            'end_date': self.lm.tr("end_date", 'Bitiş')
        }
        self.projects_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        for col in columns:
            self.projects_tree.heading(col, text=column_headers.get(col, col), anchor='center')
            self.projects_tree.column(col, width=120, anchor='center')

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.projects_tree.yview)
        self.projects_tree.configure(yscrollcommand=scrollbar.set)

        self.projects_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def create_categories_tab(self) -> None:
        """Kategoriler sekmesini oluştur"""
        # Kategoriler listesi
        categories_frame = tk.Frame(self.categories_frame, bg='white')
        categories_frame.pack(fill='both', expand=True, padx=20, pady=20)

        tk.Label(categories_frame, text=self.lm.tr("project_categories", "Proje Kategorileri"), font=('Segoe UI', 12, 'bold'),
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 10))

        categories = self.manager.get_categories()

        for category in categories:
            cat_frame = tk.Frame(categories_frame, bg='#f8f9fa', relief='solid', bd=1)
            cat_frame.pack(fill='x', pady=2)

            # Renk göstergesi
            color_frame = tk.Frame(cat_frame, bg=category['color'], width=20, height=20)
            color_frame.pack(side='left', padx=10, pady=10)

            # Kategori bilgileri
            info_frame = tk.Frame(cat_frame, bg='#f8f9fa')
            info_frame.pack(side='left', fill='x', expand=True, padx=10, pady=10)

            tk.Label(info_frame, text=category['name'], font=('Segoe UI', 11, 'bold'),
                    bg='#f8f9fa', fg='#2c3e50').pack(anchor='w')

            if category['description']:
                tk.Label(info_frame, text=category['description'], font=('Segoe UI', 9),
                        bg='#f8f9fa', fg='#6c757d').pack(anchor='w')

    def create_progress_tab(self) -> None:
        """İlerleme takibi sekmesini oluştur"""
        # Butonlar
        button_frame = tk.Frame(self.progress_frame, bg='white')
        button_frame.pack(fill='x', padx=20, pady=20)

        tk.Button(button_frame, text=self.lm.tr("add_progress_record", "+ İlerleme Kaydı Ekle"), font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#dc3545', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.show_progress_form).pack(side='left')

        # İlerleme kayıtları tablosu
        table_frame = tk.Frame(self.progress_frame, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        columns = ('project', 'date', 'progress_pct', 'actual_reduction', 'cost', 'notes')
        column_headers = {
            'project': self.lm.tr("project", 'Proje'),
            'date': self.lm.tr("date", 'Tarih'),
            'progress_pct': self.lm.tr("progress_pct", 'İlerleme (%)'),
            'actual_reduction': self.lm.tr("actual_reduction", 'Gerçekleşen Azaltma'),
            'cost': self.lm.tr("cost", 'Maliyet'),
            'notes': self.lm.tr("notes", 'Notlar')
        }
        self.progress_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        for col in columns:
            self.progress_tree.heading(col, text=column_headers.get(col, col), anchor='center')
            self.progress_tree.column(col, width=120, anchor='center')

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.progress_tree.yview)
        self.progress_tree.configure(yscrollcommand=scrollbar.set)

        self.progress_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def create_reports_tab(self) -> None:
        """Raporlar sekmesini oluştur"""
        # Butonlar
        button_frame = tk.Frame(self.reports_frame, bg='white')
        button_frame.pack(fill='x', padx=20, pady=20)

        tk.Button(button_frame, text=self.lm.tr("create_report", " Rapor Oluştur"), font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#6f42c1', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.show_report_form).pack(side='left', padx=(0, 10))

        tk.Button(button_frame, text=self.lm.tr("export_to_excel", " Excel'e Aktar"), font=('Segoe UI', 11),
                 fg='white', bg='#20c997', relief='flat', cursor='hand2',
                 padx=20, pady=8, command=self.export_to_excel).pack(side='left')

        # Raporlar listesi
        reports_frame = tk.Frame(self.reports_frame, bg='white')
        reports_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        tk.Label(reports_frame, text=self.lm.tr("available_reports", "Mevcut Raporlar"), font=('Segoe UI', 12, 'bold'),
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 10))

        # Raporlar tablosu
        columns = ('report_name', 'format', 'created_date', 'size', 'status')
        column_headers = {
            'report_name': self.lm.tr("report_name", 'Rapor Adı'),
            'format': self.lm.tr("format", 'Format'),
            'created_date': self.lm.tr("created_date", 'Oluşturulma Tarihi'),
            'size': self.lm.tr("size", 'Boyut'),
            'status': self.lm.tr("status", 'Durum')
        }
        self.reports_tree = ttk.Treeview(reports_frame, columns=columns, show='headings', height=10)

        for col in columns:
            self.reports_tree.heading(col, text=column_headers.get(col, col), anchor='center')
            self.reports_tree.column(col, width=120, anchor='center')

        # Scrollbar
        scrollbar = ttk.Scrollbar(reports_frame, orient='vertical', command=self.reports_tree.yview)
        self.reports_tree.configure(yscrollcommand=scrollbar.set)

        self.reports_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def load_data(self) -> None:
        """Verileri yükle"""
        self.load_projects()
        self.load_progress_records()

    def load_projects(self) -> None:
        """Projeleri yükle"""
        # Tabloyu temizle
        for item in self.projects_tree.get_children():
            self.projects_tree.delete(item)

        projects = self.manager.get_projects(self.company_id)

        for project in projects:
            self.projects_tree.insert('', 'end', values=(
                project['project_name'],
                project['category'],
                project['status'],
                f"{project['expected_reduction']:.1f} tCO2e",
                f"{project['actual_reduction']:.1f} tCO2e",
                f"{project['progress_percentage']:.1f}%",
                project['start_date'] or '-',
                project['end_date'] or '-'
            ))

    def load_progress_records(self) -> None:
        """İlerleme kayıtlarını yükle"""
        # Tabloyu temizle
        for item in self.progress_tree.get_children():
            self.progress_tree.delete(item)

        records = self.manager.get_progress_records()

        for record in records:
            self.progress_tree.insert('', 'end', values=(
                record['project_name'],
                record['progress_date'],
                f"{record['progress_percentage']:.1f}%",
                f"{record['actual_reduction']:.1f} tCO2e",
                f"{record['cost']:.0f} TL" if record['cost'] else '-',
                record['notes'] or '-'
            ))

    def show_new_project_form(self) -> None:
        """Yeni proje ekleme formunu göster"""
        form_window = tk.Toplevel(self.parent)
        form_window.title(self.lm.tr("new_emission_project", "Yeni Emisyon Azaltma Projesi"))
        form_window.geometry("600x500")
        form_window.configure(bg='#f8f9fa')
        form_window.resizable(False, False)

        # Modal yap
        form_window.transient(self.parent)
        form_window.grab_set()

        # Pencereyi ortala
        form_window.update_idletasks()
        x = (form_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (form_window.winfo_screenheight() // 2) - (500 // 2)
        form_window.geometry(f"600x500+{x}+{y}")

        # Başlık
        header_frame = tk.Frame(form_window, bg='#28a745', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text=self.lm.tr("new_emission_project", " Yeni Emisyon Azaltma Projesi"),
                font=('Segoe UI', 16, 'bold'), fg='white', bg='#28a745').pack(expand=True)

        # Form içeriği
        form_frame = tk.Frame(form_window, bg='#f8f9fa')
        form_frame.pack(fill='both', expand=True, padx=30, pady=20)

        # Form değişkenleri
        self.form_vars = {
            'project_name': tk.StringVar(),
            'category': tk.StringVar(),
            'description': tk.StringVar(),
            'start_date': tk.StringVar(),
            'end_date': tk.StringVar(),
            'expected_reduction': tk.StringVar(),
            'budget': tk.StringVar(),
            'responsible_person': tk.StringVar()
        }

        # Proje Adı
        tk.Label(form_frame, text=self.lm.tr("project_name_label", "Proje Adı:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(form_frame, textvariable=self.form_vars['project_name'],
                font=('Segoe UI', 10), width=50).pack(fill='x', pady=(0, 15))

        # Kategori
        tk.Label(form_frame, text=self.lm.tr("category_label", "Kategori:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))

        categories = self.manager.get_categories()
        category_names = [cat['name'] for cat in categories]

        category_combo = ttk.Combobox(form_frame, textvariable=self.form_vars['category'],
                                    values=category_names, font=('Segoe UI', 10), state='readonly', width=47)
        category_combo.pack(fill='x', pady=(0, 15))

        # Açıklama
        tk.Label(form_frame, text=self.lm.tr("description_label", "Açıklama:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        description_text = tk.Text(form_frame, height=3, font=('Segoe UI', 10), width=50)
        description_text.pack(fill='x', pady=(0, 15))

        # Tarihler
        date_frame = tk.Frame(form_frame, bg='#f8f9fa')
        date_frame.pack(fill='x', pady=(0, 15))

        tk.Label(date_frame, text=self.lm.tr("start_date_label", "Başlangıç Tarihi:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(date_frame, textvariable=self.form_vars['start_date'],
                font=('Segoe UI', 10), width=20).pack(side='left', padx=(0, 20))

        tk.Label(date_frame, text=self.lm.tr("end_date_label", "Bitiş Tarihi:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(date_frame, textvariable=self.form_vars['end_date'],
                font=('Segoe UI', 10), width=20).pack(side='left')

        # Beklenen Azaltma ve Bütçe
        value_frame = tk.Frame(form_frame, bg='#f8f9fa')
        value_frame.pack(fill='x', pady=(0, 15))

        tk.Label(value_frame, text=self.lm.tr("expected_reduction_unit", "Beklenen Azaltma (tCO2e):"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(value_frame, textvariable=self.form_vars['expected_reduction'],
                font=('Segoe UI', 10), width=20).pack(side='left', padx=(0, 20))

        tk.Label(value_frame, text=self.lm.tr("budget_unit", "Bütçe (TL):"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(value_frame, textvariable=self.form_vars['budget'],
                font=('Segoe UI', 10), width=20).pack(side='left')

        # Sorumlu Kişi
        tk.Label(form_frame, text=self.lm.tr("responsible_person", "Sorumlu Kişi:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(form_frame, textvariable=self.form_vars['responsible_person'],
                font=('Segoe UI', 10), width=50).pack(fill='x', pady=(0, 20))

        # Butonlar
        button_frame = tk.Frame(form_frame, bg='#f8f9fa')
        button_frame.pack(fill='x', pady=(20, 0))

        tk.Button(button_frame, text=self.lm.tr("btn_cancel", " İptal"), font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#6c757d', relief='flat', cursor='hand2',
                 padx=30, pady=10, command=form_window.destroy).pack(side='left')

        tk.Button(button_frame, text=self.lm.tr("btn_save", " Kaydet"), font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#28a745', relief='flat', cursor='hand2',
                 padx=30, pady=10, command=lambda: self.save_project(form_window, description_text)).pack(side='right')

    def save_project(self, window, description_widget) -> None:
        """Projeyi kaydet"""
        try:
            # Validasyon
            if not self.form_vars['project_name'].get().strip():
                messagebox.showerror("Hata", "Proje adı boş olamaz!")
                return

            if not self.form_vars['category'].get().strip():
                messagebox.showerror("Hata", "Kategori seçimi zorunludur!")
                return

            # Projeyi kaydet
            project_data = {
                'project_name': self.form_vars['project_name'].get().strip(),
                'category': self.form_vars['category'].get(),
                'description': description_widget.get('1.0', tk.END).strip(),
                'start_date': self.form_vars['start_date'].get().strip() or None,
                'end_date': self.form_vars['end_date'].get().strip() or None,
                'expected_reduction': float(self.form_vars['expected_reduction'].get() or 0),
                'budget': float(self.form_vars['budget'].get() or 0),
                'responsible_person': self.form_vars['responsible_person'].get().strip() or None
            }

            project_id = self.manager.add_project(self.company_id, **project_data)

            if project_id:
                messagebox.showinfo("Başarılı", "Proje başarıyla eklendi!")
                window.destroy()
                self.load_data()
            else:
                messagebox.showerror("Hata", "Proje eklenirken hata oluştu!")

        except ValueError as e:
            messagebox.showerror("Hata", f"Geçersiz veri: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Proje kaydetme hatası: {e}")

    def show_progress_form(self) -> None:
        """İlerleme kaydı ekleme formunu göster"""
        form_window = tk.Toplevel(self.parent)
        form_window.title("İlerleme Kaydı Ekle")
        form_window.geometry("500x400")
        form_window.configure(bg='#f8f9fa')
        form_window.resizable(False, False)

        # Modal yap
        form_window.transient(self.parent)
        form_window.grab_set()

        # Pencereyi ortala
        form_window.update_idletasks()
        x = (form_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (form_window.winfo_screenheight() // 2) - (400 // 2)
        form_window.geometry(f"500x400+{x}+{y}")

        # Başlık
        header_frame = tk.Frame(form_window, bg='#dc3545', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text=" İlerleme Kaydı Ekle",
                font=('Segoe UI', 16, 'bold'), fg='white', bg='#dc3545').pack(expand=True)

        # Form içeriği
        form_frame = tk.Frame(form_window, bg='#f8f9fa')
        form_frame.pack(fill='both', expand=True, padx=30, pady=20)

        # Form değişkenleri
        self.progress_vars = {
            'project_id': tk.StringVar(),
            'progress_date': tk.StringVar(),
            'progress_percentage': tk.StringVar(),
            'actual_reduction': tk.StringVar(),
            'cost': tk.StringVar(),
            'notes': tk.StringVar()
        }

        # Proje Seçimi
        tk.Label(form_frame, text="Proje:", font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))

        projects = self.manager.get_projects(self.company_id)
        project_options = [f"{p['id']} - {p['project_name']}" for p in projects]

        project_combo = ttk.Combobox(form_frame, values=project_options,
                                   font=('Segoe UI', 10), state='readonly', width=50)
        project_combo.pack(fill='x', pady=(0, 15))

        # Tarih ve İlerleme
        row1_frame = tk.Frame(form_frame, bg='#f8f9fa')
        row1_frame.pack(fill='x', pady=(0, 15))

        tk.Label(row1_frame, text=self.lm.tr("date_label", "Tarih:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(row1_frame, textvariable=self.progress_vars['progress_date'],
                font=('Segoe UI', 10), width=20).pack(side='left', padx=(0, 20))

        tk.Label(row1_frame, text=self.lm.tr("progress_pct_label", "İlerleme (%):"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(row1_frame, textvariable=self.progress_vars['progress_percentage'],
                font=('Segoe UI', 10), width=20).pack(side='left')

        # Azaltma ve Maliyet
        row2_frame = tk.Frame(form_frame, bg='#f8f9fa')
        row2_frame.pack(fill='x', pady=(0, 15))

        tk.Label(row2_frame, text=self.lm.tr("actual_reduction_unit", "Gerçekleşen Azaltma (tCO2e):"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(row2_frame, textvariable=self.progress_vars['actual_reduction'],
                font=('Segoe UI', 10), width=20).pack(side='left', padx=(0, 20))

        tk.Label(row2_frame, text=self.lm.tr("cost_unit", "Maliyet (TL):"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        tk.Entry(row2_frame, textvariable=self.progress_vars['cost'],
                font=('Segoe UI', 10), width=20).pack(side='left')

        # Notlar
        tk.Label(form_frame, text=self.lm.tr("notes_label", "Notlar:"), font=('Segoe UI', 11, 'bold'),
                bg='#f8f9fa', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        notes_text = tk.Text(form_frame, height=3, font=('Segoe UI', 10), width=50)
        notes_text.pack(fill='x', pady=(0, 20))

        # Butonlar
        button_frame = tk.Frame(form_frame, bg='#f8f9fa')
        button_frame.pack(fill='x', pady=(20, 0))

        tk.Button(button_frame, text=" İptal", font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#6c757d', relief='flat', cursor='hand2',
                 padx=30, pady=10, command=form_window.destroy).pack(side='left')

        tk.Button(button_frame, text=" Kaydet", font=('Segoe UI', 11, 'bold'),
                 fg='white', bg='#dc3545', relief='flat', cursor='hand2',
                 padx=30, pady=10, command=lambda: self.save_progress(form_window, notes_text, project_combo)).pack(side='right')

    def save_progress(self, window, notes_widget, project_combo) -> None:
        """İlerleme kaydını kaydet"""
        try:
            # Validasyon
            if not project_combo.get().strip():
                messagebox.showerror("Hata", "Proje seçimi zorunludur!")
                return

            if not self.progress_vars['progress_date'].get().strip():
                messagebox.showerror("Hata", "Tarih boş olamaz!")
                return

            if not self.progress_vars['progress_percentage'].get().strip():
                messagebox.showerror("Hata", "İlerleme yüzdesi boş olamaz!")
                return

            # Proje ID'sini al
            project_id = int(project_combo.get().split(' - ')[0])

            # İlerleme kaydını kaydet
            progress_data = {
                'progress_date': self.progress_vars['progress_date'].get().strip(),
                'progress_percentage': float(self.progress_vars['progress_percentage'].get()),
                'actual_reduction': float(self.progress_vars['actual_reduction'].get() or 0),
                'cost': float(self.progress_vars['cost'].get() or 0),
                'notes': notes_widget.get('1.0', tk.END).strip() or None
            }

            progress_id = self.manager.add_progress(project_id, **progress_data)

            if progress_id:
                messagebox.showinfo("Başarılı", "İlerleme kaydı başarıyla eklendi!")
                window.destroy()
                self.load_data()
            else:
                messagebox.showerror("Hata", "İlerleme kaydı eklenirken hata oluştu!")

        except ValueError as e:
            messagebox.showerror("Hata", f"Geçersiz veri: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"İlerleme kaydı kaydetme hatası: {e}")

    def show_report_form(self) -> None:
        """Rapor oluşturma formunu göster"""
        # Rapor özeti oluştur
        summary = self.lm.tr("emission_report_title", "Emisyon Azaltma Raporu") + "\n\n"
        summary += f"{self.lm.tr('company_id', 'Şirket ID')}: {self.company_id}\n"
        summary += self.lm.tr("report_detail_hint", "Detaylı rapor için:\nSol menü → Raporlar → Emisyon Raporu")

        messagebox.showinfo(self.lm.tr("report_label", "Rapor"), summary)

    def export_to_excel(self) -> None:
        """Excel'e aktar"""
        import datetime

        try:
            filename = filedialog.asksaveasfilename(
                title=self.lm.tr('save_report', "Raporu Kaydet"),
                defaultextension=".xlsx",
                filetypes=[(self.lm.tr('file_excel', "Excel Dosyası"), "*.xlsx")],
                initialfile=f"emisyon_azaltma_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            if filename:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = self.lm.tr("emission_reduction", "Emisyon Azaltma")

                ws['A1'] = self.lm.tr("emission_reduction_data", "Emisyon Azaltma Verileri")
                ws['A2'] = f"{self.lm.tr('date_label', 'Tarih')}: {datetime.datetime.now().strftime('%d.%m.%Y')}"

                wb.save(filename)
                messagebox.showinfo(self.lm.tr("success", "Başarılı"), f"{self.lm.tr('excel_export_success', 'Excel dosyası oluşturuldu')}:\n{filename}")
        except Exception as e:
            messagebox.showerror(self.lm.tr("error", "Hata"), f"{self.lm.tr('export_error', 'Export hatası')}: {e}")

    def edit_project(self) -> None:
        """Proje düzenle"""
        selection = self.projects_tree.selection()
        if not selection:
            messagebox.showwarning(self.lm.tr("warning", "Uyarı"), self.lm.tr("select_project_to_edit", "Lütfen düzenlemek istediğiniz projeyi seçin!"))
            return

        messagebox.showinfo(self.lm.tr("info", "Bilgi"), self.lm.tr("edit_mode_enabled", "Proje düzenleme moduna alındı"))

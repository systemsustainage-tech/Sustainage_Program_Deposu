#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UNGC Excel Export Module
Export UNGC data to Excel format
"""
import logging
import os
from datetime import datetime
from typing import Dict, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.ungc.ungc_manager_enhanced import UNGCManagerEnhanced
from config.database import DB_PATH


class UNGCExcelExporter:
    """Export UNGC compliance data to Excel"""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.manager = UNGCManagerEnhanced(db_path)

    def export_compliance_report(self, company_id: int, period: Optional[str] = None, output_path: Optional[str] = None) -> str:
        """
        Export full compliance report to Excel
        
        Returns:
            Path to generated Excel file
        """
        if period is None:
            period = str(datetime.now().year)

        # Get company info
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        result = cursor.execute("SELECT name FROM companies WHERE id = ?", (company_id,)).fetchone()
        company_name = result[0] if result else "Company"
        conn.close()

        # Get compliance data
        overall_data = self.manager.calculate_overall_score(company_id, period)

        # Create workbook
        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, company_name, period, overall_data)

        # Sheet 2: KPI Data
        ws_kpi = wb.create_sheet("KPI Data")
        self._create_kpi_sheet(ws_kpi, company_id, period)

        # Sheet 3: Category Scores
        ws_cat = wb.create_sheet("Category Scores")
        self._create_category_sheet(ws_cat, overall_data)

        # Sheet 4: Principle Scores
        ws_prin = wb.create_sheet("Principle Scores")
        self._create_principle_sheet(ws_prin, overall_data)

        # Output path
        if output_path is None:
            os.makedirs('reports/ungc/excel', exist_ok=True)
            output_path = f"reports/ungc/excel/UNGC_Report_{company_name.replace(' ', '_')}_{period}.xlsx"

        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, ws, company_name: str, period: str, data: Dict):
        """Create summary sheet"""
        # Title
        ws['A1'] = "UNGC Compliance Report - Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Company info
        ws['A3'] = "Company:"
        ws['B3'] = company_name
        ws['A4'] = "Period:"
        ws['B4'] = period
        ws['A5'] = "Report Date:"
        ws['B5'] = datetime.now().strftime('%Y-%m-%d')

        # Overall score
        ws['A7'] = "Overall Score:"
        ws['B7'] = f"{data['overall_score']:.2f}%"
        ws['B7'].font = Font(size=14, bold=True, color="0000FF")

        ws['A8'] = "Level:"
        ws['B8'] = data['level']
        ws['B8'].font = Font(size=14, bold=True)

        # Category scores table
        ws['A10'] = "Category Scores"
        ws['A10'].font = Font(size=12, bold=True)

        row = 11
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Score (%)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        for category, score in data['category_scores'].items():
            row += 1
            ws[f'A{row}'] = category
            ws[f'B{row}'] = f"{score:.2f}"

        # Auto width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_kpi_sheet(self, ws, company_id: int, period: str):
        """Create KPI data sheet"""
        # Header
        ws['A1'] = "KPI Data"
        ws['A1'].font = Font(size=14, bold=True)

        # Table headers
        headers = ['Principle', 'KPI ID', 'KPI Name', 'Type', 'Current Value', 'Target', 'Score (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        # Get all principles and their KPIs
        row = 4
        for principle in self.manager.config.get('principles', []):
            principle_id = principle['id']
            result = self.manager.calculate_principle_score(company_id, principle_id, period)

            for kpi_data in result['kpis']:
                ws.cell(row, 1, principle_id)
                ws.cell(row, 2, kpi_data['kpi_id'])
                ws.cell(row, 3, kpi_data['name'])
                ws.cell(row, 4, kpi_data['type'])
                ws.cell(row, 5, kpi_data['current_value'] if kpi_data['current_value'] is not None else 'N/A')
                ws.cell(row, 6, kpi_data['target'] if kpi_data['target'] is not None else 'N/A')
                ws.cell(row, 7, f"{kpi_data['score']:.2f}")
                row += 1

        # Auto width
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_category_sheet(self, ws, data: Dict):
        """Create category scores sheet"""
        ws['A1'] = "Category Scores"
        ws['A1'].font = Font(size=14, bold=True)

        # Table
        row = 3
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Score (%)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        for category, score in data['category_scores'].items():
            row += 1
            ws[f'A{row}'] = category
            ws[f'B{row}'] = f"{score:.2f}"

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_principle_sheet(self, ws, data: Dict):
        """Create principle scores sheet"""
        ws['A1'] = "Principle Scores"
        ws['A1'].font = Font(size=14, bold=True)

        # Table
        row = 3
        headers = ['Principle', 'Title', 'Category', 'Score (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

        for p_score in data['principle_scores']:
            row += 1
            ws.cell(row, 1, p_score['principle_id'])
            ws.cell(row, 2, p_score['title'])
            ws.cell(row, 3, p_score['category'])
            ws.cell(row, 4, f"{p_score['score']:.2f}")

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15


if __name__ == '__main__':
    import sys
    db = sys.argv[1] if len(sys.argv) > 1 else DB_PATH

    exporter = UNGCExcelExporter(db)
    output = exporter.export_compliance_report(company_id=1)
    logging.info(f"Excel report generated: {output}")


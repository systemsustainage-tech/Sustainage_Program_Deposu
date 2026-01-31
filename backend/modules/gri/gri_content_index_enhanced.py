import logging
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GRI Content Index Enhanced - İyileştirilmiş Versiyon
- Renk kodlaması
- Grafikler ve görselleştirme
- Özet dashboard
- İlerleme takibi
"""

import os
import sqlite3
from datetime import datetime
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from config.database import DB_PATH


class GRIContentIndexEnhanced:
    """GRI Content Index - Gelişmiş Versiyon"""

    # Renk paleti
    COLORS = {
        'completed': 'C6EFCE',      # Açık yeşil
        'partial': 'FFEB9C',        # Açık sarı
        'not_reported': 'FFC7CE',   # Açık kırmızı
        'header': '4472C4',         # Koyu mavi
        'subheader': '5B9BD5',      # Orta mavi
        'border': '000000'          # Siyah
    }

    def __init__(self, db_path: str = DB_PATH) -> None:
        """İnit"""
        if not os.path.isabs(db_path):
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
            db_path = os.path.join(base_dir, db_path)
        self.db_path = db_path
        try:
            os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        except Exception as e:
            logging.error(f"Silent error caught: {str(e)}")

    def get_connection(self):
        """Veritabanı bağlantısı"""
        return sqlite3.connect(self.db_path)

    def generate_content_index(self, company_id: int = 1) -> Dict:
        """GRI Content Index oluştur"""
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            logging.info(" GRI Content Index Enhanced oluşturuluyor...")

            # Tüm göstergeleri al
            cursor.execute("""
                SELECT 
                    gi.id,
                    gi.code as disclosure_code,
                    gi.title as disclosure_title,
                    gi.description,
                    gi.unit,
                    gi.methodology,
                    gi.reporting_requirement,
                    gi.priority,
                    gi.requirement_level,
                    gi.reporting_frequency,
                    gi.data_quality,
                    gi.audit_required,
                    gi.validation_required,
                    gi.digitalization_status,
                    gi.cost_level,
                    gi.time_requirement,
                    gi.expertise_requirement,
                    gi.sustainability_impact,
                    gi.legal_compliance,
                    gi.sector_specific,
                    gi.international_standard,
                    gi.metric_type,
                    gi.scale_unit,
                    gi.data_source_system,
                    gi.reporting_format,
                    gi.tsrs_esrs_mapping,
                    gi.un_sdg_mapping,
                    gi.gri_3_3_reference,
                    gi.impact_area,
                    gi.stakeholder_group,
                    gs.code as standard_code,
                    gs.title as standard_title,
                    gs.category,
                    gs.type as standard_type,
                    gs.sub_category
                FROM gri_indicators gi
                JOIN gri_standards gs ON gi.standard_id = gs.id
                ORDER BY gs.category, gs.code, gi.code
            """)

            indicators = cursor.fetchall()

            # Kategori bazında organize et
            content_index = {
                'universal': [],
                'economic': [],
                'environmental': [],
                'social': [],
                'sector': []
            }

            for indicator in indicators:
                category = indicator[31]  # gs.category
                if category == 'Universal':
                    content_index['universal'].append(indicator)
                elif category == 'Economic':
                    content_index['economic'].append(indicator)
                elif category == 'Environmental':
                    content_index['environmental'].append(indicator)
                elif category == 'Social':
                    content_index['social'].append(indicator)
                elif category == 'Sector-Specific':
                    content_index['sector'].append(indicator)

            # Company responses al
            company_responses = self.get_company_responses(company_id)

            # İstatistikleri hesapla
            stats = self.calculate_statistics(content_index, company_responses)

            return {
                'indicators': content_index,
                'responses': company_responses,
                'statistics': stats,
                'generated_at': datetime.now().isoformat(),
                'total_indicators': len(indicators)
            }

        except Exception as e:
            logging.error(f" Content Index oluşturma hatası: {e}")
            return {}
        finally:
            conn.close()

    def get_company_responses(self, company_id: int) -> Dict:
        """Şirket yanıtlarını al"""
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT 
                    gr.indicator_id,
                    gr.period,
                    gr.response_value,
                    gr.numerical_value,
                    gr.unit,
                    gr.methodology,
                    gr.reporting_status,
                    gr.evidence_url,
                    gr.notes,
                    gi.code as disclosure_code
                FROM gri_responses gr
                JOIN gri_indicators gi ON gr.indicator_id = gi.id
                WHERE gr.company_id = ?
                ORDER BY gi.code
            """, (company_id,))

            responses = cursor.fetchall()

            response_dict = {}
            for response in responses:
                indicator_id = response[0]
                # Veri kalitesi kontrolü - kısmi mi tam mı?
                response_value = response[2]
                numerical_value = response[3]

                # Basit kalite skoru
                quality_score = 0
                if response_value and str(response_value).strip():
                    quality_score += 50
                if numerical_value is not None:
                    quality_score += 50

                response_dict[indicator_id] = {
                    'period': response[1],
                    'response_value': response[2],
                    'numerical_value': response[3],
                    'unit': response[4],
                    'methodology': response[5],
                    'reporting_status': response[6],
                    'evidence_url': response[7],
                    'notes': response[8],
                    'disclosure_code': response[9],
                    'quality_score': quality_score  # 0-100
                }

            return response_dict

        except Exception as e:
            logging.error(f" Company responses alma hatası: {e}")
            return {}
        finally:
            conn.close()

    def calculate_statistics(self, content_index: Dict, responses: Dict) -> Dict:
        """İstatistikleri hesapla"""
        stats = {}

        for category, indicators in content_index.items():
            total = len(indicators)
            completed = sum(1 for ind in indicators if ind[0] in responses and responses[ind[0]]['quality_score'] >= 90)
            partial = sum(1 for ind in indicators if ind[0] in responses and 0 < responses[ind[0]]['quality_score'] < 90)
            not_reported = total - completed - partial

            stats[category] = {
                'total': total,
                'completed': completed,
                'partial': partial,
                'not_reported': not_reported,
                'completion_rate': round((completed / total * 100) if total > 0 else 0, 1),
                'overall_rate': round(((completed + partial * 0.5) / total * 100) if total > 0 else 0, 1)
            }

        return stats

    def export_to_excel(self, output_path: str, company_id: int = 1) -> bool:
        """Content Index'i Excel'e export et - Enhanced"""
        try:
            logging.info(f" GRI Content Index Enhanced Excel export başlıyor: {output_path}")

            # Content Index oluştur
            content_data = self.generate_content_index(company_id)
            if not content_data:
                logging.info(" Content Index oluşturulamadı!")
                return False

            # Çıkış klasörünü oluştur
            out_dir = os.path.dirname(output_path)
            if out_dir:
                try:
                    os.makedirs(out_dir, exist_ok=True)
                except Exception as e:
                    logging.error(f"Silent error caught: {str(e)}")

            # Excel writer oluştur
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                # 1. Dashboard Sayfası (YENİ!)
                self.create_dashboard_sheet(writer, content_data)

                # 2. Özet Sayfası (İyileştirilmiş)
                self.create_summary_sheet_enhanced(writer, content_data)

                # 3-7. Kategori Sayfaları (Renk kodlamalı)
                self.create_category_sheet_enhanced(writer, content_data['indicators']['universal'],
                                         'Universal Standards', content_data['responses'])

                self.create_category_sheet_enhanced(writer, content_data['indicators']['economic'],
                                         'Economic Standards', content_data['responses'])

                self.create_category_sheet_enhanced(writer, content_data['indicators']['environmental'],
                                         'Environmental Standards', content_data['responses'])

                self.create_category_sheet_enhanced(writer, content_data['indicators']['social'],
                                         'Social Standards', content_data['responses'])

                self.create_category_sheet_enhanced(writer, content_data['indicators']['sector'],
                                         'Sector-Specific Standards', content_data['responses'])

                # 8. Mapping Sayfası
                self.create_mapping_sheet(writer, content_data)

            logging.info(f" GRI Content Index Enhanced başarıyla export edildi: {output_path}")
            return True

        except Exception as e:
            logging.error(f" Excel export hatası: {e}")
            import traceback
            traceback.print_exc()
            return False

    def create_dashboard_sheet(self, writer, content_data) -> None:
        """Dashboard sayfası oluştur - YENİ!"""
        logging.info(" Dashboard oluşturuluyor...")

        stats = content_data['statistics']

        # Dashboard verisi
        dashboard_data = []
        dashboard_data.append(['GRI CONTENT INDEX DASHBOARD', '', '', '', ''])
        dashboard_data.append(['Oluşturma Tarihi:', content_data['generated_at'], '', '', ''])
        dashboard_data.append(['Toplam Gösterge:', content_data['total_indicators'], '', '', ''])
        dashboard_data.append(['', '', '', '', ''])

        # Genel özet
        dashboard_data.append(['KATEGORİ', 'TOPLAM', 'TAMAMLANAN', 'KISMİ', 'EKSİK', 'TAMAMLANMA (%)'])

        category_names = {
            'universal': 'Universal Standards',
            'economic': 'Economic Standards',
            'environmental': 'Environmental Standards',
            'social': 'Social Standards',
            'sector': 'Sector-Specific'
        }

        for category, name in category_names.items():
            s = stats.get(category, {})
            dashboard_data.append([
                name,
                s.get('total', 0),
                s.get('completed', 0),
                s.get('partial', 0),
                s.get('not_reported', 0),
                s.get('completion_rate', 0)
            ])

        # Toplamlar
        total_all = sum(s.get('total', 0) for s in stats.values())
        completed_all = sum(s.get('completed', 0) for s in stats.values())
        partial_all = sum(s.get('partial', 0) for s in stats.values())
        not_reported_all = sum(s.get('not_reported', 0) for s in stats.values())
        completion_rate_all = round((completed_all / total_all * 100) if total_all > 0 else 0, 1)

        dashboard_data.append(['', '', '', '', '', ''])
        dashboard_data.append(['TOPLAM', total_all, completed_all, partial_all, not_reported_all, completion_rate_all])

        # DataFrame oluştur
        df = pd.DataFrame(dashboard_data)
        df.to_excel(writer, sheet_name='Dashboard', index=False, header=False)

        # Formatla
        worksheet = writer.sheets['Dashboard']

        # Başlık formatı
        for col in range(1, 6):
            cell = worksheet.cell(1, col)
            cell.font = Font(bold=True, size=14, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Tablo başlık formatı
        for col in range(1, 7):
            cell = worksheet.cell(5, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Veri satırlarını renklendir
        for row in range(6, 11):  # Kategori satırları
            completion_rate = worksheet.cell(row, 6).value
            if completion_rate is not None and isinstance(completion_rate, (int, float)):
                if completion_rate >= 80:
                    color = self.COLORS['completed']
                elif completion_rate >= 50:
                    color = self.COLORS['partial']
                else:
                    color = self.COLORS['not_reported']

                for col in range(1, 7):
                    cell = worksheet.cell(row, col)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Toplam satırı kalın
        for col in range(1, 7):
            cell = worksheet.cell(13, col)
            cell.font = Font(bold=True)

        # Kolon genişlikleri
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 18

        logging.info(" Dashboard oluşturuldu!")

    def create_summary_sheet_enhanced(self, writer, content_data) -> None:
        """Özet sayfası oluştur - İyileştirilmiş"""
        logging.info(" Özet sayfası oluşturuluyor...")

        stats = content_data['statistics']
        summary_data = []

        category_names = {
            'universal': 'Universal Standards',
            'economic': 'Economic Standards',
            'environmental': 'Environmental Standards',
            'social': 'Social Standards',
            'sector': 'Sector-Specific Standards'
        }

        for category, name in category_names.items():
            s = stats.get(category, {})
            summary_data.append({
                'Kategori': name,
                'Toplam Gösterge': s.get('total', 0),
                'Tamamlanan': s.get('completed', 0),
                'Kısmi': s.get('partial', 0),
                'Eksik': s.get('not_reported', 0),
                'Tamamlanma Oranı (%)': s.get('completion_rate', 0),
                'Genel Puan (%)': s.get('overall_rate', 0),
                'Son Güncelleme': content_data['generated_at']
            })

        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Özet', index=False)

        # Formatla
        worksheet = writer.sheets['Özet']

        # Başlık satırı
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Veri satırlarını renklendir
        for row_num in range(2, len(df) + 2):
            completion_rate = worksheet.cell(row_num, 6).value  # Tamamlanma Oranı kolonu

            if completion_rate >= 80:
                color = self.COLORS['completed']
            elif completion_rate >= 50:
                color = self.COLORS['partial']
            else:
                color = self.COLORS['not_reported']

            # Tüm satırı renklendir
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row_num, col_num)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Kolon genişlikleri
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logging.error(f"Silent error caught: {str(e)}")
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        logging.info(" Özet sayfası oluşturuldu!")

    def create_category_sheet_enhanced(self, writer, indicators, sheet_name, responses) -> None:
        """Kategori sayfası oluştur - Renk kodlamalı"""
        logging.info(f" {sheet_name} oluşturuluyor...")

        if not indicators:
            df = pd.DataFrame({'Mesaj': ['Bu kategoride gösterge bulunmamaktadır.']})
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            return

        sheet_data = []

        for indicator in indicators:
            indicator_id = indicator[0]
            response = responses.get(indicator_id, {})
            quality_score = response.get('quality_score', 0)

            # Durum belirleme
            if quality_score >= 90:
                status = ' Tam Raporlandı'
                status_color = 'completed'
            elif quality_score > 0:
                status = '️ Kısmi Raporlandı'
                status_color = 'partial'
            else:
                status = ' Raporlanmadı'
                status_color = 'not_reported'

            sheet_data.append({
                'Durum': status,
                'GRI Kodu': indicator[1],  # disclosure_code
                'Başlık': indicator[2],  # disclosure_title
                'Açıklama': indicator[3],  # description
                'GRI 3-3 Referansı': indicator[27],  # gri_3_3_reference
                'Öncelik': indicator[7],  # priority
                'Gereklilik': indicator[8],  # requirement_level
                'Yanıt': response.get('response_value', ''),
                'Sayısal Değer': response.get('numerical_value', ''),
                'Birim': response.get('unit', ''),
                'Kanıt': response.get('evidence_url', ''),
                'Notlar': response.get('notes', ''),
                'Sayfa No': '',  # Manuel doldurulacak
                'Kalite Skoru (%)': quality_score if response else 0,
                '_status_color': status_color  # İç kullanım
            })

        df = pd.DataFrame(sheet_data)

        # _status_color kolonunu kaldır (sadece renklendirme için kullanıldı)
        status_colors = df['_status_color'].tolist()
        df = df.drop(columns=['_status_color'])

        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Formatla
        worksheet = writer.sheets[sheet_name]

        # Başlık satırı
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Veri satırlarını renklendir
        for row_num, color_key in enumerate(status_colors, start=2):
            color = self.COLORS.get(color_key, 'FFFFFF')

            # Sadece ilk 3 kolonu renklendir (Durum, GRI Kodu, Başlık)
            for col_num in range(1, 4):
                cell = worksheet.cell(row_num, col_num)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Kolon genişlikleri
        worksheet.column_dimensions['A'].width = 20  # Durum
        worksheet.column_dimensions['B'].width = 15  # GRI Kodu
        worksheet.column_dimensions['C'].width = 40  # Başlık
        worksheet.column_dimensions['D'].width = 50  # Açıklama

        logging.info(f" {sheet_name} oluşturuldu!")

    def create_mapping_sheet(self, writer, content_data) -> None:
        """Mapping sayfası oluştur"""
        logging.info(" Mapping sayfası oluşturuluyor...")

        mapping_data = []
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            # SDG-GRI mapping
            cursor.execute("""
                SELECT 
                    mg.sdg_indicator_code,
                    mg.gri_standard,
                    mg.gri_disclosure,
                    mg.relation_type,
                    mg.notes
                FROM map_sdg_gri mg
                ORDER BY mg.sdg_indicator_code, mg.gri_standard
            """)

            sdg_gri_mappings = cursor.fetchall()

            for mapping in sdg_gri_mappings:
                relation = mapping[3] if (mapping[3] is not None and str(mapping[3]).strip().lower() != 'none') else ''
                note = mapping[4] if (mapping[4] is not None and str(mapping[4]).strip().lower() != 'none') else ''
                mapping_data.append({
                    'SDG Indicator': mapping[0],
                    'GRI Standard': mapping[1],
                    'GRI Disclosure': mapping[2],
                    'Relation Type': relation,
                    'Notes': note,
                    'Mapping Type': 'SDG-GRI'
                })

            # TSRS-GRI mapping
            cursor.execute("""
                SELECT 
                    mt.gri_standard,
                    mt.gri_disclosure,
                    mt.tsrs_section,
                    mt.tsrs_metric,
                    mt.relation_type,
                    mt.notes
                FROM map_gri_tsrs mt
                ORDER BY mt.gri_standard, mt.gri_disclosure
            """)

            tsrs_gri_mappings = cursor.fetchall()

            for mapping in tsrs_gri_mappings:
                relation = mapping[4] if (mapping[4] is not None and str(mapping[4]).strip().lower() != 'none') else ''
                note = mapping[5] if (mapping[5] is not None and str(mapping[5]).strip().lower() != 'none') else ''
                mapping_data.append({
                    'SDG Indicator': '',
                    'GRI Standard': mapping[0],
                    'GRI Disclosure': mapping[1],
                    'Relation Type': relation,
                    'Notes': note,
                    'Mapping Type': 'GRI-TSRS',
                    'TSRS Section': mapping[2],
                    'TSRS Metric': mapping[3]
                })

        except Exception as e:
            logging.error(f" Mapping verisi alma hatası: {e}")
        finally:
            conn.close()

        df = pd.DataFrame(mapping_data)
        df.to_excel(writer, sheet_name='Mappings', index=False)

        # Formatla
        worksheet = writer.sheets['Mappings']

        # Başlık satırı
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Kolon genişlikleri
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logging.error(f"Silent error caught: {str(e)}")
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        logging.info(" Mapping sayfası oluşturuldu!")


def export_gri_content_index_enhanced(output_path: str, company_id: int = 1) -> bool:
    """GRI Content Index Enhanced export fonksiyonu"""
    content_index = GRIContentIndexEnhanced()
    return content_index.export_to_excel(output_path, company_id)


if __name__ == "__main__":
    # Test export
    output_file = "gri/gri_content_index_enhanced_test.xlsx"
    logging.info(" GRI Content Index Enhanced test başlıyor...")
    if export_gri_content_index_enhanced(output_file):
        logging.info(f" GRI Content Index Enhanced başarıyla oluşturuldu: {output_file}")
    else:
        logging.info(" GRI Content Index Enhanced oluşturulamadı!")


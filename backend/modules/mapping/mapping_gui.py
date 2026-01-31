"""
Eşleştirme (Mapping) GUI
GRI-SDG-TSRS-UNGC-ISSB standartları arası eşleştirme arayüzü
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from utils.language_manager import LanguageManager
from .mapping_manager import MappingManager
from config.icons import Icons


class MappingGUI:
    """Eşleştirme GUI sınıfı"""

    def __init__(self, parent, company_id: int = 1):
        self.parent = parent
        self.company_id = company_id
        self.lm = LanguageManager()
        self.db_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "data", "sdg_desktop.sqlite"
        )

        # Manager oluştur
        self.manager = MappingManager(self.db_path)

        # GUI oluştur
        self.create_gui()

        # Verileri yükle
        self.load_mappings()

    def create_gui(self) -> None:
        """Ana GUI'yi oluştur"""
        # Ana container
        main_frame = tk.Frame(self.parent, bg='white')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Başlık
        title_frame = tk.Frame(main_frame, bg='#1e40af', height=60)
        title_frame.pack(fill='x', pady=(0, 20))
        title_frame.pack_propagate(False)

        title_label = tk.Label(
            title_frame,
            text=f"{Icons.LINK} Standart Eşleştirme (Mapping)",
            font=('Segoe UI', 16, 'bold'),
            fg='white',
            bg='#1e40af'
        )
        title_label.pack(expand=True)

        # İstatistikler
        self.create_statistics_panel(main_frame)

        # Notebook (sekmeler)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(10, 0))

        # Sekmeler
        self.create_view_mappings_tab()
        self.create_add_mapping_tab()
        self.create_import_export_tab()
        self.create_statistics_tab()

    def create_statistics_panel(self, parent) -> None:
        """İstatistik paneli"""
        stats_frame = tk.Frame(parent, bg='white')
        stats_frame.pack(fill='x', pady=(0, 10))

        # 4 KPI kartı
        self.kpi_total = self.create_kpi_card(
            stats_frame, "Toplam Eşleştirme", "0", "#1e40af"
        )
        self.kpi_total.pack(side='left', padx=5, fill='x', expand=True)

        self.kpi_verified = self.create_kpi_card(
            stats_frame, "Doğrulanmış", "0", "#10b981"
        )
        self.kpi_verified.pack(side='left', padx=5, fill='x', expand=True)

        self.kpi_gri_sdg = self.create_kpi_card(
            stats_frame, "GRI → SDG", "0", "#f59e0b"
        )
        self.kpi_gri_sdg.pack(side='left', padx=5, fill='x', expand=True)

        self.kpi_strong = self.create_kpi_card(
            stats_frame, "Güçlü Bağlantı", "0", "#8b5cf6"
        )
        self.kpi_strong.pack(side='left', padx=5, fill='x', expand=True)

        # İstatistikleri güncelle
        self.update_statistics()

    def create_kpi_card(self, parent, title: str, value: str, color: str) -> tk.Frame:
        """KPI kartı oluştur"""
        card = tk.Frame(parent, bg=color, relief='raised', bd=2)

        title_label = tk.Label(
            card, text=title, font=('Segoe UI', 10),
            bg=color, fg='white'
        )
        title_label.pack(pady=(10, 5))

        value_label = tk.Label(
            card, text=value, font=('Segoe UI', 20, 'bold'),
            bg=color, fg='white'
        )
        value_label.pack(pady=(0, 10))

        # Value label'ı card'a attribute olarak ekle
        card.value_label = value_label

        return card

    def update_statistics(self) -> None:
        """İstatistikleri güncelle"""
        stats = self.manager.get_mapping_statistics()

        self.kpi_total.value_label.config(text=str(stats.get('total_mappings', 0)))
        self.kpi_verified.value_label.config(text=str(stats.get('verified_mappings', 0)))

        by_standard = stats.get('by_standard', {})
        self.kpi_gri_sdg.value_label.config(text=str(by_standard.get('GRI-SDG', 0)))

        by_strength = stats.get('by_strength', {})
        self.kpi_strong.value_label.config(text=str(by_strength.get('strong', 0)))

    def create_view_mappings_tab(self) -> None:
        """Eşleştirmeleri Görüntüle sekmesi"""
        view_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(view_frame, text=f"{Icons.CLIPBOARD} Tüm Eşleştirmeler")

        # Filtre çerçevesi
        filter_frame = tk.LabelFrame(
            view_frame, text=f"{Icons.SEARCH} Filtreler",
            font=('Segoe UI', 10, 'bold'), bg='white'
        )
        filter_frame.pack(fill='x', padx=20, pady=10)

        # Filtre alanları
        filter_controls = tk.Frame(filter_frame, bg='white')
        filter_controls.pack(fill='x', padx=10, pady=10)

        tk.Label(filter_controls, text="Kaynak:", bg='white').pack(side='left', padx=(0, 5))
        self.filter_source_var = tk.StringVar(value="Tümü")
        source_combo = ttk.Combobox(
            filter_controls, textvariable=self.filter_source_var,
            values=["Tümü", "GRI", "SDG", "TSRS", "UNGC", "ISSB", "ESRS"],
            width=12
        )
        source_combo.pack(side='left', padx=(0, 20))

        tk.Label(filter_controls, text="Hedef:", bg='white').pack(side='left', padx=(0, 5))
        self.filter_target_var = tk.StringVar(value="Tümü")
        target_combo = ttk.Combobox(
            filter_controls, textvariable=self.filter_target_var,
            values=["Tümü", "GRI", "SDG", "TSRS", "UNGC", "ISSB", "ESRS"],
            width=12
        )
        target_combo.pack(side='left', padx=(0, 20))

        tk.Label(filter_controls, text="Güç:", bg='white').pack(side='left', padx=(0, 5))
        self.filter_strength_var = tk.StringVar(value="Tümü")
        strength_combo = ttk.Combobox(
            filter_controls, textvariable=self.filter_strength_var,
            values=["Tümü", "strong", "medium", "weak"],
            width=12
        )
        strength_combo.pack(side='left', padx=(0, 20))

        self.filter_verified_var = tk.BooleanVar()
        verified_check = tk.Checkbutton(
            filter_controls, text="Sadece Doğrulanmış",
            variable=self.filter_verified_var, bg='white'
        )
        verified_check.pack(side='left', padx=(0, 20))

        # Filtre uygula butonu
        filter_btn = tk.Button(
            filter_controls, text=f"{Icons.SEARCH} Filtrele",
            command=self.load_mappings,
            bg='#1e40af', fg='white', font=('Segoe UI', 10, 'bold'),
            relief='flat', padx=15, pady=5
        )
        filter_btn.pack(side='left')

        # Treeview çerçevesi
        tree_frame = tk.Frame(view_frame, bg='white')
        tree_frame.pack(fill='both', expand=True, padx=20, pady=(0, 10))

        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side='right', fill='y')

        # Treeview
        columns = (
            'source_standard', 'source_code', 'source_title',
            'target_standard', 'target_code', 'target_title',
            'strength', 'verified'
        )

        self.mappings_tree = ttk.Treeview(
            tree_frame, columns=columns, show='headings',
            yscrollcommand=scrollbar.set
        )

        # Başlıklar
        self.mappings_tree.heading('source_standard', text='Kaynak')
        self.mappings_tree.heading('source_code', text='Kod')
        self.mappings_tree.heading('source_title', text='Başlık')
        self.mappings_tree.heading('target_standard', text='Hedef')
        self.mappings_tree.heading('target_code', text='Kod')
        self.mappings_tree.heading('target_title', text='Başlık')
        self.mappings_tree.heading('strength', text='Güç')
        self.mappings_tree.heading('verified', text='Doğrulandı')

        # Sütun genişlikleri
        self.mappings_tree.column('source_standard', width=80)
        self.mappings_tree.column('source_code', width=100)
        self.mappings_tree.column('source_title', width=200)
        self.mappings_tree.column('target_standard', width=80)
        self.mappings_tree.column('target_code', width=100)
        self.mappings_tree.column('target_title', width=200)
        self.mappings_tree.column('strength', width=80)
        self.mappings_tree.column('verified', width=80)

        self.mappings_tree.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.mappings_tree.yview)

        # Butonlar
        button_frame = tk.Frame(view_frame, bg='white')
        button_frame.pack(fill='x', padx=20, pady=(0, 10))

        delete_btn = tk.Button(
            button_frame, text=f"{Icons.DELETE} Seçili Olanı Sil",
            command=self.delete_selected_mapping,
            bg='#ef4444', fg='white', font=('Segoe UI', 10, 'bold'),
            relief='flat', padx=15, pady=8
        )
        delete_btn.pack(side='left', padx=5)

        edit_btn = tk.Button(
            button_frame, text=f"{Icons.EDIT} Düzenle",
            command=self.edit_selected_mapping,
            bg='#f59e0b', fg='white', font=('Segoe UI', 10, 'bold'),
            relief='flat', padx=15, pady=8
        )
        edit_btn.pack(side='left', padx=5)

        refresh_btn = tk.Button(
            button_frame, text=f"{Icons.LOADING} Yenile",
            command=self.load_mappings,
            bg='#10b981', fg='white', font=('Segoe UI', 10, 'bold'),
            relief='flat', padx=15, pady=8
        )
        refresh_btn.pack(side='left', padx=5)

    def create_add_mapping_tab(self) -> None:
        """Yeni Eşleştirme Ekle sekmesi"""
        add_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(add_frame, text=f"{Icons.ADD} Yeni Eşleştirme")

        # Form çerçevesi
        form_frame = tk.LabelFrame(
            add_frame, text=f"{Icons.MEMO} Eşleştirme Bilgileri",
            font=('Segoe UI', 12, 'bold'), bg='white'
        )
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Kaynak standart
        source_group = tk.LabelFrame(
            form_frame, text="Kaynak Standart",
            font=('Segoe UI', 10, 'bold'), bg='white'
        )
        source_group.pack(fill='x', padx=10, pady=10)

        row1 = tk.Frame(source_group, bg='white')
        row1.pack(fill='x', padx=10, pady=5)

        tk.Label(row1, text="Standart:", width=12, anchor='w', bg='white').pack(side='left')
        self.add_source_standard_var = tk.StringVar()
        source_std_combo = ttk.Combobox(
            row1, textvariable=self.add_source_standard_var,
            values=["GRI", "SDG", "TSRS", "UNGC", "ISSB", "ESRS", "TCFD", "SASB"],
            width=15
        )
        source_std_combo.pack(side='left', padx=(0, 20))

        tk.Label(row1, text="Kod:", width=8, anchor='w', bg='white').pack(side='left')
        self.add_source_code_entry = tk.Entry(row1, width=20)
        self.add_source_code_entry.pack(side='left')

        row2 = tk.Frame(source_group, bg='white')
        row2.pack(fill='x', padx=10, pady=5)

        tk.Label(row2, text="Başlık:", width=12, anchor='w', bg='white').pack(side='left')
        self.add_source_title_entry = tk.Entry(row2, width=50)
        self.add_source_title_entry.pack(side='left', fill='x', expand=True)

        row3 = tk.Frame(source_group, bg='white')
        row3.pack(fill='x', padx=10, pady=5)

        tk.Label(row3, text="Açıklama:", width=12, anchor='w', bg='white').pack(side='left', anchor='n', pady=5)
        self.add_source_desc_text = tk.Text(row3, height=3, width=50)
        self.add_source_desc_text.pack(side='left', fill='both', expand=True)

        # Hedef standart
        target_group = tk.LabelFrame(
            form_frame, text="Hedef Standart",
            font=('Segoe UI', 10, 'bold'), bg='white'
        )
        target_group.pack(fill='x', padx=10, pady=10)

        row4 = tk.Frame(target_group, bg='white')
        row4.pack(fill='x', padx=10, pady=5)

        tk.Label(row4, text="Standart:", width=12, anchor='w', bg='white').pack(side='left')
        self.add_target_standard_var = tk.StringVar()
        target_std_combo = ttk.Combobox(
            row4, textvariable=self.add_target_standard_var,
            values=["GRI", "SDG", "TSRS", "UNGC", "ISSB", "ESRS", "TCFD", "SASB"],
            width=15
        )
        target_std_combo.pack(side='left', padx=(0, 20))

        tk.Label(row4, text="Kod:", width=8, anchor='w', bg='white').pack(side='left')
        self.add_target_code_entry = tk.Entry(row4, width=20)
        self.add_target_code_entry.pack(side='left')

        row5 = tk.Frame(target_group, bg='white')
        row5.pack(fill='x', padx=10, pady=5)

        tk.Label(row5, text="Başlık:", width=12, anchor='w', bg='white').pack(side='left')
        self.add_target_title_entry = tk.Entry(row5, width=50)
        self.add_target_title_entry.pack(side='left', fill='x', expand=True)

        row6 = tk.Frame(target_group, bg='white')
        row6.pack(fill='x', padx=10, pady=5)

        tk.Label(row6, text="Açıklama:", width=12, anchor='w', bg='white').pack(side='left', anchor='n', pady=5)
        self.add_target_desc_text = tk.Text(row6, height=3, width=50)
        self.add_target_desc_text.pack(side='left', fill='both', expand=True)

        # Eşleştirme özellikleri
        mapping_group = tk.LabelFrame(
            form_frame, text="Eşleştirme Özellikleri",
            font=('Segoe UI', 10, 'bold'), bg='white'
        )
        mapping_group.pack(fill='x', padx=10, pady=10)

        row7 = tk.Frame(mapping_group, bg='white')
        row7.pack(fill='x', padx=10, pady=5)

        tk.Label(row7, text="Tip:", width=12, anchor='w', bg='white').pack(side='left')
        self.add_type_var = tk.StringVar(value="direct")
        type_combo = ttk.Combobox(
            row7, textvariable=self.add_type_var,
            values=["direct", "indirect", "partial"],
            width=15
        )
        type_combo.pack(side='left', padx=(0, 20))

        tk.Label(row7, text="Güç:", width=8, anchor='w', bg='white').pack(side='left')
        self.add_strength_var = tk.StringVar(value="medium")
        strength_combo = ttk.Combobox(
            row7, textvariable=self.add_strength_var,
            values=["strong", "medium", "weak"],
            width=15
        )
        strength_combo.pack(side='left')

        row8 = tk.Frame(mapping_group, bg='white')
        row8.pack(fill='x', padx=10, pady=5)

        self.add_verified_var = tk.BooleanVar()
        verified_check = tk.Checkbutton(
            row8, text="Doğrulandı",
            variable=self.add_verified_var, bg='white'
        )
        verified_check.pack(side='left')

        row9 = tk.Frame(mapping_group, bg='white')
        row9.pack(fill='x', padx=10, pady=5)

        tk.Label(row9, text="Notlar:", width=12, anchor='w', bg='white').pack(side='left', anchor='n', pady=5)
        self.add_notes_text = tk.Text(row9, height=3, width=50)
        self.add_notes_text.pack(side='left', fill='both', expand=True)

        # Butonlar
        button_frame = tk.Frame(form_frame, bg='white')
        button_frame.pack(fill='x', padx=10, pady=10)

        save_btn = tk.Button(
            button_frame, text=f"{Icons.SAVE} Kaydet",
            command=self.save_new_mapping,
            bg='#10b981', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        save_btn.pack(side='left', padx=5)

        clear_btn = tk.Button(
            button_frame, text=f"{Icons.LOADING} Temizle",
            command=self.clear_add_form,
            bg='#6b7280', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        clear_btn.pack(side='left', padx=5)

    def create_import_export_tab(self) -> None:
        """İçe/Dışa Aktar sekmesi"""
        import_export_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(import_export_frame, text="📁 İçe/Dışa Aktar")

        # Export bölümü
        export_group = tk.LabelFrame(
            import_export_frame, text=f"{Icons.OUTBOX} Excel'e Aktar",
            font=('Segoe UI', 12, 'bold'), bg='white'
        )
        export_group.pack(fill='x', padx=20, pady=20)

        export_info = tk.Label(
            export_group,
            text="Eşleştirmeleri Excel dosyasına aktarabilirsiniz.\nFiltre seçenekleriyle belirli eşleştirmeleri dışa aktarabilirsiniz.",
            bg='white', justify='left'
        )
        export_info.pack(padx=20, pady=10, anchor='w')

        export_btn = tk.Button(
            export_group, text="📥 Excel'e Aktar",
            command=self.export_to_excel,
            bg='#10b981', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        export_btn.pack(padx=20, pady=10, anchor='w')

        # Import bölümü
        import_group = tk.LabelFrame(
            import_export_frame, text=f"{Icons.OUTBOX} Excel'den İçe Aktar",
            font=('Segoe UI', 12, 'bold'), bg='white'
        )
        import_group.pack(fill='x', padx=20, pady=20)

        import_info = tk.Label(
            import_group,
            text="Excel dosyasından eşleştirme içe aktarabilirsiniz.\nDosya formatı: Kaynak Standart, Kaynak Kod, Kaynak Başlık, Hedef Standart, Hedef Kod, Hedef Başlık, Tip, Güç, Doğrulandı, Notlar",
            bg='white', justify='left'
        )
        import_info.pack(padx=20, pady=10, anchor='w')

        import_btn = tk.Button(
            import_group, text=f"{Icons.OUTBOX} Excel'den İçe Aktar",
            command=self.import_from_excel,
            bg='#3b82f6', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        import_btn.pack(padx=20, pady=10, anchor='w')

        # Şablon indirme
        template_group = tk.LabelFrame(
            import_export_frame, text=f"{Icons.CLIPBOARD} Şablon",
            font=('Segoe UI', 12, 'bold'), bg='white'
        )
        template_group.pack(fill='x', padx=20, pady=20)

        template_info = tk.Label(
            template_group,
            text="İçe aktarma için örnek Excel şablonu indirebilirsiniz.",
            bg='white', justify='left'
        )
        template_info.pack(padx=20, pady=10, anchor='w')

        template_btn = tk.Button(
            template_group, text=f"{Icons.FILE} Şablon İndir",
            command=self.download_template,
            bg='#f59e0b', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        template_btn.pack(padx=20, pady=10, anchor='w')

    def create_statistics_tab(self) -> None:
        """İstatistikler sekmesi"""
        stats_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(stats_frame, text=f"{Icons.REPORT} İstatistikler")

        # İstatistik paneli
        stats_display = tk.Frame(stats_frame, bg='white')
        stats_display.pack(fill='both', expand=True, padx=20, pady=20)

        # Başlık
        title = tk.Label(
            stats_display, text=f"{Icons.REPORT} Eşleştirme İstatistikleri",
            font=('Segoe UI', 16, 'bold'), bg='white'
        )
        title.pack(pady=(0, 20))

        # İstatistik kutusu
        self.stats_text = tk.Text(
            stats_display, font=('Courier', 11),
            bg='#f9fafb', relief='solid', bd=1,
            wrap='word'
        )
        self.stats_text.pack(fill='both', expand=True)

        # Yenile butonu
        refresh_stats_btn = tk.Button(
            stats_frame, text=f"{Icons.LOADING} İstatistikleri Yenile",
            command=self.update_statistics_display,
            bg='#1e40af', fg='white', font=('Segoe UI', 12, 'bold'),
            relief='flat', padx=30, pady=10
        )
        refresh_stats_btn.pack(pady=10)

        # İstatistikleri göster
        self.update_statistics_display()

    def load_mappings(self) -> None:
        """Eşleştirmeleri yükle"""
        # Treeview'i temizle
        for item in self.mappings_tree.get_children():
            self.mappings_tree.delete(item)

        # Filtreleri hazırla
        filters = {}

        source = self.filter_source_var.get()
        if source != "Tümü":
            filters['source_standard'] = source

        target = self.filter_target_var.get()
        if target != "Tümü":
            filters['target_standard'] = target

        strength = self.filter_strength_var.get()
        if strength != "Tümü":
            filters['mapping_strength'] = strength

        if self.filter_verified_var.get():
            filters['verified_only'] = True

        # Verileri getir
        mappings = self.manager.get_all_mappings(filters)

        # Treeview'e ekle
        for mapping in mappings:
            self.mappings_tree.insert('', 'end', values=(
                mapping['source_standard'],
                mapping['source_code'],
                mapping['source_title'][:50] + '...' if len(mapping['source_title'] or '') > 50 else mapping['source_title'],
                mapping['target_standard'],
                mapping['target_code'],
                mapping['target_title'][:50] + '...' if len(mapping['target_title'] or '') > 50 else mapping['target_title'],
                mapping['mapping_strength'],
                'Evet' if mapping['verified'] else 'Hayır'
            ), tags=(str(mapping['id']),))

        # İstatistikleri güncelle
        self.update_statistics()

    def save_new_mapping(self) -> None:
        """Yeni eşleştirme kaydet"""
        # Validasyon
        if not self.add_source_standard_var.get():
            messagebox.showerror("Hata", "Kaynak standart seçiniz!")
            return

        if not self.add_source_code_entry.get():
            messagebox.showerror("Hata", "Kaynak kod giriniz!")
            return

        if not self.add_target_standard_var.get():
            messagebox.showerror("Hata", "Hedef standart seçiniz!")
            return

        if not self.add_target_code_entry.get():
            messagebox.showerror("Hata", "Hedef kod giriniz!")
            return

        # Veri hazırla
        mapping_data = {
            'source_standard': self.add_source_standard_var.get(),
            'source_code': self.add_source_code_entry.get(),
            'source_title': self.add_source_title_entry.get(),
            'source_description': self.add_source_desc_text.get('1.0', tk.END).strip(),
            'target_standard': self.add_target_standard_var.get(),
            'target_code': self.add_target_code_entry.get(),
            'target_title': self.add_target_title_entry.get(),
            'target_description': self.add_target_desc_text.get('1.0', tk.END).strip(),
            'mapping_type': self.add_type_var.get(),
            'mapping_strength': self.add_strength_var.get(),
            'verified': 1 if self.add_verified_var.get() else 0,
            'notes': self.add_notes_text.get('1.0', tk.END).strip()
        }

        # Kaydet
        if self.manager.add_mapping(mapping_data):
            messagebox.showinfo("Başarılı", "Eşleştirme kaydedildi!")
            self.clear_add_form()
            self.load_mappings()
        else:
            messagebox.showerror("Hata", "Eşleştirme kaydedilemedi!")

    def clear_add_form(self) -> None:
        """Form alanlarını temizle"""
        self.add_source_standard_var.set('')
        self.add_source_code_entry.delete(0, tk.END)
        self.add_source_title_entry.delete(0, tk.END)
        self.add_source_desc_text.delete('1.0', tk.END)
        self.add_target_standard_var.set('')
        self.add_target_code_entry.delete(0, tk.END)
        self.add_target_title_entry.delete(0, tk.END)
        self.add_target_desc_text.delete('1.0', tk.END)
        self.add_type_var.set('direct')
        self.add_strength_var.set('medium')
        self.add_verified_var.set(False)
        self.add_notes_text.delete('1.0', tk.END)

    def delete_selected_mapping(self) -> None:
        """Seçili eşleştirmeyi sil"""
        selected = self.mappings_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen silinecek eşleştirmeyi seçiniz!")
            return

        if not messagebox.askyesno("Onay", "Seçili eşleştirmeyi silmek istediğinize emin misiniz?"):
            return

        # ID'yi al
        mapping_id = int(self.mappings_tree.item(selected[0])['tags'][0])

        # Sil
        if self.manager.delete_mapping(mapping_id):
            messagebox.showinfo("Başarılı", "Eşleştirme silindi!")
            self.load_mappings()
        else:
            messagebox.showerror("Hata", "Eşleştirme silinemedi!")

    def edit_selected_mapping(self) -> None:
        """Seçili eşleştirmeyi düzenle"""
        selected = self.mappings_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek eşleştirmeyi seçiniz!")
            return

        # Seçili item'ın değerlerini al
        item = self.mappings_tree.item(selected[0])
        values = item['values']
        mapping_id = values[0]

        # Mapping bilgilerini al
        mapping = self.manager.get_mapping_by_id(mapping_id)
        if not mapping:
            messagebox.showerror("Hata", "Eşleştirme bulunamadı!")
            return

        # Düzenleme dialogu oluştur
        dialog = tk.Toplevel(self.parent)
        dialog.title("Eşleştirme Düzenle")
        dialog.geometry("600x400")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Form frame
        form_frame = ttk.LabelFrame(dialog, text="Eşleştirme Bilgileri", padding=20)
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Kaynak Standart
        ttk.Label(form_frame, text="Kaynak Standart:").grid(row=0, column=0, sticky='w', pady=5)
        source_var = tk.StringVar(value=mapping[1])  # source_standard
        source_combo = ttk.Combobox(form_frame, textvariable=source_var, values=self.standards, state='readonly', width=25)
        source_combo.grid(row=0, column=1, sticky='ew', pady=5, padx=(10, 0))

        # Kaynak Kod
        ttk.Label(form_frame, text="Kaynak Kod:").grid(row=1, column=0, sticky='w', pady=5)
        source_code_var = tk.StringVar(value=mapping[2])  # source_code
        ttk.Entry(form_frame, textvariable=source_code_var, width=27).grid(row=1, column=1, sticky='ew', pady=5, padx=(10, 0))

        # Hedef Standart
        ttk.Label(form_frame, text="Hedef Standart:").grid(row=2, column=0, sticky='w', pady=5)
        target_var = tk.StringVar(value=mapping[3])  # target_standard
        target_combo = ttk.Combobox(form_frame, textvariable=target_var, values=self.standards, state='readonly', width=25)
        target_combo.grid(row=2, column=1, sticky='ew', pady=5, padx=(10, 0))

        # Hedef Kod
        ttk.Label(form_frame, text="Hedef Kod:").grid(row=3, column=0, sticky='w', pady=5)
        target_code_var = tk.StringVar(value=mapping[4])  # target_code
        ttk.Entry(form_frame, textvariable=target_code_var, width=27).grid(row=3, column=1, sticky='ew', pady=5, padx=(10, 0))

        # Eşleştirme Tipi
        ttk.Label(form_frame, text="Eşleştirme Tipi:").grid(row=4, column=0, sticky='w', pady=5)
        type_var = tk.StringVar(value=mapping[5])  # mapping_type
        type_combo = ttk.Combobox(form_frame, textvariable=type_var,
                                 values=["exact", "partial", "related", "similar"],
                                 state='readonly', width=25)
        type_combo.grid(row=4, column=1, sticky='ew', pady=5, padx=(10, 0))

        # Açıklama
        ttk.Label(form_frame, text="Açıklama:").grid(row=5, column=0, sticky='nw', pady=5)
        description_text = tk.Text(form_frame, height=4, width=27)
        description_text.insert('1.0', mapping[6] or '')  # description
        description_text.grid(row=5, column=1, sticky='ew', pady=5, padx=(10, 0))

        form_frame.columnconfigure(1, weight=1)

        def save_changes():
            """Değişiklikleri kaydet"""
            # Güncellenmiş verileri al
            updated_data = {
                'source_standard': source_var.get(),
                'source_code': source_code_var.get().strip(),
                'target_standard': target_var.get(),
                'target_code': target_code_var.get().strip(),
                'mapping_type': type_var.get(),
                'notes': description_text.get('1.0', tk.END).strip()
            }

            # Validasyon
            if not all([updated_data['source_standard'], updated_data['source_code'],
                       updated_data['target_standard'], updated_data['target_code']]):
                messagebox.showerror("Hata", "Lütfen tüm zorunlu alanları doldurunuz!")
                return

            # Güncelle
            if self.manager.update_mapping(mapping_id, updated_data):
                messagebox.showinfo("Başarılı", "Eşleştirme güncellendi!")
                dialog.destroy()
                self.load_mappings()
            else:
                messagebox.showerror("Hata", "Eşleştirme güncellenemedi!")

        # Butonlar
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill='x', padx=20, pady=(0, 20))

        ttk.Button(button_frame, text=self.lm.tr("btn_save", "Kaydet"), command=save_changes).pack(side='left', padx=5)
        ttk.Button(button_frame, text=self.lm.tr("btn_cancel", "İptal"), command=dialog.destroy).pack(side='left', padx=5)

    def export_to_excel(self) -> None:
        """Excel'e aktar"""
        filepath = filedialog.asksaveasfilename(
            title=self.lm.tr('export_excel', "Excel'e Aktar"),
            defaultextension=".xlsx",
            filetypes=[(self.lm.tr('file_excel', "Excel Dosyası"), "*.xlsx"), (self.lm.tr('all_files', "Tüm Dosyalar"), "*.*")],
            initialfile=f"eslestirmeler_{self.filter_source_var.get()}_{self.filter_target_var.get()}.xlsx"
        )

        if not filepath:
            return

        # Filtreleri hazırla
        filters = {}
        if self.filter_source_var.get() != "Tümü":
            filters['source_standard'] = self.filter_source_var.get()
        if self.filter_target_var.get() != "Tümü":
            filters['target_standard'] = self.filter_target_var.get()

        # Export
        if self.manager.export_to_excel(filepath, filters):
            messagebox.showinfo("Başarılı", f"Eşleştirmeler Excel'e aktarıldı!\n\n{filepath}")
        else:
            messagebox.showerror("Hata", "Excel'e aktarma başarısız!")

    def import_from_excel(self) -> None:
        """Excel'den içe aktar"""
        filepath = filedialog.askopenfilename(
            title=self.lm.tr('import_excel', "Excel'den İçe Aktar"),
            filetypes=[(self.lm.tr('file_excel', "Excel Dosyası"), "*.xlsx"), (self.lm.tr('all_files', "Tüm Dosyalar"), "*.*")]
        )

        if not filepath:
            return

        # Import
        success, error = self.manager.import_from_excel(filepath)

        messagebox.showinfo(
            "Tamamlandı",
            f"İçe aktarma tamamlandı!\n\nBaşarılı: {success}\nHatalı: {error}"
        )

        # Verileri yenile
        self.load_mappings()

    def download_template(self) -> None:
        """Şablon indir"""
        filepath = filedialog.asksaveasfilename(
            title=self.lm.tr('save_template', "Şablonu Kaydet"),
            defaultextension=".xlsx",
            filetypes=[(self.lm.tr('file_excel', "Excel Dosyası"), "*.xlsx")],
            initialfile="eslestirme_sablonu.xlsx"
        )

        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Eşleştirmeler"

            # Başlıklar
            headers = [
                'Kaynak Standart', 'Kaynak Kod', 'Kaynak Başlık',
                'Hedef Standart', 'Hedef Kod', 'Hedef Başlık',
                'Eşleştirme Tipi', 'Güç', 'Doğrulandı', 'Notlar'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Örnek veri
            ws.append(['GRI', 'GRI 302-1', 'Energy consumption', 'SDG', 'SDG 7.3', 'Energy efficiency', 'direct', 'strong', 'Evet', 'Örnek eşleştirme'])

            wb.save(filepath)
            messagebox.showinfo("Başarılı", f"Şablon indirildi!\n\n{filepath}")

        except Exception as e:
            messagebox.showerror("Hata", f"Şablon oluşturma hatası:\n{str(e)}")

    def update_statistics_display(self) -> None:
        """İstatistikleri güncelle ve göster"""
        self.stats_text.delete('1.0', tk.END)

        stats = self.manager.get_mapping_statistics()

        output = []
        output.append("=" * 60)
        output.append("EŞLEŞTİRME İSTATİSTİKLERİ")
        output.append("=" * 60)
        output.append("")
        output.append(f"Toplam Eşleştirme: {stats.get('total_mappings', 0)}")
        output.append(f"Doğrulanmış: {stats.get('verified_mappings', 0)}")
        output.append("")
        output.append("-" * 60)
        output.append("STANDART BAZINDA:")
        output.append("-" * 60)

        by_standard = stats.get('by_standard', {})
        for key, count in sorted(by_standard.items(), key=lambda x: x[1], reverse=True):
            output.append(f"  {key}: {count}")

        output.append("")
        output.append("-" * 60)
        output.append("GÜÇ SEVİYESİ BAZINDA:")
        output.append("-" * 60)

        by_strength = stats.get('by_strength', {})
        for key, count in sorted(by_strength.items(), key=lambda x: x[1], reverse=True):
            output.append(f"  {key}: {count}")

        output.append("")
        output.append("=" * 60)

        self.stats_text.insert('1.0', '\n'.join(output))


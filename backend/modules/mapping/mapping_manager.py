"""
Eşleştirme Yöneticisi
GRI-SDG-TSRS-UNGC-ISSB standartları arası eşleştirme yönetimi
"""

import logging
import sqlite3
from typing import Any, Dict, List, Optional, Tuple

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False



class MappingManager:
    """Standart eşleştirme yönetimi"""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.create_tables()
        self.initialize_default_mappings()

    def create_tables(self) -> None:
        """Eşleştirme tablolarını oluştur"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Ana eşleştirme tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS standard_mappings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_standard TEXT NOT NULL,
                    source_code TEXT NOT NULL,
                    source_title TEXT,
                    source_description TEXT,
                    target_standard TEXT NOT NULL,
                    target_code TEXT NOT NULL,
                    target_title TEXT,
                    target_description TEXT,
                    mapping_type TEXT DEFAULT 'direct',
                    mapping_strength TEXT DEFAULT 'strong',
                    notes TEXT,
                    verified BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Özel eşleştirmeler (kullanıcı tarafından eklenen)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS custom_mappings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company_id INTEGER,
                    source_standard TEXT NOT NULL,
                    source_code TEXT NOT NULL,
                    target_standard TEXT NOT NULL,
                    target_code TEXT NOT NULL,
                    mapping_rationale TEXT,
                    created_by TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (company_id) REFERENCES companies (id)
                )
            ''')

            # Eşleştirme önerileri
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS mapping_suggestions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_standard TEXT NOT NULL,
                    source_code TEXT NOT NULL,
                    target_standard TEXT NOT NULL,
                    target_code TEXT NOT NULL,
                    confidence_score REAL,
                    suggestion_reason TEXT,
                    status TEXT DEFAULT 'pending',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            conn.commit()
            conn.close()
            logging.info("[OK] Eşleştirme tabloları oluşturuldu")

        except Exception as e:
            logging.error(f"[HATA] Eşleştirme tabloları oluşturma hatası: {e}")

    def initialize_default_mappings(self) -> None:
        """Varsayılan eşleştirmeleri yükle"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Mevcut eşleştirme sayısını kontrol et
            cursor.execute("SELECT COUNT(*) FROM standard_mappings")
            count = cursor.fetchone()[0]

            if count > 0:
                conn.close()
                return  # Zaten yüklenmiş

            # Örnek GRI-SDG eşleştirmeleri
            default_mappings = [
                # GRI 302 (Enerji) -> SDG 7 (Erişilebilir ve Temiz Enerji)
                ('GRI', 'GRI 302-1', 'Energy consumption within the organization',
                 'Total fuel consumption, electricity, heating, cooling, and steam',
                 'SDG', 'SDG 7.3', 'Affordable and Clean Energy',
                 'Double the global rate of improvement in energy efficiency',
                 'direct', 'strong', 'Energy efficiency metrics'),

                # GRI 305 (Emisyonlar) -> SDG 13 (İklim Eylemi)
                ('GRI', 'GRI 305-1', 'Direct GHG emissions (Scope 1)',
                 'Gross direct greenhouse gas emissions in metric tons of CO2 equivalent',
                 'SDG', 'SDG 13.2', 'Climate Action',
                 'Integrate climate change measures into national policies',
                 'direct', 'strong', 'GHG emissions tracking'),

                # GRI 303 (Su) -> SDG 6 (Temiz Su ve Sanitasyon)
                ('GRI', 'GRI 303-3', 'Water withdrawal',
                 'Total water withdrawal from all sources',
                 'SDG', 'SDG 6.4', 'Clean Water and Sanitation',
                 'Substantially increase water-use efficiency',
                 'direct', 'strong', 'Water management'),

                # GRI 401 (İstihdam) -> SDG 8 (İnsana Yakışır İş)
                ('GRI', 'GRI 401-1', 'New employee hires and employee turnover',
                 'Total number and rate of new employee hires and employee turnover',
                 'SDG', 'SDG 8.5', 'Decent Work and Economic Growth',
                 'Full and productive employment and decent work for all',
                 'direct', 'strong', 'Employment metrics'),

                # GRI 403 (İSG) -> SDG 8 (İnsana Yakışır İş)
                ('GRI', 'GRI 403-9', 'Work-related injuries',
                 'Number and rate of fatalities, high-consequence and recordable work-related injuries',
                 'SDG', 'SDG 8.8', 'Decent Work and Economic Growth',
                 'Protect labour rights and promote safe working environments',
                 'direct', 'strong', 'Occupational health and safety'),

                # GRI 404 (Eğitim) -> SDG 4 (Nitelikli Eğitim)
                ('GRI', 'GRI 404-1', 'Average hours of training per year per employee',
                 'Average hours of training that employees have undertaken',
                 'SDG', 'SDG 4.4', 'Quality Education',
                 'Increase the number of people with relevant skills for employment',
                 'direct', 'strong', 'Training and development'),

                # GRI 405 (Çeşitlilik) -> SDG 5 (Toplumsal Cinsiyet Eşitliği)
                ('GRI', 'GRI 405-1', 'Diversity of governance bodies and employees',
                 'Percentage of individuals within governance bodies and employees per category',
                 'SDG', 'SDG 5.5', 'Gender Equality',
                 'Ensure women\'s full participation in leadership',
                 'direct', 'strong', 'Gender diversity'),

                # GRI 306 (Atık) -> SDG 12 (Sorumlu Üretim ve Tüketim)
                ('GRI', 'GRI 306-3', 'Waste generated',
                 'Total weight of waste generated and waste diverted from disposal',
                 'SDG', 'SDG 12.5', 'Responsible Consumption and Production',
                 'Substantially reduce waste generation',
                 'direct', 'strong', 'Waste management'),

                # GRI 413 (Yerel Topluluklar) -> SDG 11 (Sürdürülebilir Şehirler)
                ('GRI', 'GRI 413-1', 'Operations with local community engagement',
                 'Percentage of operations with implemented local community engagement',
                 'SDG', 'SDG 11.3', 'Sustainable Cities and Communities',
                 'Enhance inclusive and sustainable urbanization',
                 'direct', 'medium', 'Community engagement'),

                # GRI 201 (Ekonomik Performans) -> SDG 8 (Ekonomik Büyüme)
                ('GRI', 'GRI 201-1', 'Direct economic value generated and distributed',
                 'Direct economic value generated and distributed including revenues',
                 'SDG', 'SDG 8.1', 'Decent Work and Economic Growth',
                 'Sustain per capita economic growth',
                 'direct', 'medium', 'Economic contribution'),
            ]

            # TSRS → ESRS başlangıç eşleştirmeleri
            tsrs_esrs = [
                ('TSRS', 'TSRS E1', 'İklim Değişikliği',
                 'Enerji, emisyon ve iklimle ilgili TSRS göstergeleri',
                 'ESRS', 'ESRS E1', 'İklim Değişikliği',
                 'ESRS E1 kapsamındaki iklim açıklamaları',
                 'direct', 'strong', 'TSRS çevre standartları ↔ ESRS E1'),
                ('TSRS', 'TSRS E2', 'Kirlilik',
                 'Kirletici emisyonlar ve yönetimi',
                 'ESRS', 'ESRS E2', 'Kirlilik',
                 'ESRS E2 kapsamındaki kirlilik açıklamaları',
                 'direct', 'strong', 'TSRS çevre standartları ↔ ESRS E2'),
                ('TSRS', 'TSRS E3', 'Su ve Deniz Kaynakları',
                 'Su tüketimi ve kaynak yönetimi',
                 'ESRS', 'ESRS E3', 'Su ve Deniz Kaynakları',
                 'ESRS E3 kapsamındaki su ve deniz açıklamaları',
                 'direct', 'strong', 'TSRS çevre standartları ↔ ESRS E3')
            ]
            cursor.executemany('''
                INSERT INTO standard_mappings 
                (source_standard, source_code, source_title, source_description,
                 target_standard, target_code, target_title, target_description,
                 mapping_type, mapping_strength, notes, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', tsrs_esrs)

            cursor.executemany('''
                INSERT INTO standard_mappings 
                (source_standard, source_code, source_title, source_description,
                 target_standard, target_code, target_title, target_description,
                 mapping_type, mapping_strength, notes, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', default_mappings)

            conn.commit()
            conn.close()
            logging.info(f"[OK] {len(default_mappings)} varsayılan eşleştirme yüklendi")

        except Exception as e:
            logging.error(f"[HATA] Varsayılan eşleştirme yükleme hatası: {e}")

    def get_all_mappings(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Tüm eşleştirmeleri getir"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            query = "SELECT * FROM standard_mappings WHERE 1=1"
            params = []

            if filters:
                if 'source_standard' in filters:
                    query += " AND source_standard = ?"
                    params.append(filters['source_standard'])

                if 'target_standard' in filters:
                    query += " AND target_standard = ?"
                    params.append(filters['target_standard'])

                if 'mapping_strength' in filters:
                    query += " AND mapping_strength = ?"
                    params.append(filters['mapping_strength'])

                if 'verified_only' in filters and filters['verified_only']:
                    query += " AND verified = 1"

            query += " ORDER BY source_code, target_code"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()

            mappings = []
            for row in rows:
                mappings.append({
                    'id': row[0],
                    'source_standard': row[1],
                    'source_code': row[2],
                    'source_title': row[3],
                    'source_description': row[4],
                    'target_standard': row[5],
                    'target_code': row[6],
                    'target_title': row[7],
                    'target_description': row[8],
                    'mapping_type': row[9],
                    'mapping_strength': row[10],
                    'notes': row[11],
                    'verified': row[12]
                })

            return mappings

        except Exception as e:
            logging.error(f"[HATA] Eşleştirme getirme hatası: {e}")
            return []

    def add_mapping(self, mapping_data: Dict[str, Any]) -> bool:
        """Yeni eşleştirme ekle"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                INSERT INTO standard_mappings 
                (source_standard, source_code, source_title, source_description,
                 target_standard, target_code, target_title, target_description,
                 mapping_type, mapping_strength, notes, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                mapping_data.get('source_standard'),
                mapping_data.get('source_code'),
                mapping_data.get('source_title', ''),
                mapping_data.get('source_description', ''),
                mapping_data.get('target_standard'),
                mapping_data.get('target_code'),
                mapping_data.get('target_title', ''),
                mapping_data.get('target_description', ''),
                mapping_data.get('mapping_type', 'direct'),
                mapping_data.get('mapping_strength', 'medium'),
                mapping_data.get('notes', ''),
                mapping_data.get('verified', 0)
            ))

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            logging.error(f"[HATA] Eşleştirme ekleme hatası: {e}")
            return False

    def get_mapping_by_id(self, mapping_id: int) -> Optional[Tuple]:
        """ID'ye göre eşleştirme getir"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, source_standard, source_code, target_standard, 
                       target_code, mapping_type, notes
                FROM standard_mappings 
                WHERE id = ?
            """, (mapping_id,))
            result = cursor.fetchone()
            conn.close()
            return result
        except Exception as e:
            logging.error(f"[HATA] Eşleştirme getirme hatası: {e}")
            return None

    def delete_mapping(self, mapping_id: int) -> bool:
        """Eşleştirme sil"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM standard_mappings WHERE id = ?", (mapping_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logging.error(f"[HATA] Eşleştirme silme hatası: {e}")
            return False

    def update_mapping(self, mapping_id: int, mapping_data: Dict[str, Any]) -> bool:
        """Eşleştirme güncelle"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                UPDATE standard_mappings SET
                    source_standard = ?,
                    source_code = ?,
                    source_title = ?,
                    source_description = ?,
                    target_standard = ?,
                    target_code = ?,
                    target_title = ?,
                    target_description = ?,
                    mapping_type = ?,
                    mapping_strength = ?,
                    notes = ?,
                    verified = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (
                mapping_data.get('source_standard'),
                mapping_data.get('source_code'),
                mapping_data.get('source_title', ''),
                mapping_data.get('source_description', ''),
                mapping_data.get('target_standard'),
                mapping_data.get('target_code'),
                mapping_data.get('target_title', ''),
                mapping_data.get('target_description', ''),
                mapping_data.get('mapping_type', 'direct'),
                mapping_data.get('mapping_strength', 'medium'),
                mapping_data.get('notes', ''),
                mapping_data.get('verified', 0),
                mapping_id
            ))

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            logging.error(f"[HATA] Eşleştirme güncelleme hatası: {e}")
            return False

    def get_mapping_statistics(self) -> Dict[str, Any]:
        """Eşleştirme istatistikleri"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Toplam eşleştirme
            cursor.execute("SELECT COUNT(*) FROM standard_mappings")
            total = cursor.fetchone()[0]

            # Doğrulanmış eşleştirmeler
            cursor.execute("SELECT COUNT(*) FROM standard_mappings WHERE verified = 1")
            verified = cursor.fetchone()[0]

            # Standart bazlı
            cursor.execute("""
                SELECT source_standard, target_standard, COUNT(*) 
                FROM standard_mappings 
                GROUP BY source_standard, target_standard
            """)
            by_standard = cursor.fetchall()

            # Güç bazlı
            cursor.execute("""
                SELECT mapping_strength, COUNT(*) 
                FROM standard_mappings 
                GROUP BY mapping_strength
            """)
            by_strength = cursor.fetchall()

            conn.close()

            return {
                'total_mappings': total,
                'verified_mappings': verified,
                'by_standard': {f"{row[0]}-{row[1]}": row[2] for row in by_standard},
                'by_strength': {row[0]: row[1] for row in by_strength}
            }

        except Exception as e:
            logging.error(f"[HATA] İstatistik hesaplama hatası: {e}")
            return {}

    def export_to_excel(self, filepath: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Eşleştirmeleri Excel'e aktar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            mappings = self.get_all_mappings(filters)

            wb = Workbook()
            ws = wb.active
            ws.title = "Eşleştirmeler"

            # Başlıklar
            headers = [
                'Kaynak Standart', 'Kaynak Kod', 'Kaynak Başlık',
                'Hedef Standart', 'Hedef Kod', 'Hedef Başlık',
                'Eşleştirme Tipi', 'Güç', 'Doğrulandı', 'Notlar'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Veri
            for row_idx, mapping in enumerate(mappings, 2):
                ws.cell(row_idx, 1, mapping['source_standard'])
                ws.cell(row_idx, 2, mapping['source_code'])
                ws.cell(row_idx, 3, mapping['source_title'])
                ws.cell(row_idx, 4, mapping['target_standard'])
                ws.cell(row_idx, 5, mapping['target_code'])
                ws.cell(row_idx, 6, mapping['target_title'])
                ws.cell(row_idx, 7, mapping['mapping_type'])
                ws.cell(row_idx, 8, mapping['mapping_strength'])
                ws.cell(row_idx, 9, 'Evet' if mapping['verified'] else 'Hayır')
                ws.cell(row_idx, 10, mapping['notes'] or '')

            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 30

            wb.save(filepath)
            return True

        except Exception as e:
            logging.error(f"[HATA] Excel export hatası: {e}")
            return False

    def import_from_excel(self, filepath: str) -> Tuple[int, int]:
        """Excel'den eşleştirme içe aktar"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filepath)
            ws = wb.active

            success_count = 0
            error_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[3] or not row[4]:
                    error_count += 1
                    continue

                mapping_data = {
                    'source_standard': row[0],
                    'source_code': row[1],
                    'source_title': row[2] or '',
                    'target_standard': row[3],
                    'target_code': row[4],
                    'target_title': row[5] or '',
                    'mapping_type': row[6] or 'direct',
                    'mapping_strength': row[7] or 'medium',
                    'verified': 1 if row[8] == 'Evet' else 0,
                    'notes': row[9] or ''
                }

                if self.add_mapping(mapping_data):
                    success_count += 1
                else:
                    error_count += 1

            return (success_count, error_count)

        except Exception as e:
            logging.error(f"[HATA] Excel import hatası: {e}")
            return (0, 0)

    def generate_suggestions(self) -> int:
        """Otomatik eşleştirme önerileri oluştur (TF-IDF ile)"""
        if not SKLEARN_AVAILABLE:
            logging.warning("scikit-learn bulunamadı, öneri oluşturulamıyor.")
            return 0

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 1. Mevcut tüm standart maddelerini topla
            cursor.execute("SELECT source_standard, source_code, source_description FROM standard_mappings")
            sources = cursor.fetchall()
            
            cursor.execute("SELECT target_standard, target_code, target_description FROM standard_mappings")
            targets = cursor.fetchall()
            
            # Benzersizleri al
            unique_sources = list(set(sources))
            unique_targets = list(set(targets))
            
            if not unique_sources or not unique_targets:
                conn.close()
                return 0

            # 2. Metinleri hazırla
            source_docs = [s[2] for s in unique_sources if s[2]]
            target_docs = [t[2] for t in unique_targets if t[2]]
            
            if not source_docs or not target_docs:
                conn.close()
                return 0

            # 3. TF-IDF ve Benzerlik
            vectorizer = TfidfVectorizer(stop_words='english')
            # Tüm corpus
            all_docs = source_docs + target_docs
            vectorizer.fit(all_docs)
            
            tfidf_source = vectorizer.transform(source_docs)
            tfidf_target = vectorizer.transform(target_docs)
            
            similarity_matrix = cosine_similarity(tfidf_source, tfidf_target)
            
            suggestions_count = 0
            
            # 4. Önerileri bul ve kaydet
            for i, src in enumerate(unique_sources):
                for j, tgt in enumerate(unique_targets):
                    score = float(similarity_matrix[i][j])
                    
                    # Eşik değer ve aynı standart kontrolü
                    if score > 0.3 and src[0] != tgt[0]:
                        # Zaten var mı kontrol et
                        cursor.execute("""
                            SELECT 1 FROM standard_mappings 
                            WHERE source_code = ? AND target_code = ?
                        """, (src[1], tgt[1]))
                        exists = cursor.fetchone()
                        
                        if not exists:
                            # Zaten önerilmiş mi?
                            cursor.execute("""
                                SELECT 1 FROM mapping_suggestions 
                                WHERE source_code = ? AND target_code = ?
                            """, (src[1], tgt[1]))
                            suggested = cursor.fetchone()
                            
                            if not suggested:
                                cursor.execute("""
                                    INSERT INTO mapping_suggestions 
                                    (source_standard, source_code, target_standard, target_code, 
                                     confidence_score, suggestion_reason, status)
                                    VALUES (?, ?, ?, ?, ?, ?, 'pending')
                                """, (src[0], src[1], tgt[0], tgt[1], score, f"Similarity: {score:.2f}"))
                                suggestions_count += 1
            
            conn.commit()
            conn.close()
            return suggestions_count

        except Exception as e:
            logging.error(f"[HATA] Öneri oluşturma hatası: {e}")
            return 0

    def get_suggestions(self, status: str = 'pending') -> List[Dict[str, Any]]:
        """Önerileri listele"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, source_standard, source_code, target_standard, target_code, 
                       confidence_score, suggestion_reason, status, created_at
                FROM mapping_suggestions
                WHERE status = ?
                ORDER BY confidence_score DESC
            """, (status,))
            
            rows = cursor.fetchall()
            conn.close()
            
            suggestions = []
            for row in rows:
                suggestions.append({
                    'id': row[0],
                    'source_standard': row[1],
                    'source_code': row[2],
                    'target_standard': row[3],
                    'target_code': row[4],
                    'confidence_score': row[5],
                    'suggestion_reason': row[6],
                    'status': row[7],
                    'created_at': row[8]
                })
            return suggestions
            
        except Exception as e:
            logging.error(f"[HATA] Öneri listeleme hatası: {e}")
            return []

    def approve_suggestion(self, suggestion_id: int) -> bool:
        """Öneriyi onayla ve standard_mappings'e taşı"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Öneriyi al
            cursor.execute("SELECT * FROM mapping_suggestions WHERE id = ?", (suggestion_id,))
            row = cursor.fetchone()
            
            if not row:
                conn.close()
                return False
                
            # Mapping verisi (Indices based on SELECT * or schema)
            # Schema: id, source_std, source_code, target_std, target_code, score, reason, status, created
            # Indices: 0, 1, 2, 3, 4, 5, 6, 7, 8
            
            # Standard mapping ekle
            # Description'ları bulmak için standard_mappings'den tekrar sorgulamak gerekebilir 
            # veya boş bırakabiliriz. Şimdilik boş bırakalım veya mevcutlardan bulmaya çalışalım.
            
            # Kaynak tanımını bul
            cursor.execute("SELECT source_title, source_description FROM standard_mappings WHERE source_code = ? LIMIT 1", (row[2],))
            src_info = cursor.fetchone() or ('', '')
            
            # Hedef tanımını bul
            cursor.execute("SELECT target_title, target_description FROM standard_mappings WHERE target_code = ? LIMIT 1", (row[4],))
            tgt_info = cursor.fetchone() or ('', '')

            cursor.execute('''
                INSERT INTO standard_mappings 
                (source_standard, source_code, source_title, source_description,
                 target_standard, target_code, target_title, target_description,
                 mapping_type, mapping_strength, notes, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'direct', 'medium', ?, 1)
            ''', (
                row[1], row[2], src_info[0], src_info[1],
                row[3], row[4], tgt_info[0], tgt_info[1],
                f"Auto-generated from suggestion (Score: {row[5]:.2f})"
            ))
            
            # Öneri durumunu güncelle
            cursor.execute("UPDATE mapping_suggestions SET status = 'approved' WHERE id = ?", (suggestion_id,))
            
            conn.commit()
            conn.close()
            return True
            
        except Exception as e:
            logging.error(f"[HATA] Öneri onaylama hatası: {e}")
            return False

    def reject_suggestion(self, suggestion_id: int) -> bool:
        """Öneriyi reddet"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("UPDATE mapping_suggestions SET status = 'rejected' WHERE id = ?", (suggestion_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logging.error(f"[HATA] Öneri reddetme hatası: {e}")
            return False


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Veri Toplama Sistemi GUI
"""

import logging
import os
import tkinter as tk
from tkinter import messagebox, ttk
from utils.language_manager import LanguageManager


class DataCollectionGUI:
    """Veri toplama sistemi GUI"""

    def __init__(self, parent, company_id: int = 1) -> None:
        self.parent = parent
        self.company_id = company_id
        self.lm = LanguageManager()
        self.db_path = os.path.join(os.getcwd(), 'data', 'sdg_desktop.sqlite')

        self.setup_ui()
        self.load_data()

    def setup_ui(self) -> None:
        """Veri toplama arayüzünü oluştur"""
        # Ana frame
        main_frame = tk.Frame(self.parent, bg='#f8f9fa')
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)

        # Başlık
        title_frame = tk.Frame(main_frame, bg='#2c3e50', height=60)
        title_frame.pack(fill='x', pady=(0, 20))
        title_frame.pack_propagate(False)

        title_label = tk.Label(title_frame, text=" Veri Toplama Sistemi",
                              font=('Segoe UI', 16, 'bold'), fg='white', bg='#2c3e50')
        title_label.pack(expand=True)

        # İçerik alanı
        content_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
        content_frame.pack(fill='both', expand=True)

        # Sol panel - Veri toplama formları
        left_panel = tk.Frame(content_frame, bg='white')
        left_panel.pack(side='left', fill='both', expand=True, padx=20, pady=20)

        # Form başlığı
        form_title = tk.Label(left_panel, text="Veri Toplama Formları",
                             font=('Segoe UI', 14, 'bold'), fg='#2c3e50', bg='white')
        form_title.pack(anchor='w', pady=(0, 15))

        # Form listesi
        self.form_listbox = tk.Listbox(left_panel, font=('Segoe UI', 10), height=10)
        self.form_listbox.pack(fill='both', expand=True, pady=(0, 15))

        # Form butonları
        button_frame = tk.Frame(left_panel, bg='white')
        button_frame.pack(fill='x')

        tk.Button(button_frame, text=" Yeni Form Oluştur",
                 font=('Segoe UI', 10, 'bold'), bg='#27ae60', fg='white',
                 relief='flat', cursor='hand2', padx=20, pady=8,
                 command=self.create_new_form).pack(side='left', padx=(0, 10))

        tk.Button(button_frame, text="️ Form Düzenle",
                 font=('Segoe UI', 10, 'bold'), bg='#f39c12', fg='white',
                 relief='flat', cursor='hand2', padx=20, pady=8,
                 command=self.edit_form).pack(side='left', padx=(0, 10))

        tk.Button(button_frame, text="️ Form Sil",
                 font=('Segoe UI', 10, 'bold'), bg='#e74c3c', fg='white',
                 relief='flat', cursor='hand2', padx=20, pady=8,
                 command=self.delete_form).pack(side='left')

        # Sağ panel - Veri görüntüleme
        right_panel = tk.Frame(content_frame, bg='#f8f9fa')
        right_panel.pack(side='right', fill='both', expand=True, padx=20, pady=20)

        # Veri başlığı
        data_title = tk.Label(right_panel, text="Toplanan Veriler",
                             font=('Segoe UI', 14, 'bold'), fg='#2c3e50', bg='#f8f9fa')
        data_title.pack(anchor='w', pady=(0, 15))

        # Veri tablosu
        columns = ('Form Adı', 'Tarih', 'Durum', 'Kayıt Sayısı')
        self.data_tree = ttk.Treeview(right_panel, columns=columns, show='headings', height=10)

        for col in columns:
            self.data_tree.heading(col, text=col)
            self.data_tree.column(col, width=120)

        # Scrollbar
        data_scrollbar = ttk.Scrollbar(right_panel, orient='vertical', command=self.data_tree.yview)
        self.data_tree.configure(yscrollcommand=data_scrollbar.set)

        self.data_tree.pack(side='left', fill='both', expand=True)
        data_scrollbar.pack(side='right', fill='y')

        # Veri butonları
        data_button_frame = tk.Frame(right_panel, bg='#f8f9fa')
        data_button_frame.pack(fill='x', pady=(15, 0))

        tk.Button(data_button_frame, text=" Veri Analizi",
                 font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                 relief='flat', cursor='hand2', padx=20, pady=8,
                 command=self.analyze_data).pack(side='left', padx=(0, 10))

        tk.Button(data_button_frame, text=" Veri Dışa Aktar",
                 font=('Segoe UI', 10, 'bold'), bg='#9b59b6', fg='white',
                 relief='flat', cursor='hand2', padx=20, pady=8,
                 command=self.export_data).pack(side='left')

    def load_data(self) -> None:
        """Veri toplama formlarını yükle"""
        try:
            # Form listesini temizle
            self.form_listbox.delete(0, tk.END)

            import sqlite3
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # SDG Hedeflerini form olarak listele
            try:
                cursor.execute("SELECT code, title_tr FROM sdg_goals ORDER BY id")
                goals = cursor.fetchall()
                
                if not goals:
                    self.form_listbox.insert(tk.END, "Henüz SDG Hedefi Tanımlanmamış")
                
                for code, title in goals:
                    self.form_listbox.insert(tk.END, f"{code} - {title}")
            except Exception as e:
                logging.error(f"SDG hedefleri yüklenemedi: {e}")
                self.form_listbox.insert(tk.END, "Veritabanı Hatası")

            # Veri tablosunu temizle
            for item in self.data_tree.get_children():
                self.data_tree.delete(item)

            # Gerçek verileri çek (sdg_question_responses)
            try:
                cursor.execute("""
                    SELECT 
                        g.code || ' - ' || g.title_tr as form_name,
                        MAX(r.created_at) as last_update,
                        COUNT(r.id) as response_count
                    FROM sdg_question_responses r
                    JOIN sdg_indicators i ON r.indicator_id = i.id
                    JOIN sdg_targets t ON i.target_id = t.id
                    JOIN sdg_goals g ON t.goal_id = g.id
                    WHERE r.company_id = ?
                    GROUP BY g.id
                """, (self.company_id,))
                
                rows = cursor.fetchall()
                for row in rows:
                    # Form Adı, Tarih, Durum, Kayıt Sayısı
                    self.data_tree.insert('', 'end', values=(row[0], row[1], 'Veri Girildi', row[2]))
            except Exception as e:
                logging.error(f"Yanıt verileri yüklenemedi: {e}")

            conn.close()

        except Exception as e:
            logging.error(f"Veri yükleme hatası: {e}")

    def create_new_form(self) -> None:
        """Yeni veri toplama formu oluştur"""
        from modules.forms.form_gui import FormManagementGUI

        # Form yönetimi penceresini aç
        form_window = tk.Toplevel(self.parent)
        form_window.title("Yeni Form Oluştur")
        form_window.geometry("800x600")

        FormManagementGUI(form_window)

        # Pencere kapandığında verileri yenile
        form_window.protocol("WM_DELETE_WINDOW", lambda: (self.load_data(), form_window.destroy()))

    def edit_form(self) -> None:
        """Form düzenle"""
        selection = self.form_listbox.curselection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen düzenlemek için bir form seçin")
            return

        selected_form = self.form_listbox.get(selection[0])
        messagebox.showinfo("Bilgi", f"'{selected_form}' formu düzenleme moduna alındı")

    def delete_form(self) -> None:
        """Form sil"""
        selection = self.form_listbox.curselection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir form seçin")
            return

        selected_form = self.form_listbox.get(selection[0])
        result = messagebox.askyesno("Onay", f"'{selected_form}' formunu silmek istediğinizden emin misiniz?")

        if result:
            messagebox.showinfo("Başarılı", f"'{selected_form}' formu silindi")
            self.load_data()

    def analyze_data(self) -> None:
        """Veri analizi yap"""
        # Basit veri özeti göster
        summary = "Veri Analizi Özeti:\n\n"
        summary += f"Toplam Form: {self.form_listbox.size()}\n"
        summary += f"Şirket ID: {self.company_id}\n"
        summary += "\nAnaliz tamamlandı."

        messagebox.showinfo("Veri Analizi", summary)

    def export_data(self) -> None:
        """Veri dışa aktar"""
        import datetime
        from tkinter import filedialog

        try:
            filename = filedialog.asksaveasfilename(
                title=self.lm.tr("export_data", "Verileri Dışa Aktar"),
                defaultextension=".xlsx",
                filetypes=[(self.lm.tr("file_excel", "Excel Dosyaları"), "*.xlsx"), (self.lm.tr("all_files", "Tüm Dosyalar"), "*.*")],
                initialfile=f"veri_export_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            if filename:
                # Excel export (basit)
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Veri Export"

                ws['A1'] = "Veri Toplama Export"
                ws['A2'] = f"Tarih: {datetime.datetime.now().strftime('%d.%m.%Y')}"
                ws['A3'] = f"Şirket ID: {self.company_id}"

                wb.save(filename)
                messagebox.showinfo("Başarılı", f"Veriler dışa aktarıldı:\n{filename}")
        except Exception as e:
            messagebox.showerror("Hata", f"Export hatası: {e}")

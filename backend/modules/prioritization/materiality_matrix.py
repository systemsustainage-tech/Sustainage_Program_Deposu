#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MATERİALİTE MATRİSİ GÖRSELLEŞTIRME MODÜLÜ
- Scatter plot matris oluşturma
- Otomatik materyal konu belirleme
- Renk kodlamalı göstergeler
- Excel/PDF export
"""

import logging
import os
import sqlite3
from datetime import datetime
from typing import Dict, List, Tuple

import matplotlib.patches as mpatches
import pandas as pd
from matplotlib.figure import Figure
from config.database import DB_PATH


class MaterialityMatrix:
    """Materialite Matrisi Görselleştirme Sınıfı"""

    # Renk paleti
    COLORS = {
        'high': '#e74c3c',      # Kırmızı - Yüksek öncelik
        'medium': '#f39c12',    # Turuncu - Orta öncelik
        'low': '#95a5a6',       # Gri - Düşük öncelik
        'grid': '#bdc3c7',      # Açık gri - Grid çizgileri
        'threshold': '#e74c3c'  # Kırmızı - Eşik çizgisi
    }

    # Öncelik eşikleri (1-5 skalasında)
    THRESHOLDS = {
        'high': 3.5,    # 3.5 ve üzeri yüksek
        'medium': 2.5   # 2.5-3.5 arası orta
    }

    def __init__(self, db_path: str = DB_PATH) -> None:
        """Init"""
        if not os.path.isabs(db_path):
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
            db_path = os.path.join(base_dir, db_path)
        self.db_path = db_path

    def get_connection(self):
        """Veritabanı bağlantısı"""
        return sqlite3.connect(self.db_path)

    def get_materiality_topics(self, company_id: int) -> List[Dict]:
        """Materialite konularını getir"""
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT 
                    id,
                    topic_name,
                    category,
                    description,
                    stakeholder_impact,
                    business_impact,
                    priority_score,
                    sdg_mapping
                FROM materiality_topics
                WHERE company_id = ?
                ORDER BY priority_score DESC
            """, (company_id,))

            topics = []
            for row in cursor.fetchall():
                topics.append({
                    'id': row[0],
                    'topic_name': row[1],
                    'category': row[2],
                    'description': row[3],
                    'stakeholder_impact': row[4] or 0,
                    'business_impact': row[5] or 0,
                    'priority_score': row[6] or 0,
                    'sdg_mapping': row[7]
                })

            return topics

        except Exception as e:
            logging.info(f" Konular getirilemedi: {e}")
            return []
        finally:
            conn.close()

    def determine_priority(self, stakeholder_impact: float, business_impact: float) -> str:
        """Öncelik seviyesi belirle"""
        # Ortalama değer
        avg_impact = (stakeholder_impact + business_impact) / 2

        if avg_impact >= self.THRESHOLDS['high']:
            return 'high'
        elif avg_impact >= self.THRESHOLDS['medium']:
            return 'medium'
        else:
            return 'low'

    def get_priority_color(self, priority: str) -> str:
        """Öncelik rengini getir"""
        return self.COLORS.get(priority, self.COLORS['low'])

    def create_matrix_plot(self, company_id: int, figsize: Tuple[int, int] = (10, 8)) -> Figure:
        """Materialite matrisini oluştur"""
        logging.info(f" Materialite matrisi oluşturuluyor (Company ID: {company_id})...")

        # Konuları getir
        topics = self.get_materiality_topics(company_id)

        if not topics:
            logging.info("️ Hiç materyal konu bulunamadı!")
            # Boş figür oluştur
            fig = Figure(figsize=figsize)
            ax = fig.add_subplot(111)
            ax.text(0.5, 0.5, 'Henüz materyal konu eklenmemiş.\n\nÖncelikle konuları ekleyin.',
                   ha='center', va='center', fontsize=14, color='gray')
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            return fig

        # Figure oluştur
        fig = Figure(figsize=figsize, facecolor='white')
        ax = fig.add_subplot(111)

        # Verileri hazırla
        x_values = []  # Stakeholder Impact
        y_values = []  # Business Impact
        colors = []
        labels = []
        sizes = []

        for topic in topics:
            x_values.append(topic['stakeholder_impact'])
            y_values.append(topic['business_impact'])

            priority = self.determine_priority(
                topic['stakeholder_impact'],
                topic['business_impact']
            )
            colors.append(self.get_priority_color(priority))
            labels.append(topic['topic_name'])

            # Boyut - öncelik skoruna göre
            size = 100 + (topic['priority_score'] * 10)
            sizes.append(size)

        # Şekil uyumluluğu kontrolü
        n_x = len(x_values)
        n_y = len(y_values)
        n_s = len(sizes)
        n_c = len(colors)
        n_l = len(labels)
        if not (n_x == n_y == n_s == n_c == n_l):
            raise ValueError(
                f"Shape mismatch: x={n_x}, y={n_y}, sizes={n_s}, colors={n_c}, labels={n_l}"
            )
        # Scatter plot çiz
        ax.scatter(x_values, y_values, c=colors, s=sizes, alpha=0.6,
                   edgecolors='black', linewidths=1.5)

        # Eşik çizgileri (3.5 üzeri yüksek öncelik)
        ax.axhline(y=self.THRESHOLDS['high'], color=self.COLORS['threshold'],
                  linestyle='--', linewidth=1, alpha=0.3, label='Yüksek Öncelik Eşiği')
        ax.axvline(x=self.THRESHOLDS['high'], color=self.COLORS['threshold'],
                  linestyle='--', linewidth=1, alpha=0.3)

        # Orta öncelik eşikleri
        ax.axhline(y=self.THRESHOLDS['medium'], color=self.COLORS['medium'],
                  linestyle=':', linewidth=1, alpha=0.2)
        ax.axvline(x=self.THRESHOLDS['medium'], color=self.COLORS['medium'],
                  linestyle=':', linewidth=1, alpha=0.2)

        # Etiketler ekle (sadece yüksek öncelikli konular)
        for i, (x, y, label, color) in enumerate(zip(x_values, y_values, labels, colors)):
            if color == self.COLORS['high']:  # Sadece kırmızıları etiketle
                ax.annotate(label, (x, y), xytext=(5, 5), textcoords='offset points',
                          fontsize=8, fontweight='bold',
                          bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                                  edgecolor=color, alpha=0.7))

        # Grid
        ax.grid(True, linestyle=':', alpha=0.3, color=self.COLORS['grid'])

        # Eksen ayarları
        ax.set_xlim(0, 5.5)
        ax.set_ylim(0, 5.5)
        ax.set_xlabel('Paydaş Etkisi (Stakeholder Impact)', fontsize=12, fontweight='bold')
        ax.set_ylabel('İşletme Etkisi (Business Impact)', fontsize=12, fontweight='bold')
        ax.set_title('Materialite Matrisi - Öncelikli Konular',
                    fontsize=14, fontweight='bold', pad=20)

        # Lejant
        high_patch = mpatches.Patch(color=self.COLORS['high'], label='Yüksek Öncelik (>3.5)')
        medium_patch = mpatches.Patch(color=self.COLORS['medium'], label='Orta Öncelik (2.5-3.5)')
        low_patch = mpatches.Patch(color=self.COLORS['low'], label='Düşük Öncelik (<2.5)')
        ax.legend(handles=[high_patch, medium_patch, low_patch], loc='upper left', fontsize=10)

        # Bölge etiketleri
        ax.text(4.5, 4.5, 'YÜKSEKÖNCELİK', ha='center', va='center',
               fontsize=10, color=self.COLORS['high'], alpha=0.3, fontweight='bold')
        ax.text(1.5, 1.5, 'DÜŞÜK\nÖNCELİK', ha='center', va='center',
               fontsize=10, color=self.COLORS['low'], alpha=0.3, fontweight='bold')

        # Layout ayarları
        fig.tight_layout()

        logging.info(f" Matris oluşturuldu: {len(topics)} konu")
        return fig

    def export_to_excel(self, company_id: int, output_path: str) -> bool:
        """Materialite analizi Excel'e export et"""
        try:
            logging.info(f" Materialite analizi Excel export başlıyor: {output_path}")

            # Konuları getir
            topics = self.get_materiality_topics(company_id)

            if not topics:
                logging.info("️ Export edilecek konu bulunamadı!")
                return False

            # DataFrame oluştur
            data = []
            for topic in topics:
                priority = self.determine_priority(
                    topic['stakeholder_impact'],
                    topic['business_impact']
                )

                # Öncelik etiketi
                if priority == 'high':
                    priority_label = ' Yüksek Öncelik'
                elif priority == 'medium':
                    priority_label = ' Orta Öncelik'
                else:
                    priority_label = ' Düşük Öncelik'

                data.append({
                    'Konu': topic['topic_name'],
                    'Kategori': topic['category'],
                    'Açıklama': topic['description'],
                    'Paydaş Etkisi (1-5)': topic['stakeholder_impact'],
                    'İşletme Etkisi (1-5)': topic['business_impact'],
                    'Öncelik Skoru': topic['priority_score'],
                    'Öncelik Seviyesi': priority_label,
                    'SDG İlişkilendirmesi': topic['sdg_mapping']
                })

            df = pd.DataFrame(data)

            # Excel writer oluştur
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. Analiz Sayfası
                df.to_excel(writer, sheet_name='Materialite Analizi', index=False)

                # 2. Özet Sayfası
                summary_data = {
                    'Metrik': [
                        'Toplam Konu Sayısı',
                        'Yüksek Öncelikli',
                        'Orta Öncelikli',
                        'Düşük Öncelikli',
                        'Ortalama Paydaş Etkisi',
                        'Ortalama İşletme Etkisi',
                        'Analiz Tarihi'
                    ],
                    'Değer': [
                        len(topics),
                        len([t for t in topics if self.determine_priority(t['stakeholder_impact'], t['business_impact']) == 'high']),
                        len([t for t in topics if self.determine_priority(t['stakeholder_impact'], t['business_impact']) == 'medium']),
                        len([t for t in topics if self.determine_priority(t['stakeholder_impact'], t['business_impact']) == 'low']),
                        round(sum(t['stakeholder_impact'] for t in topics) / len(topics), 2),
                        round(sum(t['business_impact'] for t in topics) / len(topics), 2),
                        datetime.now().strftime('%Y-%m-%d')
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Özet', index=False)

                # Formatla

                # Analiz sayfası formatı
                worksheet = writer.sheets['Materialite Analizi']
                from openpyxl.styles import Alignment, Font, PatternFill

                # Başlık satırı
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(1, col_num)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for row_num in range(2, len(df) + 2):
                    priority_cell = worksheet.cell(row_num, 7)
                    val = str(priority_cell.value or '')
                    if 'Yüksek' in val:
                        color = 'FFC7CE'
                    elif 'Orta' in val:
                        color = 'FFEB9C'
                    else:
                        color = 'E8E8E8'
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row_num, col_num)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                # Kolon genişlikleri
                worksheet.column_dimensions['A'].width = 30  # Konu
                worksheet.column_dimensions['B'].width = 20  # Kategori
                worksheet.column_dimensions['C'].width = 50  # Açıklama
                worksheet.column_dimensions['D'].width = 20  # Paydaş
                worksheet.column_dimensions['E'].width = 20  # İşletme
                worksheet.column_dimensions['F'].width = 15  # Skor
                worksheet.column_dimensions['G'].width = 20  # Öncelik
                worksheet.column_dimensions['H'].width = 30  # SDG

            logging.info(f" Excel export başarılı: {output_path}")
            return True

        except Exception as e:
            logging.error(f" Excel export hatası: {e}")
            import traceback
            traceback.print_exc()
            return False

    def save_matrix_image(self, company_id: int, output_path: str) -> bool:
        """Matris görselini PNG olarak kaydet"""
        try:
            logging.info(f"️ Matris görseli kaydediliyor: {output_path}")

            # Matris oluştur
            fig = self.create_matrix_plot(company_id, figsize=(12, 10))

            # Kaydet
            fig.savefig(output_path, dpi=300, bbox_inches='tight',
                       facecolor='white', edgecolor='none')

            logging.info(f" Görsel kaydedildi: {output_path}")
            return True

        except Exception as e:
            logging.error(f" Görsel kaydetme hatası: {e}")
            return False

    def get_material_topics(self, company_id: int, threshold: float = 3.5) -> List[Dict]:
        """Materyal (öncelikli) konuları getir"""
        topics = self.get_materiality_topics(company_id)

        material_topics = []
        for topic in topics:
            priority = self.determine_priority(
                topic['stakeholder_impact'],
                topic['business_impact']
            )

            if priority == 'high':
                material_topics.append(topic)

        return material_topics

    def generate_materiality_report(self, company_id: int) -> str:
        """Materialite raporu metni oluştur"""
        topics = self.get_materiality_topics(company_id)
        material_topics = self.get_material_topics(company_id)

        report = f"""
MATERYAL KONU ANALİZİ RAPORU
================================

Analiz Tarihi: {datetime.now().strftime('%d.%m.%Y')}
Şirket ID: {company_id}

GENEL ÖZET
----------
Toplam Konu: {len(topics)}
Materyal (Yüksek Öncelikli) Konu: {len(material_topics)}
Materyal Konu Oranı: {round(len(material_topics)/len(topics)*100, 1) if topics else 0}%

MATERYAL KONULAR (Yüksek Öncelik)
----------------------------------
"""

        for i, topic in enumerate(material_topics, 1):
            report += f"""
{i}. {topic['topic_name']}
   Kategori: {topic['category']}
   Paydaş Etkisi: {topic['stakeholder_impact']}/5
   İşletme Etkisi: {topic['business_impact']}/5
   Öncelik Skoru: {topic['priority_score']}
   {'-' * 50}
"""

        return report


if __name__ == "__main__":
    # Test
    logging.info(" Materialite Matrix Test Başlıyor...")

    matrix = MaterialityMatrix()

    # Test: Matris oluştur
    fig = matrix.create_matrix_plot(company_id=1)

    # Test: Görsel kaydet
    output_image = "modules/prioritization/materiality_matrix_test.png"
    if matrix.save_matrix_image(company_id=1, output_path=output_image):
        logging.info(f" Test görseli oluşturuldu: {output_image}")

    # Test: Excel export
    output_excel = "modules/prioritization/materiality_analysis_test.xlsx"
    if matrix.export_to_excel(company_id=1, output_path=output_excel):
        logging.info(f" Test Excel oluşturuldu: {output_excel}")

    # Test: Rapor
    report = matrix.generate_materiality_report(company_id=1)
    logging.info(report)

    logging.info(" Test tamamlandı!")


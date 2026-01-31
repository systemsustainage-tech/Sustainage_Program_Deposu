import logging
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TCFD GUI - Kullanıcı Arayüzü
- 4 ana bölüm: Governance, Strategy, Risk Management, Metrics
- İklim risk yönetimi
- Senaryo analizi
- Finansal etki hesaplama
- Rapor oluşturma
"""

import csv
import json
import os
import sys
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, scrolledtext, ttk
from typing import Dict, List, Optional

import requests
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

from utils.language_manager import LanguageManager
from utils.ui_theme import apply_theme

# Modül importları
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


from modules.tcfd.tcfd_calculator import TCFDCalculator
from modules.tcfd.tcfd_manager import TCFDManager
from modules.tcfd.tcfd_report_generator import TCFDReportGenerator


class TCFDGUI:
    """TCFD modülü ana GUI"""

    def __init__(self, parent, company_id: int = 1, user=None, db_path: str = None, main_app=None):
        """
        Args:
            parent: Ana pencere
            company_id: Firma ID
            user: Kullanıcı bilgileri (tuple: id, username, display_name)
            db_path: Veritabanı yolu
            main_app: Ana uygulama referansı
        """
        self.parent = parent
        self.company_id = company_id
        self.main_app = main_app
        
        # Kullanıcı bilgisi
        if user is not None:
            self.user = user
        elif main_app is not None and hasattr(main_app, 'user'):
            self.user = main_app.user
        else:
            self.user = None

        # Veritabanı yolu
        if db_path:
            self.db_path = db_path
        elif main_app is not None and hasattr(main_app, 'db_path'):
            self.db_path = main_app.db_path
        else:
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            self.db_path = os.path.join(base_dir, "data", "sdg_desktop.sqlite")

        # Manager ve Calculator
        try:
            self.lm = LanguageManager()
            self.manager = TCFDManager(self.db_path)
            self.calculator = TCFDCalculator()
        except Exception as e:
            # Fallback if self.lm is not initialized
            error_title = "Hata"
            error_msg = "TCFD Manager başlatılamadı"
            if hasattr(self, 'lm'):
                try:
                    error_title = self.lm.tr('error', "Hata")
                    error_msg = self.lm.tr('tcfd_init_error', 'TCFD Manager başlatılamadı')
                except Exception as e:
                    logging.warning(f"Translation failed in error handler: {e}")
            messagebox.showerror(error_title, f"{error_msg}:\n{e}")
            return

        # Raporlama yılı (varsayılan: bu yıl)
        self.current_year = datetime.now().year

        # Tema
        self.colors = {
            'primary': '#2c3e50',   # Standard Dark Blue
            'secondary': '#34495e', # Standard Secondary Blue
            'danger': '#e74c3c',
            'warning': '#f1c40f',
            'info': '#3498db',
            'success': '#2ecc71',
            'bg': '#f0f2f5',        # Standard Background
            'white': '#FFFFFF'
        }

        self.setup_ui()

    def setup_ui(self) -> None:
        """Ana UI oluştur"""
        apply_theme(self.parent)
        # Ana frame
        self.main_frame = tk.Frame(self.parent, bg=self.colors['bg'])
        self.main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Üst bar (başlık ve yıl seçici)
        self.create_header()

        # Tab control (4 ana bölüm)
        self.create_tabs()

    def create_header(self) -> None:
        """Üst başlık bar"""
        header = tk.Frame(self.main_frame, bg=self.colors['primary'], height=80)
        header.pack(fill='x', side='top', pady=(0, 20))
        header.pack_propagate(False)

        # Geri Butonu
        back_btn = tk.Button(
            header,
            text=self.lm.tr("back_button", "< Geri"),
            command=self.go_back,
            bg=self.colors['secondary'],
            fg='white',
            font=('Segoe UI', 10, 'bold'),
            relief='raised',
            bd=1,
            activebackground=self.colors['success'],
            activeforeground='white',
            cursor='hand2'
        )
        back_btn.pack(side='left', padx=(20, 0), pady=10)

        # Başlık
        title_frame = tk.Frame(header, bg=self.colors['primary'])
        title_frame.pack(side='left', padx=20, pady=10)

        tk.Label(
            title_frame,
            text=self.lm.tr("tcfd_title", " TCFD - İklim İle İlgili Finansal Açıklamalar"),
            font=('Segoe UI', 16, 'bold'),
            bg=self.colors['primary'],
            fg='white'
        ).pack(anchor='w')

        tk.Label(
            title_frame,
            text="Task Force on Climate-related Financial Disclosures",
            font=('Segoe UI', 9),
            bg=self.colors['primary'],
            fg='#E0F2E0'
        ).pack(anchor='w')

        # Sağ taraf - yıl seçici ve işlemler
        right_frame = tk.Frame(header, bg=self.colors['primary'])
        right_frame.pack(side='right', padx=20, pady=10)

        ttk.Button(right_frame, text=self.lm.tr("report_center", " Rapor Merkezi"), style='Primary.TButton',
                   command=self.open_report_center_tcfd).pack(side='left', padx=5)

        # Yıl seçici
        year_frame = tk.Frame(right_frame, bg=self.colors['white'], bd=1, relief='solid')
        year_frame.pack(side='left', padx=5)

        tk.Label(
            year_frame,
            text=self.lm.tr("reporting_year", "Raporlama Yılı:"),
            font=('Segoe UI', 9),
            bg=self.colors['white']
        ).pack(side='left', padx=5)

        self.year_var = tk.IntVar(value=self.current_year)
        year_combo = ttk.Combobox(
            year_frame,
            textvariable=self.year_var,
            values=list(range(2020, 2031)),
            width=8,
            state='readonly'
        )
        year_combo.pack(side='left', padx=5, pady=5)
        year_combo.bind('<<ComboboxSelected>>', lambda e: self.on_year_changed())

        # Rapor butonu
        tk.Button(
            right_frame,
            text=self.lm.tr("create_report", " Rapor Oluştur"),
            command=self.create_report,
            bg=self.colors['info'],
            fg='white',
            font=('Segoe UI', 10, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=15,
            pady=8
        ).pack(side='left', padx=5)

    def go_back(self) -> None:
        """Ana ekrana dön"""
        try:
            self.main_frame.destroy()
        except Exception as e:
            logging.error(f"Geri dönme hatası: {e}")

    def open_report_center_tcfd(self) -> None:
        try:
            from modules.reporting.report_center_gui import ReportCenterGUI
            win = tk.Toplevel(self.parent)
            gui = ReportCenterGUI(win, self.company_id)
            try:
                gui.module_filter_var.set('tcfd')
                gui.refresh_reports()
            except Exception as e:
                logging.error(f"Error filtering reports for tcfd: {e}")
        except Exception as e:
            messagebox.showerror(self.lm.tr('error', "Hata"), f"{self.lm.tr('report_center_error', 'Rapor Merkezi açılamadı')}:\n{e}")
            logging.error(f"Error opening report center: {e}")

    def create_tabs(self) -> None:
        """4 ana bölüm için sekmeler"""
        tab_frame = tk.Frame(self.main_frame, bg=self.colors['bg'])
        tab_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Notebook (tab control)
        self.notebook = ttk.Notebook(tab_frame)
        self.notebook.pack(fill='both', expand=True)

        # Style
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook.Tab', font=('Segoe UI', 10), padding=[20, 10])

        # 4 ana tab
        self.tab_governance = tk.Frame(self.notebook, bg=self.colors['bg'])
        self.tab_strategy = tk.Frame(self.notebook, bg=self.colors['bg'])
        self.tab_risk = tk.Frame(self.notebook, bg=self.colors['bg'])
        self.tab_metrics = tk.Frame(self.notebook, bg=self.colors['bg'])

        self.notebook.add(self.tab_governance, text=self.lm.tr("tcfd_tab_governance", "️ Yönetişim (Governance)"))
        self.notebook.add(self.tab_strategy, text=self.lm.tr("tcfd_tab_strategy", " Strateji (Strategy)"))
        self.notebook.add(self.tab_risk, text=self.lm.tr("tcfd_tab_risk", "️ Risk Yönetimi"))
        self.notebook.add(self.tab_metrics, text=self.lm.tr("tcfd_tab_metrics", " Metrikler ve Hedefler"))

        # Her sekmeyi oluştur
        self.create_governance_tab()
        self.create_strategy_tab()
        self.create_risk_tab()
        self.create_metrics_tab()

    # ========================================================================
    # 1. GOVERNANCE TAB
    # ========================================================================

    def create_governance_tab(self) -> None:
        """Yönetişim sekmesi"""
        # Scroll frame
        canvas = tk.Canvas(self.tab_governance, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_governance, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=self.colors['bg'])

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # İçerik
        content = tk.Frame(scroll_frame, bg=self.colors['bg'])
        content.pack(fill='both', expand=True, padx=10, pady=10)

        # Başlık
        tk.Label(
            content,
            text=self.lm.tr("tcfd_gov_title", "Yönetişim - Yönetim Kurulu ve Üst Yönetim Gözetimi"),
            font=('Segoe UI', 14, 'bold'),
            bg=self.colors['bg']
        ).pack(anchor='w', pady=(0, 10))

        # Section 1: Yönetim Kurulu
        self.create_section_card(content, self.lm.tr("tcfd_gov_sec1", "Yönetim Kurulu Gözetimi"), [
            (self.lm.tr("tcfd_gov_q1", "YK'nın İklim Risklerini Gözetimi:"), 'board_oversight', 'text', 5),
            (self.lm.tr("tcfd_gov_q2", "YK Toplantı Sıklığı:"), 'board_frequency', 'combo',
             [self.lm.tr("freq_annual", "Yıllık"), self.lm.tr("freq_semiannual", "Yarı Yıllık"), 
              self.lm.tr("freq_quarterly", "Çeyreklik"), self.lm.tr("freq_monthly", "Aylık")]),
            (self.lm.tr("tcfd_gov_q3", "YK'da İklim Uzmanlığı:"), 'board_expertise', 'text', 3),
        ])

        # Section 2: Üst Yönetim
        self.create_section_card(content, self.lm.tr("tcfd_gov_sec2", "Üst Yönetimin Rolü"), [
            (self.lm.tr("tcfd_gov_q4", "Üst Yönetimin İklim Konusundaki Rolü:"), 'management_role', 'text', 5),
            (self.lm.tr("tcfd_gov_q5", "Organizasyon Yapısı:"), 'management_structure', 'text', 3),
            (self.lm.tr("tcfd_gov_q6", "Sorumluluk Dağılımı:"), 'management_responsibility', 'text', 3),
        ])

        # Section 3: Komiteler
        self.create_section_card(content, self.lm.tr("tcfd_gov_sec3", "İklim Komiteleri ve Birimler"), [
            (self.lm.tr("tcfd_gov_q7", "İklim Komitesi Var mı?"), 'climate_committee', 'check', None),
            (self.lm.tr("tcfd_gov_q8", "Komite/Birim Adı:"), 'committee_name', 'entry', None),
            (self.lm.tr("tcfd_gov_q9", "Komite Üyeleri:"), 'committee_members', 'text', 3),
            (self.lm.tr("tcfd_gov_q10", "Sorumlu Yönetici:"), 'responsible_executive', 'entry', None),
        ])

        # Section 4: Entegrasyon
        self.create_section_card(content, self.lm.tr("tcfd_gov_sec4", "Stratejiye Entegrasyon"), [
            (self.lm.tr("tcfd_gov_q11", "İklim Konularının Stratejiye Entegrasyonu:"), 'strategy_integration', 'text', 4),
            (self.lm.tr("tcfd_gov_q12", "Risk Yönetimine Entegrasyon:"), 'risk_integration', 'text', 4),
        ])

        # Kaydet butonu
        btn_frame = tk.Frame(content, bg=self.colors['bg'])
        btn_frame.pack(fill='x', pady=20)

        tk.Button(
            btn_frame,
            text=self.lm.tr("save_gov_data", " Yönetişim Verilerini Kaydet"),
            command=self.save_governance,
            bg=self.colors['success'],
            fg='white',
            font=('Segoe UI', 11, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=30,
            pady=12
        ).pack()

        # Veri storage
        self.gov_widgets = {}

    def create_section_card(self, parent, title: str, fields: List[tuple]) -> None:
        """
        Bölüm kartı oluştur
        
        Args:
            parent: Üst widget
            title: Bölüm başlığı
            fields: [(label, key, type, param), ...]
                    type: 'entry', 'text', 'combo', 'check'
        """
        # Kart çerçevesi
        card = tk.Frame(parent, bg=self.colors['white'], relief='raised', bd=1)
        card.pack(fill='x', pady=10)

        # Başlık
        header = tk.Frame(card, bg=self.colors['primary'], height=40)
        header.pack(fill='x')
        header.pack_propagate(False)

        tk.Label(
            header,
            text=title,
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white'
        ).pack(side='left', padx=15, pady=8)

        # İçerik
        content = tk.Frame(card, bg=self.colors['white'])
        content.pack(fill='both', expand=True, padx=15, pady=15)

        # Alanları oluştur
        for field_info in fields:
            label_text = field_info[0]
            key = field_info[1]
            field_type = field_info[2]
            param = field_info[3] if len(field_info) > 3 else None

            field_frame = tk.Frame(content, bg=self.colors['white'])
            field_frame.pack(fill='x', pady=5)

            # Label
            tk.Label(
                field_frame,
                text=label_text,
                font=('Segoe UI', 10),
                bg=self.colors['white'],
                anchor='w'
            ).pack(anchor='w', pady=(0, 3))

            # Widget
            if field_type == 'entry':
                widget = tk.Entry(
                    field_frame,
                    font=('Segoe UI', 10),
                    relief='solid',
                    bd=1
                )
                widget.pack(fill='x')

            elif field_type == 'text':
                rows = param or 3
                widget = scrolledtext.ScrolledText(
                    field_frame,
                    font=('Segoe UI', 10),
                    height=rows,
                    relief='solid',
                    bd=1,
                    wrap='word'
                )
                widget.pack(fill='x')

            elif field_type == 'combo':
                widget = ttk.Combobox(
                    field_frame,
                    values=param or [],
                    font=('Segoe UI', 10),
                    state='readonly'
                )
                widget.pack(fill='x')
                if param:
                    widget.set(param[0])

            elif field_type == 'check':
                var = tk.BooleanVar()
                widget = tk.Checkbutton(
                    field_frame,
                    text="Evet",
                    variable=var,
                    font=('Segoe UI', 10),
                    bg=self.colors['white']
                )
                widget.pack(anchor='w')
                widget.var = var  # Reference kaydet

            else:
                widget = None

            # Widget'ı kaydet
            if not hasattr(self, 'gov_widgets'):
                self.gov_widgets = {}
            self.gov_widgets[key] = widget

    def save_governance(self) -> None:
        """Yönetişim verilerini kaydet"""
        try:
            data = {}

            for key, widget in self.gov_widgets.items():
                if isinstance(widget, tk.Entry):
                    data[key] = widget.get()
                elif isinstance(widget, scrolledtext.ScrolledText):
                    data[key] = widget.get('1.0', 'end-1c')
                elif isinstance(widget, ttk.Combobox):
                    data[key] = widget.get()
                elif isinstance(widget, tk.Checkbutton):
                    data[key] = widget.var.get()

            # Kaydet
            success, message = self.manager.save_governance(
                self.company_id,
                self.year_var.get(),
                data
            )

            if success:
                messagebox.showinfo("Başarılı", message)
            else:
                messagebox.showerror("Hata", message)

        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası:\n{e}")

    def load_governance(self) -> None:
        """Yönetişim verilerini yükle"""
        try:
            data = self.manager.get_governance(self.company_id, self.year_var.get())

            if not data:
                return

            for key, widget in self.gov_widgets.items():
                value = data.get(key)
                if value is None:
                    continue

                if isinstance(widget, tk.Entry):
                    widget.delete(0, 'end')
                    widget.insert(0, str(value))
                elif isinstance(widget, scrolledtext.ScrolledText):
                    widget.delete('1.0', 'end')
                    widget.insert('1.0', str(value))
                elif isinstance(widget, ttk.Combobox):
                    widget.set(str(value))
                elif isinstance(widget, tk.Checkbutton):
                    widget.var.set(bool(value))

        except Exception as e:
            logging.info(f"[WARN] Yönetişim verileri yüklenemedi: {e}")

    def _load_scenario_data(self) -> Dict:
        try:
            module_dir = os.path.dirname(os.path.abspath(__file__))
            scenarios_path = os.path.join(module_dir, 'data', 'climate_scenarios.json')
            if os.path.exists(scenarios_path):
                with open(scenarios_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logging.info(f"[WARN] Senaryo verileri yüklenemedi: {e}")
        return {}

    def _get_scenario_carbon_price(self, scenario_data: Dict, scenario_name: str, year: int) -> Optional[float]:
        try:
            cmp = scenario_data.get('scenario_comparison_metrics', {})
            prices = cmp.get('carbon_price', {}).get('scenarios', {}).get(scenario_name)
            if not prices:
                return None
            ykey = str(year)
            if ykey in prices:
                return float(prices[ykey])
            keys = sorted(int(k) for k in prices.keys())
            if not keys:
                return None
            closest = min(keys, key=lambda k: abs(k - year))
            return float(prices[str(closest)])
        except Exception:
            return None

    def generate_scenario_timeseries(self) -> None:
        try:
            scen = self.scenario_name_var.get().strip()
            if not scen:
                messagebox.showwarning("Uyarı", "Senaryo seçiniz")
                return
            metrics = self.manager.get_metrics(self.company_id, int(self.year_var.get()))
            if not metrics:
                messagebox.showwarning("Uyarı", "Önce metrikleri kaydedin")
                return
            emissions = {
                'scope1': metrics.get('scope1_emissions') or 0,
                'scope2': metrics.get('scope2_emissions') or 0,
                'scope3': metrics.get('scope3_emissions') or 0,
            }
            sdata = self._load_scenario_data()
            years = list(range(2025, 2036))
            cprices = {y: (self._get_scenario_carbon_price(sdata, scen, y) or (metrics.get('internal_carbon_price') or 0)) for y in years}
            series = self.calculator.calculate_carbon_price_scenario_impact(emissions, cprices, start_year=2025, end_year=2035)
            for i in self.scenario_table.get_children():
                self.scenario_table.delete(i)
            xs = []
            ys = []
            for y in years:
                impact = series.get(y, {})
                price = cprices.get(y, 0)
                cost = impact.get('total_cost', 0)
                net = 0 - cost
                self.scenario_table.insert('', 'end', values=(y, f"${price:,.0f}", f"${cost:,.0f}", f"${net:,.0f}"))
                xs.append(y)
                ys.append(net)
            self.scenario_ax.clear()
            self.scenario_ax.plot(xs, ys, marker='o', color='#1E90FF')
            self.scenario_ax.set_title(f"Net Etki ({scen})")
            self.scenario_ax.set_ylabel("USD")
            self.scenario_ax.set_xlabel("Yıl")
            self.scenario_ax.grid(True, alpha=0.3)
            self.scenario_canvas.draw()
        except Exception as e:
            messagebox.showerror("Hata", f"Zaman serisi hesaplama hatası:\n{e}")

    # ========================================================================
    # 2. STRATEGY TAB
    # ========================================================================

    def create_strategy_tab(self) -> None:
        """Strateji sekmesi"""
        canvas = tk.Canvas(self.tab_strategy, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_strategy, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=self.colors['bg'])

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        content = tk.Frame(scroll_frame, bg=self.colors['bg'])
        content.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(
            content,
            text=self.lm.tr("tcfd_strat_title", "Strateji - Riskler, Fırsatlar ve Etkiler"),
            font=('Segoe UI', 14, 'bold'),
            bg=self.colors['bg']
        ).pack(anchor='w', pady=(0, 10))

        # Bölüm kartları
        self.strategy_widgets = {}

        self.create_section_card(content, self.lm.tr("tcfd_strat_sec1", "Kısa/Orta/Uzun Vadeli Riskler"), [
            (self.lm.tr("tcfd_strat_q1", "Kısa Vadeli Riskler:"), 'short_term_risks', 'text', 4),
            (self.lm.tr("tcfd_strat_q2", "Orta Vadeli Riskler:"), 'medium_term_risks', 'text', 4),
            (self.lm.tr("tcfd_strat_q3", "Uzun Vadeli Riskler:"), 'long_term_risks', 'text', 4),
        ])

        self.create_section_card(content, self.lm.tr("tcfd_strat_sec2", "Kısa/Orta/Uzun Vadeli Fırsatlar"), [
            (self.lm.tr("tcfd_strat_q4", "Kısa Vadeli Fırsatlar:"), 'short_term_opportunities', 'text', 4),
            (self.lm.tr("tcfd_strat_q5", "Orta Vadeli Fırsatlar:"), 'medium_term_opportunities', 'text', 4),
            (self.lm.tr("tcfd_strat_q6", "Uzun Vadeli Fırsatlar:"), 'long_term_opportunities', 'text', 4),
        ])

        self.create_section_card(content, self.lm.tr("tcfd_strat_sec3", "Etki Değerlendirmesi"), [
            (self.lm.tr("tcfd_strat_q7", "İş Etkisi:"), 'business_impact', 'text', 4),
            (self.lm.tr("tcfd_strat_q8", "Stratejiye Etki:"), 'strategy_impact', 'text', 4),
            (self.lm.tr("tcfd_strat_q9", "Finansal Etki:"), 'financial_impact', 'text', 4),
        ])

        self.create_section_card(content, self.lm.tr("tcfd_strat_sec4", "Dayanıklılık ve Uyum"), [
            (self.lm.tr("tcfd_strat_q10", "Dayanıklılık Değerlendirmesi:"), 'resilience_assessment', 'text', 4),
            (self.lm.tr("tcfd_strat_q11", "Uyum Planları:"), 'adaptation_plans', 'text', 4),
        ])

        sensitivity_card = tk.Frame(content, bg=self.colors['white'], relief='raised', bd=1)
        sensitivity_card.pack(fill='x', pady=10)
        sh = tk.Frame(sensitivity_card, bg=self.colors['primary'], height=40)
        sh.pack(fill='x')
        sh.pack_propagate(False)
        tk.Label(sh, text=self.lm.tr("tcfd_sensitivity_title", "Senaryo Hassasiyet Analizi"), font=('Segoe UI', 11, 'bold'), bg=self.colors['primary'], fg='white').pack(side='left', padx=15, pady=8)
        sc = tk.Frame(sensitivity_card, bg=self.colors['white'])
        sc.pack(fill='both', expand=True, padx=15, pady=15)

        if not hasattr(self, 'sensitivity_vars'):
            self.sensitivity_vars = {}
        if not hasattr(self, 'sensitivity_results'):
            self.sensitivity_results = {}

        cf_frame = tk.Frame(sc, bg=self.colors['white'])
        cf_frame.pack(fill='x', pady=6)
        tk.Label(cf_frame, text=self.lm.tr("carbon_price_adj", "Karbon Fiyat Ayarı (%)"), bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['carbon_pct'] = tk.IntVar(value=0)
        tk.Scale(cf_frame, from_=-20, to=20, orient='horizontal', variable=self.sensitivity_vars['carbon_pct']).pack(fill='x')

        ep_frame = tk.Frame(sc, bg=self.colors['white'])
        ep_frame.pack(fill='x', pady=6)
        tk.Label(ep_frame, text=self.lm.tr("energy_price_adj", "Enerji Fiyat Ayarı (%)"), bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['energy_pct'] = tk.IntVar(value=0)
        tk.Scale(ep_frame, from_=-20, to=20, orient='horizontal', variable=self.sensitivity_vars['energy_pct']).pack(fill='x')

        bp_frame = tk.Frame(sc, bg=self.colors['white'])
        bp_frame.pack(fill='x', pady=6)
        tk.Label(bp_frame, text=self.lm.tr("current_energy_price", "Mevcut Enerji Fiyatı (USD/MWh)"), bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['current_energy_price'] = tk.Entry(bp_frame, font=('Segoe UI', 10), relief='solid', bd=1)
        self.sensitivity_vars['current_energy_price'].insert(0, '100')
        self.sensitivity_vars['current_energy_price'].pack(fill='x')

        btns = tk.Frame(sc, bg=self.colors['white'])
        btns.pack(fill='x', pady=8)
        tk.Button(btns, text=self.lm.tr("calculate", " Hesapla"), command=self.calculate_sensitivity, bg=self.colors['info'], fg='white', relief='flat', padx=20, pady=8).pack(side='left')
        tk.Button(btns, text=self.lm.tr("import", " İçe Aktar"), command=self.import_sensitivity_data, bg=self.colors['success'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=6)
        tk.Button(btns, text=self.lm.tr("create_template_xlsx", " Şablon Oluştur (XLSX)"), command=self.export_sensitivity_template_xlsx, bg=self.colors['secondary'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=6)

        url_frame = tk.Frame(sc, bg=self.colors['white'])
        url_frame.pack(fill='x', pady=6)
        tk.Label(url_frame, text=self.lm.tr("data_url", "Veri URL (JSON)"), bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['data_url'] = tk.Entry(url_frame, font=('Segoe UI', 10), relief='solid', bd=1)
        self.sensitivity_vars['data_url'].pack(side='left', fill='x', expand=True)
        tk.Button(url_frame, text=self.lm.tr("fetch", " Getir"), command=self.fetch_sensitivity_data, bg=self.colors['warning'], fg='white', relief='flat', padx=16, pady=6).pack(side='left', padx=6)

        res = tk.Frame(sc, bg=self.colors['white'])
        res.pack(fill='x', pady=10)
        self.sensitivity_results['carbon_cost'] = tk.Label(res, text=self.lm.tr("carbon_cost_label", "Karbon Maliyeti: -"), bg=self.colors['white'], font=('Segoe UI', 10))
        self.sensitivity_results['carbon_cost'].pack(anchor='w')
        self.sensitivity_results['energy_change'] = tk.Label(res, text=self.lm.tr("energy_cost_change_label", "Enerji Maliyet Değişimi: -"), bg=self.colors['white'], font=('Segoe UI', 10))
        self.sensitivity_results['energy_change'].pack(anchor='w')
        self.sensitivity_results['net_impact'] = tk.Label(res, text=self.lm.tr("net_impact_label", "Net Etki: -"), bg=self.colors['white'], font=('Segoe UI', 11, 'bold'))
        self.sensitivity_results['net_impact'].pack(anchor='w', pady=(5,0))

        scenario_card = tk.Frame(content, bg=self.colors['white'], relief='raised', bd=1)
        scenario_card.pack(fill='x', pady=10)
        sch = tk.Frame(scenario_card, bg=self.colors['primary'], height=40)
        sch.pack(fill='x')
        sch.pack_propagate(False)
        tk.Label(sch, text=self.lm.tr("scenario_time_series", "Senaryo Zaman Serisi (2025–2035)"), font=('Segoe UI', 11, 'bold'), bg=self.colors['primary'], fg='white').pack(side='left', padx=15, pady=8)
        scc = tk.Frame(scenario_card, bg=self.colors['white'])
        scc.pack(fill='both', expand=True, padx=15, pady=15)
        tk.Label(scc, text=self.lm.tr("scenario_selection", "Senaryo Seçimi"), bg=self.colors['white']).grid(row=0, column=0, sticky='w')
        self.scenario_name_var = tk.StringVar(value="Net Zero 2050")
        scen_values = ["Net Zero 2050", "Below 2°C", "Current Policies", "Hot House World"]
        scen_combo = ttk.Combobox(scc, textvariable=self.scenario_name_var, values=scen_values, state='readonly', width=28)
        scen_combo.grid(row=0, column=1, padx=8, pady=6, sticky='w')
        tk.Button(scc, text=self.lm.tr("generate", " Oluştur"), command=self.generate_scenario_timeseries, bg=self.colors['info'], fg='white').grid(row=0, column=2, padx=8, pady=6)
        self.scenario_table = ttk.Treeview(scc, columns=("year","price","cost","net"), show='headings')
        for col, text in [("year", self.lm.tr("year", "Yıl")),("price", self.lm.tr("carbon_price", "Karbon Fiyatı")),("cost", self.lm.tr("carbon_cost_col", "Karbon Maliyeti")),("net", self.lm.tr("net_impact", "Net Etki"))]:
            self.scenario_table.heading(col, text=text)
            self.scenario_table.column(col, width=140)
        self.scenario_table.grid(row=1, column=0, columnspan=3, sticky='nsew', pady=10)
        scc.grid_rowconfigure(1, weight=1)
        scc.grid_columnconfigure(2, weight=1)
        fig = Figure(figsize=(6,2.4), dpi=100)
        self.scenario_fig = fig
        self.scenario_ax = fig.add_subplot(111)
        self.scenario_canvas = FigureCanvasTkAgg(fig, master=scc)
        self.scenario_canvas.get_tk_widget().grid(row=2, column=0, columnspan=3, sticky='nsew')

        # Kaydet butonu
        btn_frame = tk.Frame(content, bg=self.colors['bg'])
        btn_frame.pack(fill='x', pady=20)
        tk.Button(
            btn_frame,
            text=self.lm.tr("save_strat_data", " Strateji Verilerini Kaydet"),
            command=self.save_strategy,
            bg=self.colors['success'],
            fg='white',
            font=('Segoe UI', 11, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=30,
            pady=12
        ).pack()

    # ========================================================================
    # 3. RISK MANAGEMENT TAB
    # ========================================================================

    def create_risk_tab(self) -> None:
        """Risk yönetimi sekmesi"""
        frame = tk.Frame(self.tab_risk, bg=self.colors['bg'])
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Form
        form_card = tk.Frame(frame, bg=self.colors['white'], relief='raised', bd=1)
        form_card.pack(fill='x')

        header = tk.Frame(form_card, bg=self.colors['primary'], height=40)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text=self.lm.tr("tcfd_risk_title", "İklim Riski Tanımla"), font=('Segoe UI', 11, 'bold'), bg=self.colors['primary'], fg='white').pack(side='left', padx=15, pady=8)

        content = tk.Frame(form_card, bg=self.colors['white'])
        content.pack(fill='both', expand=True, padx=15, pady=15)
        self.risk_widgets = {}

        # Kategoriler ve türler
        category_frame = tk.Frame(content, bg=self.colors['white'])
        category_frame.pack(fill='x', pady=5)
        tk.Label(category_frame, text=self.lm.tr("risk_category", "Risk Kategorisi:"), bg=self.colors['white']).pack(anchor='w')
        self.risk_widgets['risk_category'] = ttk.Combobox(category_frame, values=["Transition", "Physical"], state='readonly')
        self.risk_widgets['risk_category'].pack(fill='x')
        self.risk_widgets['risk_category'].set("Transition")

        type_frame = tk.Frame(content, bg=self.colors['white'])
        type_frame.pack(fill='x', pady=5)
        tk.Label(type_frame, text=self.lm.tr("risk_type", "Risk Türü:"), bg=self.colors['white']).pack(anchor='w')
        self.risk_widgets['risk_type'] = ttk.Combobox(type_frame, values=["policy_and_legal", "technology", "market", "reputation"], state='readonly')
        self.risk_widgets['risk_type'].pack(fill='x')
        self.risk_widgets['risk_type'].set("policy_and_legal")

        # Temel metin alanları
        for label, key in [
            (self.lm.tr("risk_name", "Risk Adı:"), 'risk_name'),
            (self.lm.tr("risk_desc", "Açıklama:"), 'risk_description'),
            (self.lm.tr("current_controls", "Mevcut Kontroller:"), 'current_controls'),
            (self.lm.tr("mitigation_actions", "Azaltım Aksiyonları:"), 'mitigation_actions'),
            (self.lm.tr("responsible_party", "Sorumlu Birim/Kişi:"), 'responsible_party'),
            (self.lm.tr("notes", "Notlar:"), 'notes'),
        ]:
            f = tk.Frame(content, bg=self.colors['white'])
            f.pack(fill='x', pady=5)
            tk.Label(f, text=label, bg=self.colors['white']).pack(anchor='w')
            if key in ('risk_description', 'current_controls', 'mitigation_actions', 'notes'):
                w = scrolledtext.ScrolledText(f, height=3, font=('Segoe UI', 10), relief='solid', bd=1, wrap='word')
            else:
                w = tk.Entry(f, font=('Segoe UI', 10), relief='solid', bd=1)
            w.pack(fill='x')
            self.risk_widgets[key] = w

        # Sayısal ve seçimler
        for label, key, values in [
            (self.lm.tr("likelihood", "Olasılık:"), 'likelihood', ["Very Low", "Low", "Medium", "High", "Very High"]),
            (self.lm.tr("impact", "Etki:"), 'impact', ["Very Low", "Low", "Medium", "High", "Very High"]),
            (self.lm.tr("time_horizon", "Zaman Ufku:"), 'time_horizon', ["Short", "Medium", "Long"]),
            (self.lm.tr("status", "Durum:"), 'status', ["Identified", "Analyzing", "Mitigating", "Closed"]),
        ]:
            f = tk.Frame(content, bg=self.colors['white'])
            f.pack(fill='x', pady=5)
            tk.Label(f, text=label, bg=self.colors['white']).pack(anchor='w')
            w = ttk.Combobox(f, values=values, state='readonly')
            w.set(values[0])
            w.pack(fill='x')
            self.risk_widgets[key] = w

        # Skor ve finans
        for label, key in [
            (self.lm.tr("likelihood_score", "Olasılık Skoru (1-5):"), 'likelihood_score'),
            (self.lm.tr("impact_score", "Etki Skoru (1-5):"), 'impact_score'),
            (self.lm.tr("financial_impact_low", "Tahmini Finansal Etki - Düşük:"), 'financial_impact_low'),
            (self.lm.tr("financial_impact_high", "Tahmini Finansal Etki - Yüksek:"), 'financial_impact_high'),
        ]:
            f = tk.Frame(content, bg=self.colors['white'])
            f.pack(fill='x', pady=5)
            tk.Label(f, text=label, bg=self.colors['white']).pack(anchor='w')
            w = tk.Entry(f, font=('Segoe UI', 10), relief='solid', bd=1)
            w.pack(fill='x')
            self.risk_widgets[key] = w

        # Tarih
        f = tk.Frame(content, bg=self.colors['white'])
        f.pack(fill='x', pady=5)
        tk.Label(f, text=self.lm.tr("action_deadline", "Aksiyon Son Tarihi (YYYY-MM-DD):"), bg=self.colors['white']).pack(anchor='w')
        self.risk_widgets['action_deadline'] = tk.Entry(f, font=('Segoe UI', 10), relief='solid', bd=1)
        self.risk_widgets['action_deadline'].pack(fill='x')

        btns = tk.Frame(content, bg=self.colors['white'])
        btns.pack(fill='x', pady=10)
        tk.Button(btns, text=self.lm.tr("add_risk", " Risk Ekle"), command=self.add_risk, bg=self.colors['success'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=5)
        tk.Button(btns, text=self.lm.tr("delete_selected_risk", " Seçili Riski Sil"), command=self.delete_selected_risk, bg=self.colors['danger'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=5)

        # Liste
        list_card = tk.Frame(frame, bg=self.colors['white'], relief='raised', bd=1)
        list_card.pack(fill='both', expand=True, pady=10)
        lh = tk.Frame(list_card, bg=self.colors['primary'], height=40)
        lh.pack(fill='x')
        lh.pack_propagate(False)
        tk.Label(lh, text=self.lm.tr("defined_climate_risks", "Tanımlı İklim Riskleri"), font=('Segoe UI', 11, 'bold'), bg=self.colors['primary'], fg='white').pack(side='left', padx=15, pady=8)

        lc = tk.Frame(list_card, bg=self.colors['white'])
        lc.pack(fill='both', expand=True, padx=15, pady=15)
        self.risk_tree = ttk.Treeview(lc, columns=("name", "category", "rating", "score", "time", "status"), show='headings')
        for col, text in [("name", "Risk Adı"), ("category", "Kategori"), ("rating", "Derece"), ("score", "Skor"), ("time", "Zaman"), ("status", "Durum")]:
            self.risk_tree.heading(col, text=text)
            self.risk_tree.column(col, width=140)
        self.risk_tree.pack(fill='both', expand=True)

        self.load_risks()

    # ========================================================================
    # 4. METRICS TAB
    # ========================================================================

    def create_metrics_tab(self) -> None:
        """Metrikler ve hedefler sekmesi"""
        canvas = tk.Canvas(self.tab_metrics, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_metrics, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=self.colors['bg'])
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        content = tk.Frame(scroll_frame, bg=self.colors['bg'])
        content.pack(fill='both', expand=True, padx=10, pady=10)
        tk.Label(content, text=self.lm.tr("metrics_title", "Metrikler"), font=('Segoe UI', 14, 'bold'), bg=self.colors['bg']).pack(anchor='w', pady=(0, 10))

        self.metrics_widgets = {}
        # Emisyonlar
        self.create_section_card(content, "GHG Emisyonları", [
            ("Scope 1 (tCO2e):", 'scope1_emissions', 'entry', None),
            ("Scope 2 (tCO2e):", 'scope2_emissions', 'entry', None),
            ("Scope 3 (tCO2e):", 'scope3_emissions', 'entry', None),
            ("Toplam Emisyon:", 'total_emissions', 'entry', None),
        ])

        # Yoğunluk
        self.create_section_card(content, "Emisyon Yoğunluğu", [
            ("Yoğunluk Değeri:", 'emissions_intensity', 'entry', None),
            ("Metriği:", 'intensity_metric', 'entry', None),
        ])

        # Enerji/Su
        self.create_section_card(content, "Enerji ve Su", [
            ("Toplam Enerji (MWh):", 'total_energy_consumption', 'entry', None),
            ("Yenilenebilir Enerji (%):", 'renewable_energy_pct', 'entry', None),
            ("Enerji Yoğunluğu:", 'energy_intensity', 'entry', None),
            ("Su Tüketimi (m3):", 'water_consumption', 'entry', None),
            ("Su Yoğunluğu:", 'water_intensity', 'entry', None),
        ])

        # Karbon fiyatlama ve finansal metrikler
        self.create_section_card(content, "Karbon Fiyatlama ve Finansal Metrikler", [
            ("Dahili Karbon Fiyatı ($/tCO2e):", 'internal_carbon_price', 'entry', None),
            ("Kapsam (%):", 'carbon_price_coverage', 'entry', None),
            ("İklim İlgili Gelir (mn):", 'climate_related_revenue', 'entry', None),
            ("İşletme Gideri (mn):", 'climate_related_opex', 'entry', None),
            ("Yatırım (mn):", 'climate_related_capex', 'entry', None),
        ])

        self.create_section_card(content, "Diğer Metrikler (JSON)", [
            ("Diğer Metrikler:", 'other_metrics', 'text', 4),
        ])

        btn_frame = tk.Frame(content, bg=self.colors['bg'])
        btn_frame.pack(fill='x', pady=10)
        tk.Button(btn_frame, text=" Metrikleri Kaydet", command=self.save_metrics, bg=self.colors['success'], fg='white', relief='flat', padx=20, pady=10).pack(side='left')

        sens_card = tk.Frame(content, bg=self.colors['white'], relief='raised', bd=1)
        sens_card.pack(fill='x', pady=10)
        sf = tk.Frame(sens_card, bg=self.colors['white'])
        sf.pack(fill='both', expand=True, padx=15, pady=15)
        tk.Label(sf, text="Hassasiyet Analizi (Karbon/Enerji)", font=('Segoe UI', 12, 'bold'), bg=self.colors['white']).pack(anchor='w')

        self.sensitivity_vars = {}
        self.sensitivity_results = {}

        row1 = tk.Frame(sf, bg=self.colors['white'])
        row1.pack(fill='x', pady=6)
        tk.Label(row1, text="Karbon Fiyat Hassasiyeti (%)", bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['carbon_pct'] = tk.IntVar(value=0)
        carbon_scale = ttk.Scale(row1, from_=-30, to=30, orient='horizontal', variable=self.sensitivity_vars['carbon_pct'], command=lambda v: self.calculate_sensitivity())
        carbon_scale.pack(fill='x')
        self.sensitivity_results['carbon_cost'] = tk.Label(row1, text="Karbon Maliyeti: $0", bg=self.colors['white'])
        self.sensitivity_results['carbon_cost'].pack(anchor='w', pady=4)

        row2 = tk.Frame(sf, bg=self.colors['white'])
        row2.pack(fill='x', pady=6)
        tk.Label(row2, text="Enerji Fiyat Hassasiyeti (%)", bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['energy_pct'] = tk.IntVar(value=0)
        energy_scale = ttk.Scale(row2, from_=-30, to=30, orient='horizontal', variable=self.sensitivity_vars['energy_pct'], command=lambda v: self.calculate_sensitivity())
        energy_scale.pack(fill='x')
        self.sensitivity_results['energy_change'] = tk.Label(row2, text="Enerji Maliyet Değişimi: $0", bg=self.colors['white'])
        self.sensitivity_results['energy_change'].pack(anchor='w', pady=4)

        row3 = tk.Frame(sf, bg=self.colors['white'])
        row3.pack(fill='x', pady=6)
        tk.Label(row3, text="Mevcut Enerji Fiyatı (USD/MWh)", bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['current_energy_price'] = tk.Entry(row3, font=('Segoe UI', 10), relief='solid', bd=1)
        self.sensitivity_vars['current_energy_price'].insert(0, "100")
        self.sensitivity_vars['current_energy_price'].pack(fill='x')
        self.sensitivity_results['net_impact'] = tk.Label(row3, text="Net Etki: $0", bg=self.colors['white'])
        self.sensitivity_results['net_impact'].pack(anchor='w', pady=4)

        row4 = tk.Frame(sf, bg=self.colors['white'])
        row4.pack(fill='x', pady=6)
        tk.Button(row4, text=" Hesapla", command=self.calculate_sensitivity, bg=self.colors['info'], fg='white', relief='flat').pack(side='left', padx=5)
        tk.Button(row4, text=" Şablon Dışa Aktar", command=self.export_sensitivity_template_xlsx, bg=self.colors['warning'], fg='white', relief='flat').pack(side='left', padx=5)
        tk.Button(row4, text=" Veri İçe Aktar", command=self.import_sensitivity_data, bg=self.colors['success'], fg='white', relief='flat').pack(side='left', padx=5)
        url_frame = tk.Frame(sf, bg=self.colors['white'])
        url_frame.pack(fill='x', pady=6)
        tk.Label(url_frame, text="Veri URL", bg=self.colors['white']).pack(anchor='w')
        self.sensitivity_vars['data_url'] = tk.Entry(url_frame, font=('Segoe UI', 10), relief='solid', bd=1)
        self.sensitivity_vars['data_url'].pack(fill='x')
        tk.Button(url_frame, text=" URL’den Çek", command=self.fetch_sensitivity_data, bg=self.colors['secondary'], fg='white', relief='flat').pack(side='left', pady=6)

        # Targets
        tk.Label(content, text="Hedefler", font=('Segoe UI', 14, 'bold'), bg=self.colors['bg']).pack(anchor='w', pady=(20, 10))
        target_card = tk.Frame(content, bg=self.colors['white'], relief='raised', bd=1)
        target_card.pack(fill='x')
        tc = tk.Frame(target_card, bg=self.colors['white'])
        tc.pack(fill='both', expand=True, padx=15, pady=15)
        self.target_widgets = {}

        for label, key in [
            ("Hedef Adı:", 'target_name'),
            ("Kategori:", 'target_category'),
            ("Tür:", 'target_type'),
            ("Baz Yıl:", 'baseline_year'),
            ("Baz Değer:", 'baseline_value'),
            ("Hedef Yıl:", 'target_year'),
            ("Hedef Değer:", 'target_value'),
            ("Azaltma %:", 'reduction_pct'),
            ("Kapsam:", 'scope'),
            ("Sınır:", 'boundary'),
            ("Güncel Değer:", 'current_value'),
            ("İlerleme %:", 'progress_pct'),
        ]:
            f = tk.Frame(tc, bg=self.colors['white'])
            f.pack(fill='x', pady=3)
            tk.Label(f, text=label, bg=self.colors['white']).pack(anchor='w')
            w = tk.Entry(f, font=('Segoe UI', 10), relief='solid', bd=1)
            w.pack(fill='x')
            self.target_widgets[key] = w

        # On track ve SBTi
        check_frame = tk.Frame(tc, bg=self.colors['white'])
        check_frame.pack(fill='x', pady=5)
        self.target_widgets['on_track_var'] = tk.BooleanVar(value=True)
        tk.Checkbutton(check_frame, text="Hedefte", variable=self.target_widgets['on_track_var'], bg=self.colors['white']).pack(side='left')
        self.target_widgets['sbti_var'] = tk.BooleanVar(value=False)
        tk.Checkbutton(check_frame, text="SBTi Onaylı", variable=self.target_widgets['sbti_var'], bg=self.colors['white']).pack(side='left', padx=10)

        btns = tk.Frame(tc, bg=self.colors['white'])
        btns.pack(fill='x', pady=8)
        tk.Button(btns, text=" Hedef Ekle", command=self.add_target, bg=self.colors['success'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=5)
        tk.Button(btns, text=" Seçili Hedefi Sil", command=self.delete_selected_target, bg=self.colors['danger'], fg='white', relief='flat', padx=20, pady=8).pack(side='left', padx=5)

        # Target list
        self.target_tree = ttk.Treeview(content, columns=("name", "category", "type", "year", "value", "progress", "status"), show='headings')
        for col, text in [("name", "Hedef"), ("category", "Kategori"), ("type", "Tür"), ("year", "Hedef Yıl"), ("value", "Değer"), ("progress", "İlerleme %"), ("status", "Durum")]:
            self.target_tree.heading(col, text=text)
            self.target_tree.column(col, width=120)
        self.target_tree.pack(fill='both', expand=True, pady=10)

        self.load_metrics()
        self.load_targets()

    # ========================================================================
    # GENEL FONKSİYONLAR
    # ========================================================================

    def on_year_changed(self) -> None:
        """Yıl değiştiğinde verileri yeniden yükle"""
        self.load_governance()
        self.load_strategy()
        self.load_risks()
        self.load_metrics()
        self.load_targets()

    def create_report(self) -> None:
        """TCFD raporu oluştur"""
        try:
            # Dosya kaydetme diyaloğu
            default_name = f"TCFD_Raporu_{self.company_id}_{self.year_var.get()}.pdf"
            output_path = filedialog.asksaveasfilename(
                title=self.lm.tr('save_tcfd_report', "TCFD Raporu Kaydet"),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[(self.lm.tr('file_pdf', "PDF Dosyası"), "*.pdf")]
            )

            if not output_path:
                return

            # Generator
            generator = TCFDReportGenerator(self.manager, self.calculator)
            
            # Firma adı
            company_name = self.lm.tr('sustainability_report', "Sürdürülebilirlik Raporu")
            if self.user and len(self.user) > 2:
                company_name = f"Firma {self.company_id}"

            # Rapor oluştur
            success = generator.generate_pdf_report(
                self.company_id,
                int(self.year_var.get()),
                company_name,
                output_path
            )

            if success:
                messagebox.showinfo(self.lm.tr('success', "Başarılı"), f"{self.lm.tr('report_created_success', 'Rapor başarıyla oluşturuldu')}:\n{output_path}")
                try:
                    os.startfile(output_path)
                except Exception as e:
                    logging.error(f"Silent error caught: {str(e)}")
            else:
                messagebox.showerror(self.lm.tr('error', "Hata"), self.lm.tr('report_creation_failed_lib', "Rapor oluşturulamadı. Lütfen ReportLab kütüphanesinin kurulu olduğundan emin olun."))

        except Exception as e:
            messagebox.showerror(self.lm.tr('error', "Hata"), f"{self.lm.tr('report_creation_error', 'Rapor oluşturma hatası')}:\n{e}")

    # ============================
    # Strategy helpers
    # ============================
    def save_strategy(self) -> None:
        try:
            data = {}
            for key, widget in getattr(self, 'gov_widgets', {}).items():
                # gov_widgets also used by governance; keep separate for strategy
                pass
            for key, widget in self.strategy_widgets.items():
                if isinstance(widget, tk.Entry):
                    data[key] = widget.get()
                elif isinstance(widget, scrolledtext.ScrolledText):
                    data[key] = widget.get('1.0', 'end-1c')
                else:
                    data[key] = getattr(widget, 'get', lambda: None)()
            success, message = self.manager.save_strategy(self.company_id, self.year_var.get(), data)
            if success:
                messagebox.showinfo("Başarılı", message)
            else:
                messagebox.showerror("Hata", message)
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası:\n{e}")

    def calculate_sensitivity(self) -> None:
        try:
            scope1 = 0.0
            scope2 = 0.0
            scope3 = 0.0
            carbon_price_base = 0.0
            total_energy = 0.0
            renewable_pct = 0.0

            if hasattr(self, 'metrics_widgets'):
                def _get_float(key):
                    w = self.metrics_widgets.get(key)
                    if isinstance(w, tk.Entry):
                        val = w.get()
                        try:
                            return float(val) if val not in (None, '') else 0.0
                        except ValueError:
                            return 0.0
                    return 0.0

                scope1 = _get_float('scope1_emissions')
                scope2 = _get_float('scope2_emissions')
                scope3 = _get_float('scope3_emissions')
                carbon_price_base = _get_float('internal_carbon_price')
                total_energy = _get_float('total_energy_consumption')
                renewable_pct = _get_float('renewable_energy_pct')

            carbon_adj = self.sensitivity_vars['carbon_pct'].get() if 'carbon_pct' in self.sensitivity_vars else 0
            energy_adj = self.sensitivity_vars['energy_pct'].get() if 'energy_pct' in self.sensitivity_vars else 0
            cep_widget = self.sensitivity_vars.get('current_energy_price')
            current_energy_price = 0.0
            if isinstance(cep_widget, tk.Entry):
                try:
                    val = cep_widget.get()
                    current_energy_price = float(val) if val not in (None, '') else 0.0
                except ValueError:
                    current_energy_price = 0.0

            adjusted_carbon_price = carbon_price_base * (1 + carbon_adj / 100)
            future_energy_price = current_energy_price * (1 + energy_adj / 100)

            carbon_impact = self.calculator.calculate_carbon_price_impact(
                scope1_emissions=scope1,
                scope2_emissions=scope2,
                scope3_emissions=scope3,
                carbon_price=adjusted_carbon_price,
                scope3_included=True
            )

            energy_impact = self.calculator.calculate_energy_cost_impact(
                current_energy_consumption=total_energy,
                current_energy_price=current_energy_price,
                future_energy_price=future_energy_price,
                renewable_transition_pct=renewable_pct,
                renewable_price_premium=0.0
            )

            total_costs = carbon_impact.get('total_cost', 0.0) + max(0.0, energy_impact.get('cost_change', 0.0))
            net_impact = 0.0 - total_costs

            if 'carbon_cost' in self.sensitivity_results:
                self.sensitivity_results['carbon_cost'].configure(text=f"Karbon Maliyeti: ${carbon_impact.get('total_cost', 0.0):,.0f}")
            if 'energy_change' in self.sensitivity_results:
                self.sensitivity_results['energy_change'].configure(text=f"Enerji Maliyet Değişimi: ${energy_impact.get('cost_change', 0.0):,.0f}")
            if 'net_impact' in self.sensitivity_results:
                self.sensitivity_results['net_impact'].configure(text=f"Net Etki: ${net_impact:,.0f}")

        except Exception as e:
            messagebox.showerror("Hata", f"Hesaplama hatası:\n{e}")

    def import_sensitivity_data(self) -> None:
        try:
            path = filedialog.askopenfilename(title=self.lm.tr('select_data_file', "Veri Dosyası Seç"), filetypes=[(self.lm.tr('file_json', "JSON Dosyaları"), "*.json"), (self.lm.tr('file_csv', "CSV Dosyası"), "*.csv"), (self.lm.tr('filetype_all', "Tüm Dosyalar"), "*.*")])
            if not path:
                return
            data = {}
            if path.lower().endswith('.json'):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            elif path.lower().endswith('.csv'):
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    if rows:
                        data = rows[0]
            elif path.lower().endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                headers = [c.value for c in ws[1] if c.value]
                row_dict = {}
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    values = [cell.value for cell in row[:len(headers)]]
                    if any(v not in (None, '') for v in values):
                        row_dict = {headers[i]: values[i] for i in range(len(values))}
                        break
                data = row_dict
            self.apply_imported_data(data)
        except Exception as e:
            messagebox.showerror("Hata", f"İçe aktarma hatası:\n{e}")

    def fetch_sensitivity_data(self) -> None:
        try:
            url_widget = self.sensitivity_vars.get('data_url')
            url = url_widget.get() if isinstance(url_widget, tk.Entry) else None
            if not url:
                return
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            self.apply_imported_data(data)
        except Exception as e:
            messagebox.showerror("Hata", f"Veri getirme hatası:\n{e}")

    def apply_imported_data(self, data: Dict) -> None:
        try:
            def set_entry(w, val):
                if isinstance(w, tk.Entry) and val is not None:
                    w.delete(0, 'end')
                    w.insert(0, str(val))
            m = getattr(self, 'metrics_widgets', {})
            set_entry(m.get('internal_carbon_price'), data.get('internal_carbon_price') or data.get('carbon_price'))
            set_entry(m.get('scope1_emissions'), data.get('scope1_emissions'))
            set_entry(m.get('scope2_emissions'), data.get('scope2_emissions'))
            set_entry(m.get('scope3_emissions'), data.get('scope3_emissions'))
            set_entry(m.get('total_energy_consumption'), data.get('total_energy_consumption') or data.get('energy_consumption'))
            set_entry(m.get('renewable_energy_pct'), data.get('renewable_energy_pct'))
            cep = self.sensitivity_vars.get('current_energy_price')
            set_entry(cep, data.get('current_energy_price') or data.get('energy_price'))
        except Exception as e:
            messagebox.showerror("Hata", f"Veri uygulama hatası:\n{e}")

    def export_sensitivity_template_xlsx(self) -> None:
        try:
            default_name = f"tcfd_sensitivity_template_{self.year_var.get()}.xlsx"
            path = filedialog.asksaveasfilename(title=self.lm.tr('save_template', "Şablon Kaydet"), defaultextension=".xlsx", initialfile=default_name, filetypes=[(self.lm.tr('excel_files', "Excel Dosyaları"), "*.xlsx")])
            if not path:
                return
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "SensitivityData"
            headers = [
                'internal_carbon_price',
                'scope1_emissions',
                'scope2_emissions',
                'scope3_emissions',
                'total_energy_consumption',
                'renewable_energy_pct',
                'current_energy_price'
            ]
            ws.append(headers)
            wb.save(path)
            messagebox.showinfo("Şablon", f"Şablon oluşturuldu:\n{path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Şablon oluşturma hatası:\n{e}")

    def load_strategy(self) -> None:
        try:
            data = self.manager.get_strategy(self.company_id, self.year_var.get())
            if not data:
                return
            for key, widget in self.strategy_widgets.items():
                value = data.get(key)
                if value is None:
                    continue
                if isinstance(widget, tk.Entry):
                    widget.delete(0, 'end')
                    widget.insert(0, str(value))
                elif isinstance(widget, scrolledtext.ScrolledText):
                    widget.delete('1.0', 'end')
                    widget.insert('1.0', str(value))
        except Exception as e:
            logging.info(f"[WARN] Strateji verileri yüklenemedi: {e}")

    # ============================
    # Risk helpers
    # ============================
    def add_risk(self) -> None:
        try:
            d = {}
            for key, w in self.risk_widgets.items():
                if isinstance(w, tk.Entry):
                    d[key] = w.get() or None
                elif isinstance(w, scrolledtext.ScrolledText):
                    d[key] = w.get('1.0', 'end-1c') or None
                elif isinstance(w, ttk.Combobox):
                    d[key] = w.get() or None
            # numeric
            for k in ('likelihood_score', 'impact_score', 'financial_impact_low', 'financial_impact_high'):
                if d.get(k) not in (None, ''):
                    try:
                        d[k] = float(d[k])
                    except ValueError:
                        d[k] = None
            ok, msg, rid = self.manager.add_climate_risk(self.company_id, self.year_var.get(), d)
            if ok:
                messagebox.showinfo("Başarılı", msg)
                self.load_risks()
            else:
                messagebox.showerror("Hata", msg)
        except Exception as e:
            messagebox.showerror("Hata", f"Risk ekleme hatası:\n{e}")

    def load_risks(self) -> None:
        try:
            self.risk_tree.delete(*self.risk_tree.get_children())
            risks = self.manager.get_climate_risks(self.company_id, self.year_var.get())
            for r in risks:
                self.risk_tree.insert('', 'end', iid=r['id'], values=(
                    r.get('risk_name'), r.get('risk_category'), r.get('risk_rating'), r.get('risk_score'), r.get('time_horizon'), r.get('status')
                ))
        except Exception as e:
            logging.info(f"[WARN] Riskler yüklenemedi: {e}")

    def delete_selected_risk(self) -> None:
        try:
            sel = self.risk_tree.selection()
            if not sel:
                return
            rid = int(sel[0])
            ok, msg = self.manager.delete_climate_risk(rid)
            if ok:
                self.risk_tree.delete(sel[0])
                messagebox.showinfo("Başarılı", msg)
            else:
                messagebox.showerror("Hata", msg)
        except Exception as e:
            messagebox.showerror("Hata", f"Silme hatası:\n{e}")

    # ============================
    # Metrics helpers
    # ============================
    def save_metrics(self) -> None:
        try:
            d = {}
            for key, w in self.metrics_widgets.items():
                if isinstance(w, tk.Entry):
                    d[key] = w.get() or None
                elif isinstance(w, scrolledtext.ScrolledText):
                    d[key] = w.get('1.0', 'end-1c') or None
            # Convert numerics when possible
            for k in [
                'scope1_emissions','scope2_emissions','scope3_emissions','total_emissions',
                'emissions_intensity','total_energy_consumption','renewable_energy_pct','energy_intensity',
                'water_consumption','water_intensity','internal_carbon_price','carbon_price_coverage',
                'climate_related_revenue','climate_related_opex','climate_related_capex'
            ]:
                if d.get(k) not in (None, ''):
                    try:
                        d[k] = float(d[k])
                    except ValueError as e:
                        logging.error(f'Silent error in tcfd_gui.py: {str(e)}')
            ok, msg = self.manager.save_metrics(self.company_id, self.year_var.get(), d)
            if ok:
                messagebox.showinfo("Başarılı", msg)
            else:
                messagebox.showerror("Hata", msg)
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası:\n{e}")

    def load_metrics(self) -> None:
        try:
            data = self.manager.get_metrics(self.company_id, self.year_var.get())
            if not data:
                return
            for key, w in self.metrics_widgets.items():
                v = data.get(key)
                if v is None:
                    continue
                if isinstance(w, tk.Entry):
                    w.delete(0, 'end')
                    w.insert(0, str(v))
                elif isinstance(w, scrolledtext.ScrolledText):
                    w.delete('1.0', 'end')
                    w.insert('1.0', str(v))
        except Exception as e:
            logging.info(f"[WARN] Metrikler yüklenemedi: {e}")

    def add_target(self) -> None:
        try:
            d = {k: w.get() for k, w in self.target_widgets.items() if isinstance(w, tk.Entry)}
            d['on_track'] = self.target_widgets['on_track_var'].get()
            d['sbti_approved'] = self.target_widgets['sbti_var'].get()
            # numeric conversions
            for k in ('baseline_year','target_year'):
                if d.get(k):
                    try:
                        d[k] = int(d[k])
                    except ValueError as e:
                        logging.error(f'Silent error in tcfd_gui.py: {str(e)}')
            for k in ('baseline_value','target_value','reduction_pct','current_value','progress_pct'):
                if d.get(k):
                    try:
                        d[k] = float(d[k])
                    except ValueError as e:
                        logging.error(f'Silent error in tcfd_gui.py: {str(e)}')
            ok, msg, tid = self.manager.add_target(self.company_id, d)
            if ok:
                messagebox.showinfo("Başarılı", msg)
                self.load_targets()
            else:
                messagebox.showerror("Hata", msg)
        except Exception as e:
            messagebox.showerror("Hata", f"Hedef ekleme hatası:\n{e}")

    def load_targets(self) -> None:
        try:
            self.target_tree.delete(*self.target_tree.get_children())
            targets = self.manager.get_targets(self.company_id)
            for t in targets:
                self.target_tree.insert('', 'end', iid=t['id'], values=(
                    t.get('target_name'), t.get('target_category'), t.get('target_type'), t.get('target_year'),
                    t.get('target_value'), t.get('progress_pct'), t.get('status')
                ))
        except Exception as e:
            logging.info(f"[WARN] Hedefler yüklenemedi: {e}")

    def delete_selected_target(self) -> None:
        try:
            sel = self.target_tree.selection()
            if not sel:
                return
            tid = int(sel[0])
            ok, msg = self.manager.delete_target(tid)
            if ok:
                self.target_tree.delete(sel[0])
                messagebox.showinfo("Başarılı", msg)
            else:
                messagebox.showerror("Hata", msg)
        except Exception as e:
            messagebox.showerror("Hata", f"Silme hatası:\n{e}")


# Test
def main():
    """Test için ana fonksiyon"""
    root = tk.Tk()
    root.title("TCFD Modülü - Test")
    root.geometry("1200x800")

    # Test kullanıcısı
    user = (1, "test_user", "Test User")
    company_id = 1

    # TCFD GUI
    TCFDGUI(root, user, company_id)

    root.mainloop()


if __name__ == "__main__":
    main()


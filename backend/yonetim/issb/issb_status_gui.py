#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import logging
"""
ISSB Durum Panosu GUI (Yönetim)

Şirket/yıl filtreleri ile IFRS S1/S2 uyumluluk KPI’larını gösterir.
"""
import os
import sqlite3
import tkinter as tk
from tkinter import messagebox, ttk
from typing import Optional

from modules.issb.issb_manager import ISSBManager
from modules.issb.issb_report_generator import ISSBReportGenerator
from modules.validation.advanced_validator import AdvancedDataValidator
from utils.tooltip import add_rich_tooltip
from config.icons import Icons


class ISSBStatusDashboardGUI:
    """ISSB Durum Panosu - Yönetim sekmesi"""

    def __init__(self, parent, default_company_id: int = 1) -> None:
        self.parent = parent
        self.company_id = default_company_id
        self.period_year = tk.StringVar(value="2025")

        # Varsayılan şirket DB yolu
        self.db_path_template = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data", "companies", "{company_id}", "company.db"))
        self.manager = self._build_manager(self.company_id)

        self._setup_ui()
        self._load_metrics()

    def _build_manager(self, company_id: int) -> Optional[ISSBManager]:
        """Seçilen şirket için ISSBManager oluştur"""
        db_path = self.db_path_template.format(company_id=company_id)
        try:
            # DB yoksa kullanıcıya bilgi göster, yine de manager döndürelim ki tablo oluşturabilsin
            if not os.path.exists(db_path):
                os.makedirs(os.path.dirname(db_path), exist_ok=True)
            return ISSBManager(db_path)
        except Exception as e:
            messagebox.showerror("Hata", f"ISSBManager başlatılamadı: {e}")
            return None

    def _setup_ui(self) -> None:
        """Arayüzü kur"""
        container = tk.Frame(self.parent, bg='#ffffff')
        container.pack(fill='both', expand=True)

        # Başlık alanı
        header = tk.Frame(container, bg='#2c3e50', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text=f"{Icons.REPORT} ISSB Durum Panosu", font=('Segoe UI', 16, 'bold'), fg='white', bg='#2c3e50').pack(side='left', padx=20)

        # Filtre alanı
        filters = tk.Frame(container, bg='#f8f9fa', relief='solid', bd=1)
        filters.pack(fill='x', padx=20, pady=15)

        tk.Label(filters, text="Şirket ID:", font=('Segoe UI', 10, 'bold'), bg='#f8f9fa', fg='#2c3e50').grid(row=0, column=0, padx=10, pady=10, sticky='w')
        self.company_entry = tk.Entry(filters, width=8)
        self.company_entry.insert(0, str(self.company_id))
        self.company_entry.grid(row=0, column=1, padx=5, pady=10, sticky='w')
        try:
            add_rich_tooltip(self.company_entry, title="Şirket ID", text="Takip edilen şirketin sistem içindeki numarası.", example="Örn: 1")
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

        tk.Label(filters, text="Rapor Yılı:", font=('Segoe UI', 10, 'bold'), bg='#f8f9fa', fg='#2c3e50').grid(row=0, column=2, padx=15, pady=10, sticky='w')
        self.period_entry = tk.Entry(filters, textvariable=self.period_year, width=8)
        self.period_entry.grid(row=0, column=3, padx=5, pady=10, sticky='w')
        try:
            add_rich_tooltip(self.period_entry, title="Rapor Yılı", text="ISSB rapor yılı (YYYY).", example="Örn: 2025")
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

        refresh_btn = tk.Button(filters, text=" Yenile ", font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white', relief='flat', cursor='hand2', command=self._on_refresh)
        refresh_btn.grid(row=0, column=4, padx=15, pady=8)
        validate_btn = tk.Button(filters, text=" Validasyonları Çalıştır ", font=('Segoe UI', 10, 'bold'), bg='#16a34a', fg='white', relief='flat', cursor='hand2', command=self._run_validations)
        validate_btn.grid(row=0, column=5, padx=5, pady=8)

        tk.Label(filters, text="Durum:", font=('Segoe UI', 10, 'bold'), bg='#f8f9fa', fg='#2c3e50').grid(row=1, column=0, padx=10, pady=8, sticky='w')
        self.status_filter_var = tk.StringVar(value="")
        self.status_filter = ttk.Combobox(filters, textvariable=self.status_filter_var, values=["","Beklemede","Devam","Tamamlandı"], width=18, state='readonly')
        self.status_filter.grid(row=1, column=1, padx=5, pady=8, sticky='w')
        try:
            add_rich_tooltip(self.status_filter, title="Aksiyon Durumu", text="Excel aksiyon planındaki durumlara göre filtreler.", example="Beklemede, Devam, Tamamlandı")
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

        self.overdue_only_var = tk.BooleanVar(value=False)
        self.overdue_only = ttk.Checkbutton(filters, text="Sadece Son Tarihi Geçmiş", variable=self.overdue_only_var)
        self.overdue_only.grid(row=1, column=2, padx=15, pady=8, sticky='w')

        tk.Label(filters, text="Departman:", font=('Segoe UI', 10, 'bold'), bg='#f8f9fa', fg='#2c3e50').grid(row=1, column=3, padx=10, pady=8, sticky='w')
        self.department_var = tk.StringVar(value="")
        self.department_filter = ttk.Combobox(filters, textvariable=self.department_var, values=[], width=18, state='disabled')
        self.department_filter.grid(row=1, column=4, padx=5, pady=8, sticky='w')

        tk.Label(filters, text="Sorumlu:", font=('Segoe UI', 10, 'bold'), bg='#f8f9fa', fg='#2c3e50').grid(row=1, column=5, padx=10, pady=8, sticky='w')
        self.responsible_var = tk.StringVar(value="")
        self.responsible_filter = ttk.Combobox(filters, textvariable=self.responsible_var, values=[], width=18, state='disabled')
        self.responsible_filter.grid(row=1, column=6, padx=5, pady=8, sticky='w')

        # KPI kartları alanı
        kpi_container = tk.Frame(container, bg='#ffffff')
        kpi_container.pack(fill='x', padx=20, pady=(0, 10))

        self.s1_card = self._create_kpi_card(kpi_container, title="IFRS S1 Uyumluluk", color='#10b981')
        self.s1_card.pack(side='left', fill='x', expand=True, padx=10)
        self.s2_card = self._create_kpi_card(kpi_container, title="IFRS S2 Uyumluluk", color='#f59e0b')
        self.s2_card.pack(side='left', fill='x', expand=True, padx=10)
        self.overall_card = self._create_kpi_card(kpi_container, title="Genel Uyumluluk", color='#6366f1')
        self.overall_card.pack(side='left', fill='x', expand=True, padx=10)
        self.plan_card = self._create_kpi_card(kpi_container, title="Aksiyon Planı", color='#ef4444')
        self.plan_card.pack(side='left', fill='x', expand=True, padx=10)
        self.quality_card = self._create_kpi_card(kpi_container, title="Veri Kalitesi", color='#0ea5e9')
        self.quality_card.pack(side='left', fill='x', expand=True, padx=10)
        try:
            layout_frame = tk.Frame(kpi_container, bg='#ffffff')
            layout_frame.pack(side='right', padx=10)
            tk.Label(layout_frame, text="KPI Düzeni:", bg='#ffffff').pack(side='left')
            self.dashboard_layout_var = tk.StringVar(value='Standart')
            ttk.Combobox(layout_frame, textvariable=self.dashboard_layout_var, values=['Standart','Minimal','Detaylı'], width=10, state='readonly').pack(side='left', padx=6)
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

        # Detay alanı (özet metin)
        details = tk.Frame(container, bg='white', relief='solid', bd=1)
        details.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        tk.Label(details, text="Özet", font=('Segoe UI', 12, 'bold'), bg='white', fg='#2c3e50').pack(anchor='w', padx=15, pady=10)
        self.summary_text = tk.Text(details, height=10, bg='white', fg='#2c3e50')
        self.summary_text.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        self.summary_text.configure(state='disabled')

        # S1/S2 eksik listeleri paneli
        lists_panel = tk.Frame(container, bg='white')
        lists_panel.pack(fill='both', expand=True, padx=20, pady=(0, 10))

        # S1 Eksikler
        s1_frame = tk.LabelFrame(lists_panel, text="IFRS S1 Eksikler", font=('Segoe UI', 11, 'bold'), bg='white')
        s1_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        self.s1_tree = ttk.Treeview(s1_frame, columns=(
            'Kod','Başlık','Kategori','Durum','Son Güncelleme'
        ), show='headings', height=10)
        for col, width in [('Kod',120),('Başlık',200),('Kategori',120),('Durum',120),('Son Güncelleme',140)]:
            self.s1_tree.heading(col, text=col)
            self.s1_tree.column(col, width=width)
        self.s1_header_tips = {
            'Kod': 'IFRS S1 gereksinim kodu; tekil tanımlayıcıdır ve izlenebilirlik için kullanılır.',
            'Başlık': 'Kısa başlık; rapor bölümlerinde ve aksiyon planında referans olarak görünür.',
            'Kategori': 'Üst kategori; Genel, Risk, Strateji gibi gruplamayı belirtir.',
            'Durum': 'İlerleme durumu; Başlanmayan, Devam ediyor, Tamamlanan.',
            'Son Güncelleme': 'Son değişiklik tarihi; denetim ve takip için önemli.'
        }
        s1_scroll = ttk.Scrollbar(s1_frame, orient='vertical', command=self.s1_tree.yview)
        self.s1_tree.configure(yscrollcommand=s1_scroll.set)
        self.s1_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        s1_scroll.pack(side='right', fill='y', pady=10)
        self.s1_tree.bind('<Motion>', lambda e: self._on_tree_motion(e, self.s1_tree, self.s1_header_tips))
        self.s1_tree.bind('<Leave>', lambda e: self.hide_hover_tooltip())

        # S1 bağlantıları
        s1_links = tk.Frame(s1_frame, bg='white')
        s1_links.pack(fill='x', padx=10, pady=(0,10))
        tk.Button(s1_links, text=" DOCX’i Aç ", bg='#2563eb', fg='white', relief='flat', command=self._open_s1_docx).pack(side='left', padx=(0,10))
        tk.Button(s1_links, text=" Excel’i Aç ", bg='#16a34a', fg='white', relief='flat', command=self._open_s1_excel).pack(side='left')

        # S2 Eksikler
        s2_frame = tk.LabelFrame(lists_panel, text="IFRS S2 Eksikler", font=('Segoe UI', 11, 'bold'), bg='white')
        s2_frame.pack(side='left', fill='both', expand=True, padx=(10, 0))
        self.s2_tree = ttk.Treeview(s2_frame, columns=(
            'Kod','Başlık','Kategori','Durum','Son Güncelleme'
        ), show='headings', height=10)
        for col, width in [('Kod',120),('Başlık',200),('Kategori',120),('Durum',120),('Son Güncelleme',140)]:
            self.s2_tree.heading(col, text=col)
            self.s2_tree.column(col, width=width)
        self.s2_header_tips = {
            'Kod': 'IFRS S2 açıklama kodu; iklimle ilgili gereksinimler için tekil tanımlayıcıdır.',
            'Başlık': 'Kısa açıklama başlığı; iklim beyanlarında referans olarak kullanılır.',
            'Kategori': 'Yönetim, Risk, Metrik, Hedef ana grubu.',
            'Durum': 'İlerleme durumu; Başlanmayan, Devam ediyor, Tamamlanan.',
            'Son Güncelleme': 'Son değişiklik tarihi; doğrulama ve takip için önemli.'
        }
        s2_scroll = ttk.Scrollbar(s2_frame, orient='vertical', command=self.s2_tree.yview)
        self.s2_tree.configure(yscrollcommand=s2_scroll.set)
        self.s2_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        s2_scroll.pack(side='right', fill='y', pady=10)
        self.s2_tree.bind('<Motion>', lambda e: self._on_tree_motion(e, self.s2_tree, self.s2_header_tips))
        self.s2_tree.bind('<Leave>', lambda e: self.hide_hover_tooltip())

        # S2 bağlantıları
        s2_links = tk.Frame(s2_frame, bg='white')
        s2_links.pack(fill='x', padx=10, pady=(0,10))
        tk.Button(s2_links, text=" DOCX’i Aç ", bg='#2563eb', fg='white', relief='flat', command=self._open_s2_docx).pack(side='left', padx=(0,10))
        tk.Button(s2_links, text=" Excel’i Aç ", bg='#16a34a', fg='white', relief='flat', command=self._open_s2_excel).pack(side='left')

    def _create_kpi_card(self, parent, title: str, color: str):
        frame = tk.Frame(parent, bg='white', relief='solid', bd=1)
        tk.Label(frame, text=title, font=('Segoe UI', 12, 'bold'), bg='white', fg='#2c3e50').pack(anchor='w', padx=12, pady=(10, 0))
        value_label = tk.Label(frame, text="0.0%", font=('Segoe UI', 20, 'bold'), bg='white', fg=color)
        value_label.pack(anchor='w', padx=12, pady=(0, 10))
        # Count satırı
        count_label = tk.Label(frame, text="Tamamlanan: 0 | Devam: 0 | Başlanmayan: 0", font=('Segoe UI', 10), bg='white', fg='#64748b')
        count_label.pack(anchor='w', padx=12, pady=(0, 12))
        # Referanslar
        frame.value_label = value_label
        frame.count_label = count_label
        trend_canvas = tk.Canvas(frame, height=40, bg='white', highlightthickness=0)
        trend_canvas.pack(fill='x', padx=12, pady=(0, 12))
        frame.trend_canvas = trend_canvas
        return frame

    def _load_metrics(self) -> None:
        """KPI’ları yükle ve özet oluştur"""
        if not self.manager:
            return
        try:
            status = self.manager.get_company_issb_status(self.company_id)
            s1 = status.get('ifrs_s1', {})
            s2 = status.get('ifrs_s2', {})
            s1_rate = float(s1.get('compliance_rate', 0) or 0)
            s2_rate = float(s2.get('compliance_rate', 0) or 0)
            overall = (s1_rate + s2_rate) / 2.0

            # KPI label’larını güncelle
            self.s1_card.value_label.configure(text=f"{s1_rate:.1f}%")
            self.s1_card.count_label.configure(text=f"Tamamlanan: {s1.get('completed', 0)} | Devam: {s1.get('in_progress', 0)} | Başlanmayan: {s1.get('not_started', 0)}")

            self.s2_card.value_label.configure(text=f"{s2_rate:.1f}%")
            self.s2_card.count_label.configure(text=f"Tamamlanan: {s2.get('completed', 0)} | Devam: {s2.get('in_progress', 0)} | Başlanmayan: {s2.get('not_started', 0)}")

            self.overall_card.value_label.configure(text=f"{overall:.1f}%")
            self.overall_card.count_label.configure(text="S1+S2 ort.")

            try:
                year_text = self.period_year.get().strip()
                year_val = int(year_text) if year_text.isdigit() and len(year_text) == 4 else None
                gen = ISSBReportGenerator()
                dept = self.department_var.get().strip() or None
                resp = self.responsible_var.get().strip() or None
                stat = self.status_filter_var.get().strip() or None
                overdue = bool(self.overdue_only_var.get())
                summary = gen.get_action_plan_status_summary_with_filters(self.company_id, year_val, department=dept, responsible=resp, status_filter=stat, overdue_only=overdue) if year_val else {"tamamlanan":0,"devam":0,"beklemede":0,"son_tarih_gecmis":0}
                total = summary["tamamlanan"] + summary["devam"] + summary["beklemede"]
                rate = (summary["tamamlanan"] / total * 100.0) if total > 0 else 0.0
                self.plan_card.value_label.configure(text=f"{rate:.1f}%")
                layout_val = getattr(self, 'dashboard_layout_var', None)
                if layout_val and layout_val.get() == 'Minimal':
                    self.plan_card.count_label.configure(text=f"Tam:{summary['tamamlanan']}")
                else:
                    self.plan_card.count_label.configure(text=f"Tam:{summary['tamamlanan']} | Dev:{summary['devam']} | Bek:{summary['beklemede']} | Geç:{summary['son_tarih_gecmis']}")
                monthly = gen.get_action_plan_monthly_completion(self.company_id, year_val) if year_val else [0]*12
                targets = gen.get_action_plan_monthly_targets(self.company_id, year_val) if year_val else [0]*12
                s1m, s2m = self._get_s1_s2_monthly_completion(year_val) if year_val else ([0]*12, [0]*12)
                self._draw_trend(self.plan_card.trend_canvas, monthly, s1m, s2m, targets)
            except Exception:
                self.plan_card.value_label.configure(text="0.0%")
                self.plan_card.count_label.configure(text="Tam:0 | Dev:0 | Bek:0 | Geç:0")
                self._draw_trend(self.plan_card.trend_canvas, [0]*12, [0]*12, [0]*12, [0]*12)

            try:
                year_text = self.period_year.get().strip()
                year_val = int(year_text) if year_text.isdigit() and len(year_text) == 4 else None
                validator = AdvancedDataValidator()
                modules = ['karbon','enerji','su']
                scores = []
                if year_val:
                    for mname in modules:
                        res = validator.calculate_quality_score(self.company_id, mname, year_val)
                        scores.append(res.get('overall_score', 0.0))
                avg_score = sum(scores)/len(scores) if scores else 0.0
                vsummary = validator.get_validation_summary(self.company_id, year_val) if year_val else {
                    'total_errors': 0, 'total_warnings': 0, 'total_missing': 0, 'total_anomalies': 0
                }
                self.quality_card.value_label.configure(text=f"{avg_score:.1f}")
                self.quality_card.count_label.configure(text=f"Hata:{vsummary.get('total_errors',0)} | Uyari:{vsummary.get('total_warnings',0)} | Eksik:{vsummary.get('total_missing',0)} | Anomali:{vsummary.get('total_anomalies',0)}")
                self._draw_trend(self.quality_card.trend_canvas, [vsummary.get('total_errors',0)]*12)
            except Exception:
                self.quality_card.value_label.configure(text="0.0")
                self.quality_card.count_label.configure(text="Hata:0 | Uyari:0 | Eksik:0 | Anomali:0")
                self._draw_trend(self.quality_card.trend_canvas, [0]*12)

            # Özet metni
            summary = (
                f"Şirket: {self.company_id}\n"
                f"Rapor Yılı: {self.period_year.get()}\n"
                f"Genel Uyumluluk: {overall:.1f}%\n"
                f"S1 Uyumluluk: {s1_rate:.1f}% (T:{s1.get('total',0)}, C:{s1.get('completed',0)}, D:{s1.get('in_progress',0)}, B:{s1.get('not_started',0)})\n"
                f"S2 Uyumluluk: {s2_rate:.1f}% (T:{s2.get('total',0)}, C:{s2.get('completed',0)}, D:{s2.get('in_progress',0)}, B:{s2.get('not_started',0)})\n"
            )
            self.summary_text.configure(state='normal')
            self.summary_text.delete('1.0', tk.END)
            self.summary_text.insert('1.0', summary)
            self.summary_text.configure(state='disabled')
            # Detay listelerini yükle
            self._load_detail_lists()
        except Exception as e:
            messagebox.showerror("Hata", f"ISSB metrikleri yüklenemedi: {e}")

    def _on_refresh(self) -> None:
        """Filtrelere göre yenile"""
        try:
            # Şirket ID doğrulama
            cid_text = self.company_entry.get().strip()
            if not cid_text.isdigit():
                messagebox.showwarning("Uyarı", "Geçerli bir şirket ID girin (örn. 1).")
                return
            self.company_id = int(cid_text)

            # Yıl doğrulama
            year = self.period_year.get().strip()
            if not (len(year) == 4 and year.isdigit()):
                messagebox.showwarning("Uyarı", "Rapor yılı 4 haneli olmalı (örn. 2025).")
                return

            # Manager’ı yeniden kur ve metrikleri yükle
            self.manager = self._build_manager(self.company_id)
            self._load_action_plan_filter_options()
            self._load_metrics()
        except Exception as e:
            messagebox.showerror("Hata", f"Yenileme hatası: {e}")

    def _run_validations(self) -> None:
        try:
            year = self.period_year.get().strip()
            if not (len(year) == 4 and year.isdigit()):
                messagebox.showwarning("Uyarı", "Rapor yılı 4 haneli olmalı (örn. 2025).")
                return
            year_val = int(year)
            validator = AdvancedDataValidator()
            for mname in ['karbon','enerji','su']:
                validator.check_missing_data(self.company_id, mname)
                validator.calculate_quality_score(self.company_id, mname, year_val)
            self._load_metrics()
            messagebox.showinfo("Bilgi", "Validasyonlar çalıştırıldı ve KPI güncellendi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Validasyon çalıştırma hatası: {e}")

    def _load_action_plan_filter_options(self) -> None:
        try:
            year = self.period_year.get().strip()
            if not (len(year) == 4 and year.isdigit()):
                return
            path = os.path.join("data","reports","issb", f"issb_action_plan_{self.company_id}_{year}.xlsx")
            if not os.path.exists(path):
                self.department_filter.configure(state='disabled', values=[])
                self.responsible_filter.configure(state='disabled', values=[])
                return
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            dep_idx = headers.index("Departman") if "Departman" in headers else None
            res_idx = headers.index("Sorumlu") if "Sorumlu" in headers else None
            deps = set()
            ress = set()
            if dep_idx is not None or res_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if dep_idx is not None:
                        v = row[dep_idx]
                        if v:
                            deps.add(str(v).strip())
                    if res_idx is not None:
                        v = row[res_idx]
                        if v:
                            ress.add(str(v).strip())
            dep_values = [""] + sorted(deps)
            res_values = [""] + sorted(ress)
            if dep_values and len(dep_values) > 1:
                self.department_filter.configure(state='readonly', values=dep_values)
            else:
                self.department_filter.configure(state='disabled', values=[])
            if res_values and len(res_values) > 1:
                self.responsible_filter.configure(state='readonly', values=res_values)
            else:
                self.responsible_filter.configure(state='disabled', values=[])
        except Exception:
            self.department_filter.configure(state='disabled', values=[])
            self.responsible_filter.configure(state='disabled', values=[])

    def _draw_trend(self, canvas: tk.Canvas, values, s1_values=None, s2_values=None, target_values=None):
        try:
            canvas.delete('all')
            w = canvas.winfo_width() or 300
            h = canvas.winfo_height() or 40
            n = max(1, len(values))
            all_vals = list(values)
            if s1_values:
                all_vals += list(s1_values)
            if s2_values:
                all_vals += list(s2_values)
            if target_values:
                all_vals += list(target_values)
            maxv = max(all_vals) if all_vals else 0
            bw = max(2, int(w / (n + 2)))
            x = 6
            for v in values:
                bh = 0 if maxv == 0 else int((v / maxv) * (h - 8))
                y1 = h - 4
                y0 = y1 - bh
                canvas.create_rectangle(x, y0, x + bw, y1, fill='#3b82f6', outline='')
                x += bw + 4
            def _line(vals, color):
                if not vals:
                    return
                pts = []
                x = 6 + bw // 2
                for v in vals:
                    ly = (h - 4) - (0 if maxv == 0 else int((v / maxv) * (h - 8)))
                    pts.append((x, ly))
                    x += bw + 4
                if len(pts) > 1:
                    canvas.create_line(*[c for p in pts for c in p], fill=color, width=2)
            if s1_values:
                _line(s1_values, '#10b981')
            if s2_values:
                _line(s2_values, '#f59e0b')
            if target_values:
                # Hedef çizgisi: gri ve kesik çizgi
                pts = []
                x = 6 + bw // 2
                for v in target_values:
                    ly = (h - 4) - (0 if maxv == 0 else int((v / maxv) * (h - 8)))
                    pts.append((x, ly))
                    x += bw + 4
                if len(pts) > 1:
                    # Kesik çizgi taklidi: segmentler
                    for i in range(len(pts)-1):
                        if i % 2 == 0:
                            canvas.create_line(pts[i][0], pts[i][1], pts[i+1][0], pts[i+1][1], fill='#6b7280', width=2)
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

    def _get_s1_s2_monthly_completion(self, year: int):
        try:
            db_path = self.db_path_template.format(company_id=self.company_id)
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            s1 = [0]*12
            s2 = [0]*12
            cur.execute("""
                SELECT substr(last_updated,6,2) AS m, COUNT(*)
                FROM ifrs_s1_requirements
                WHERE company_id = ? AND status = 'Completed' AND substr(last_updated,1,4) = ?
                GROUP BY m
            """, (self.company_id, str(year)))
            for m, c in cur.fetchall():
                try:
                    idx = int(m) - 1
                    if 0 <= idx < 12:
                        s1[idx] = c
                except Exception as e:
                    logging.error(f'Silent error in issb_status_gui.py: {str(e)}')
            cur.execute("""
                SELECT substr(last_updated,6,2) AS m, COUNT(*)
                FROM ifrs_s2_climate
                WHERE company_id = ? AND status = 'Completed' AND substr(last_updated,1,4) = ?
                GROUP BY m
            """, (self.company_id, str(year)))
            for m, c in cur.fetchall():
                try:
                    idx = int(m) - 1
                    if 0 <= idx < 12:
                        s2[idx] = c
                except Exception as e:
                    logging.error(f'Silent error in issb_status_gui.py: {str(e)}')
            conn.close()
            return s1, s2
        except Exception:
            return [0]*12, [0]*12

    def _on_tree_motion(self, event, tree, tips) -> None:
        try:
            region = tree.identify_region(event.x, event.y)
            if region != 'heading':
                self.hide_hover_tooltip()
                return
            col_id = tree.identify_column(event.x)
            if not col_id:
                self.hide_hover_tooltip()
                return
            heading = tree.heading(col_id).get('text', '')
            text = tips.get(heading)
            if not text:
                self.hide_hover_tooltip()
                return
            if getattr(self, '_last_tip_text', None) == text:
                return
            self._last_tip_text = text
            self._show_hover_tooltip(tree, text, event.x_root, event.y_root)
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

    def _load_detail_lists(self) -> None:
        """S1/S2 eksik listelerini yükle"""
        try:
            # Temizle
            for item in self.s1_tree.get_children():
                self.s1_tree.delete(item)
            for item in self.s2_tree.get_children():
                self.s2_tree.delete(item)
            # S1: Not Started / In Progress
            db_path = self.db_path_template.format(company_id=self.company_id)
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT requirement_code, requirement_title, category, status, last_updated
                FROM ifrs_s1_requirements
                WHERE company_id = ? AND status IN ('Not Started','In Progress')
                ORDER BY status DESC, last_updated DESC
                LIMIT 50
                """,
                (self.company_id,)
            )
            for row in cur.fetchall():
                self.s1_tree.insert('', 'end', values=row)
            # S2: Not Started / In Progress
            cur.execute(
                """
                SELECT disclosure_code, disclosure_title, category, status, last_updated
                FROM ifrs_s2_climate
                WHERE company_id = ? AND status IN ('Not Started','In Progress')
                ORDER BY status DESC, last_updated DESC
                LIMIT 50
                """,
                (self.company_id,)
            )
            for row in cur.fetchall():
                self.s2_tree.insert('', 'end', values=row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Detay listeleri yüklenemedi: {e}")

    def _open_s1_docx(self) -> None:
        try:
            year = self.period_year.get().strip()
            path = os.path.join("data","reports","issb", f"issb_{self.company_id}_{year}.docx")
            if os.path.exists(path):
                os.startfile(path)
            else:
                messagebox.showwarning("Uyarı", f"DOCX bulunamadı: {path}")
        except Exception as e:
            messagebox.showerror("Hata", f"DOCX açılamadı: {e}")

    def _open_s1_excel(self) -> None:
        try:
            year = self.period_year.get().strip()
            path = os.path.join("data","reports","issb", f"issb_action_plan_{self.company_id}_{year}.xlsx")
            if os.path.exists(path):
                os.startfile(path)
            else:
                messagebox.showwarning("Uyarı", f"Excel bulunamadı: {path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel açılamadı: {e}")

    def _open_s2_docx(self) -> None:
        # S1 ile aynı dosya; tek raporda S1+S2 var
        self._open_s1_docx()

    def _open_s2_excel(self) -> None:
        # S1 ile aynı action plan dosyası
        self._open_s1_excel()

    def _show_hover_tooltip(self, widget, text, x_root, y_root):
        try:
            if hasattr(self, '_hover_tip') and self._hover_tip:
                self._hover_tip.destroy()
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{x_root+12}+{y_root+12}")
            frame = tk.Frame(tip, background="#ffffe0", relief='solid', borderwidth=1)
            frame.pack()
            label = tk.Label(frame, text=text, background="#ffffe0", font=('Segoe UI', 9), justify='left', wraplength=300)
            label.pack(padx=8, pady=6)
            self._hover_tip = tip
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')

    def hide_hover_tooltip(self):
        try:
            if hasattr(self, '_hover_tip') and self._hover_tip:
                self._hover_tip.destroy()
                self._hover_tip = None
            self._last_tip_text = None
        except Exception as e:
            logging.error(f'Silent error in issb_status_gui.py: {str(e)}')
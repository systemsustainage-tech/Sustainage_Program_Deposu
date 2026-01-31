#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CDP Excel Dışa Aktarım
CDP verilerini Excel formatında dışa aktarır
"""

import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .cdp_data_collector import CDPDataCollector
from config.icons import Icons


class CDPExcelExporter:
    """CDP verileri için Excel dışa aktarım"""

    def __init__(self):
        self.collector = CDPDataCollector()

    def export_all(self, company_id: int, year: int, output_path: str) -> bool:
        """Tüm CDP verilerini Excel'e aktar"""
        try:
            logging.info(f"[CDP Excel] Excel raporu oluşturuluyor: {year}")

            wb = Workbook()
            wb.remove(wb.active)  # Varsayılan sheet'i kaldır

            # Şirket bilgileri
            company_info = self.collector.collect_company_info(company_id)

            # 1. Climate Change verileri
            self._add_climate_sheet(wb, company_id, year, company_info)

            # 2. Water Security verileri
            self._add_water_sheet(wb, company_id, year, company_info)

            # 3. Forests verileri
            self._add_forests_sheet(wb, company_id, year, company_info)

            # 4. Özet sayfa
            self._add_summary_sheet(wb, company_id, year, company_info)

            # Kaydet
            wb.save(output_path)
            logging.info(f"[CDP Excel] Excel kaydedildi: {output_path}")
            return True

        except Exception as e:
            logging.error(f"[HATA] CDP Excel oluşturulamadı: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _add_climate_sheet(self, wb: Workbook, company_id: int, year: int, company_info: dict):
        """Climate Change verileri sheet'i"""
        ws = wb.create_sheet("Climate Change")

        # Başlık
        ws['A1'] = 'CDP CLIMATE CHANGE RAPORU'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"{company_info.get('name', 'Şirket')} - {year}"
        ws['A2'].font = Font(size=11, italic=True)

        # Emisyon verileri
        climate_data = self.collector.collect_climate_data(company_id, year)
        emissions = climate_data.get('emissions', {})

        headers = ['Kapsam', 'Emisyon (tCO2e)', 'Oran (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        total = emissions.get('total', 0)

        data_rows = [
            ('Scope 1', emissions.get('scope1', 0)),
            ('Scope 2', emissions.get('scope2', 0)),
            ('Scope 3', emissions.get('scope3', 0)),
            ('TOPLAM', total)
        ]

        for row_idx, (scope, value) in enumerate(data_rows, 5):
            ws.cell(row=row_idx, column=1, value=scope)
            ws.cell(row=row_idx, column=2, value=value).number_format = '#,##0.00'
            if scope != 'TOPLAM' and total > 0:
                pct = (value / total) * 100
                ws.cell(row=row_idx, column=3, value=pct / 100).number_format = '0.0%'

            # Toplam satırını bold yap
            if scope == 'TOPLAM':
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).font = Font(bold=True)

        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _add_water_sheet(self, wb: Workbook, company_id: int, year: int, company_info: dict):
        """Water Security verileri sheet'i"""
        ws = wb.create_sheet("Water Security")

        # Başlık
        ws['A1'] = 'CDP WATER SECURITY RAPORU'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"{company_info.get('name', 'Şirket')} - {year}"

        # Su verileri
        water_data = self.collector.collect_water_data(company_id, year)
        water = water_data.get('water_consumption', {})

        headers = ['Kategori', 'Miktar (m³)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')

        data_rows = [
            ('Çekilen Su', water.get('withdrawn', 0)),
            ('Tüketilen Su', water.get('consumed', 0)),
            ('Deşarj Edilen Su', water.get('discharged', 0)),
            ('Geri Dönüştürülen Su', water.get('recycled', 0))
        ]

        for row_idx, (category, value) in enumerate(data_rows, 5):
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=value).number_format = '#,##0'

        # Geri dönüşüm oranı
        withdrawn = water.get('withdrawn', 0)
        recycled = water.get('recycled', 0)
        if withdrawn > 0:
            recycling_rate = (recycled / withdrawn) * 100
            ws['A10'] = 'Geri Dönüşüm Oranı'
            ws['A10'].font = Font(bold=True)
            ws['B10'] = recycling_rate / 100
            ws['B10'].number_format = '0.0%'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _add_forests_sheet(self, wb: Workbook, company_id: int, year: int, company_info: dict):
        """Forests verileri sheet'i"""
        ws = wb.create_sheet("Forests")

        ws['A1'] = 'CDP FORESTS RAPORU'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"{company_info.get('name', 'Şirket')} - {year}"

        ws['A4'] = 'Orman Ürünleri ve Risk Değerlendirmesi'
        ws['A4'].font = Font(bold=True)

        ws['A6'] = 'Bu modül, orman ürünleri kullanımı ve ormansızlaşma risklerini değerlendirir.'

        ws.column_dimensions['A'].width = 60

    def _add_summary_sheet(self, wb: Workbook, company_id: int, year: int, company_info: dict):
        """Özet sayfa"""
        ws = wb.create_sheet("Özet", 0)  # İlk sıraya ekle

        ws['A1'] = 'CDP KAPSAMLI RAPOR - ÖZET'
        ws['A1'].font = Font(size=16, bold=True)

        ws['A3'] = 'Şirket Bilgileri'
        ws['A3'].font = Font(size=12, bold=True)
        ws['A3'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

        info_rows = [
            ('Şirket Adı', company_info.get('name', '-')),
            ('Sektör', company_info.get('sector', '-')),
            ('Ülke', company_info.get('country', 'Türkiye')),
            ('Çalışan Sayısı', company_info.get('employees', 0)),
            ('Raporlama Yılı', year)
        ]

        for row_idx, (label, value) in enumerate(info_rows, 4):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # Modül özeti
        ws['A10'] = 'CDP Modülleri'
        ws['A10'].font = Font(size=12, bold=True)
        ws['A10'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

        modules = [
            'Climate Change - İklim Değişikliği',
            'Water Security - Su Güvenliği',
            'Forests - Ormanlar'
        ]

        for row_idx, module in enumerate(modules, 11):
            ws.cell(row=row_idx, column=1, value=Icons.PASS)
            ws.cell(row=row_idx, column=2, value=module)

        # Tarih
        ws['A15'] = f'Rapor Tarihi: {datetime.now().strftime("%d.%m.%Y")}'
        ws['A15'].font = Font(italic=True, size=9)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40


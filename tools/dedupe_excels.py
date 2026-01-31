"""
Excel dosyalarını karşılaştırıp fazlalıkları temizleme aracı.

Kapsam mantığı:
- "master" dosyanın satır imzaları (tüm hücre değerlerinin normalize edilmiş birleşimi) set olarak çıkarılır.
- Diğer her Excel için satır imzaları üretilir ve master setindeki karşılama oranı hesaplanır.
- Karşılama oranı eşikten (varsayılan 0.95) büyükse dosya fazlalık kabul edilir.

Kullanım:
python tools/dedupe_excels.py --docs-dir docs --master docs/SDG_16_169_232.xlsx --threshold 0.95 [--delete]

Notlar:
- Sadece .xlsx dosyaları değerlendirilir.
- Boş satırlar ve tek hücreli satırlar gürültü azaltmak için yok sayılır.
"""

import logging
import argparse
import hashlib
import os
from typing import Set, Tuple

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def normalize_value(val) -> str:
    if val is None:
        return ""
    s = str(val)
    # Küçük harfe çevir, baş/son boşlukları kırp, iç boşlukları sıkıştır
    s = s.strip().lower()
    # Çoklu boşlukları tek boşluğa indir
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def row_signature(values: Tuple[str, ...]) -> str:
    # Satır için deterministik bir imza üret (sha1)
    joined = "|".join(values)
    return hashlib.sha1(joined.encode("utf-8", errors="ignore")).hexdigest()


def workbook_row_signatures(path: str) -> Set[str]:
    sigs: Set[str] = set()
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        logging.info(f"Dosya açılamadı: {path} -> {e}")
        return sigs

    for ws in wb.worksheets:
        try:
            for row in ws.iter_rows(values_only=True):
                vals = tuple(normalize_value(v) for v in row)
                # Gürültüyü azalt: en az iki dolu hücre olmalı ve toplam uzunluk anlamlı olmalı
                non_empty = [v for v in vals if v]
                if len(non_empty) < 2:
                    continue
                if sum(len(v) for v in non_empty) < 5:
                    continue
                sigs.add(row_signature(vals))
        except Exception as e:
            logging.info(f"Sayfa okunamadı: {path}:{ws.title} -> {e}")
            continue
    return sigs


def compute_coverage(master_sigs: Set[str], other_sigs: Set[str]) -> float:
    if not other_sigs:
        return 0.0
    matched = len(other_sigs & master_sigs)
    return matched / len(other_sigs)


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel deduplikasyon aracı")
    parser.add_argument("--docs-dir", required=True, help="Excel dosyalarının bulunduğu dizin")
    parser.add_argument("--master", required=True, help="Referans master Excel dosyası")
    parser.add_argument("--threshold", type=float, default=0.95, help="Silme eşiği (kapsam oranı)")
    parser.add_argument("--delete", action="store_true", help="Eşiği geçen dosyaları sil")
    args = parser.parse_args()

    docs_dir = args.docs_dir
    master_path = args.master
    threshold = args.threshold
    do_delete = args.delete

    if not os.path.isfile(master_path):
        raise SystemExit(f"Master dosya bulunamadı: {master_path}")
    if not os.path.isdir(docs_dir):
        raise SystemExit(f"Dizin bulunamadı: {docs_dir}")

    logging.info(f"Master yükleniyor: {master_path}")
    master_sigs = workbook_row_signatures(master_path)
    logging.info(f"Master satır imzaları: {len(master_sigs)}")

    # Diğer .xlsx dosyalarını topla
    candidates = []
    for name in os.listdir(docs_dir):
        if not name.lower().endswith(".xlsx"):
            continue
        path = os.path.join(docs_dir, name)
        # Master'ı es geç
        if os.path.abspath(path) == os.path.abspath(master_path):
            continue
        candidates.append(path)

    logging.info(f"Toplam aday dosya: {len(candidates)}")
    to_delete = []
    for path in candidates:
        other_sigs = workbook_row_signatures(path)
        coverage = compute_coverage(master_sigs, other_sigs)
        logging.info(f"{os.path.basename(path)} -> kapsam: {coverage:.4f} (satır: {len(other_sigs)})")
        if coverage >= threshold:
            to_delete.append(path)

    logging.info("\nÖzet:")
    logging.info(f"Silme eşiği: {threshold}")
    logging.info(f"Silinecek adaylar: {len(to_delete)}")
    for p in to_delete:
        logging.info(f"- {os.path.basename(p)}")

    if do_delete and to_delete:
        for p in to_delete:
            try:
                os.remove(p)
                logging.info(f"SİLİNDİ: {p}")
            except Exception as e:
                logging.info(f"Silinemedi: {p} -> {e}")
    elif do_delete:
        logging.info("Eşiği geçen dosya yok; silme yapılmadı.")


if __name__ == "__main__":
    main()

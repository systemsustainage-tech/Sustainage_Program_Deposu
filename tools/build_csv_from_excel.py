#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Excel → CSV Oluşturucu (Heuristik Algılayıcı)
- SDG_16_169_232 gibi dosyalardan hedef/alt hedef/gösterge sütunlarını otomatik algılamayı dener
- Excel yoksa veya sütunlar bulunamazsa mevcut DB’den export’a geri düşer

Kullanım:
  python tools/build_csv_from_excel.py --db data/sdg_desktop.sqlite --excel C:\path\SDG_16_169_232.xlsx --outdir data/imports/generated
"""
import logging
import argparse
import os

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def fallback_export(db, outdir) -> None:
    # çağrıyı simulate edelim
    class Args:
        def __init__(self):
            pass
    args = Args()
    args.db = db
    args.outdir = outdir
    # export_full_csvs.py içindeki main'i yeniden kullanmak yerine subprocess gerekmiyor
    # doğrudan fonksiyon çağrısı için uygun değil; bu nedenle basit bir subprocess yaklaşımı önerilir.
    # Kullanıcı doğrudan export_full_csvs.py aracını çalıştırabilir.
    logging.info("Excel bulunamadı veya algılanamadı. Lütfen export_full_csvs.py aracını kullanın:")
    logging.info(f"  python tools/export_full_csvs.py --db {db} --outdir {outdir}")

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--excel", required=True)
    ap.add_argument("--outdir", required=True)
    args = ap.parse_args()

    if not os.path.exists(args.excel):
        logging.info("Excel dosyası bulunamadı:", args.excel)
        fallback_export(args.db, args.outdir)
        return

    try:
        # Hafif bağımlılıklar ortamda yoksa hataya düşmemek için kullanıcıya yönlendirme yapıyoruz.
        import openpyxl  # type: ignore
    except Exception:
        logging.info("openpyxl yüklü değil. Heuristik Excel okuma için openpyxl gerekli.")
        logging.info("Alternatif: DB’den tam CSV export:")
        fallback_export(args.db, args.outdir)
        return

    # Heuristik sütun eşlemeleri
    aliases = {
        'sdg_no': {"SDG No","SDG","Goal No","Amaç"},
        'goal_title': {"Başlık","SDG Başlık","Goal Title","Amaç Başlık"},
        'target_code': {"Alt Hedef Kodu","Target Code","Hedef Kodu"},
        'target_title': {"Alt Hedef Tanımı","Alt Hedef Tanımı (TR)","Target Title","Target Description"},
        'indicator_code': {"Gösterge Kodu","Indicator Code","Kod"},
        'indicator_title': {"Gösterge Tanımı","Gösterge Tanımı (TR)","Indicator Title"},
        'unit': {"Birim","Unit","KPI / Metrik"},
        'frequency': {"Sıklık","Frequency","Ölçüm Sıklığı"},
        'topic': {"Konu","Topic"},
        'q1': {"Soru 1","Q1"},
        'q2': {"Soru 2","Q2"},
        'q3': {"Soru 3","Q3"},
        'default_owner': {"Sorumlu","Sorumlu Birim/Kişi"},
        'default_source': {"Kaynak","Veri Kaynağı"}
    }

    wb = openpyxl.load_workbook(args.excel)
    # Basit yaklaşım: ilk sayfada göstergeler var varsayımı
    ws = wb.active
    headers = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {}
    for i,h in enumerate(headers):
        for k,alts in aliases.items():
            if h in alts:
                header_map[k] = i
    # Minimum gerekli alanlar
    required = {'sdg_no','target_code','indicator_code','indicator_title'}
    if not required.issubset(set(header_map.keys())):
        logging.info("Excel sütunları beklenen formatta değil. DB’den export’a geçiliyor.")
        fallback_export(args.db, args.outdir)
        return

    import csv
    os.makedirs(args.outdir, exist_ok=True)
    goals_path = os.path.join(args.outdir, "sdg_goals.csv")
    targets_path = os.path.join(args.outdir, "sdg_targets.csv")
    indicators_path = os.path.join(args.outdir, "sdg_indicators.csv")
    questions_path = os.path.join(args.outdir, "question_bank.csv")

    seen_goals = set()
    seen_targets = set()
    rows_goals = []
    rows_targets = []
    rows_indicators = []
    rows_questions = []

    for row in ws.iter_rows(min_row=2):
        def val(idx) -> None:
            v = row[idx].value if idx in header_map.values() else None
            return v if v is not None else ""
        sdg_no = row[ header_map['sdg_no'] ].value
        target_code = row[ header_map['target_code'] ].value
        indicator_code = row[ header_map['indicator_code'] ].value
        indicator_title = row[ header_map['indicator_title'] ].value
        goal_title = row[ header_map.get('goal_title', header_map['sdg_no']) ].value if header_map.get('goal_title') else None
        unit = row[ header_map.get('unit', header_map['indicator_code']) ].value if header_map.get('unit') else None
        freq = row[ header_map.get('frequency', header_map['indicator_code']) ].value if header_map.get('frequency') else None
        topic = row[ header_map.get('topic', header_map['indicator_code']) ].value if header_map.get('topic') else None
        q1 = row[ header_map.get('q1', header_map['indicator_code']) ].value if header_map.get('q1') else None
        q2 = row[ header_map.get('q2', header_map['indicator_code']) ].value if header_map.get('q2') else None
        q3 = row[ header_map.get('q3', header_map['indicator_code']) ].value if header_map.get('q3') else None
        owner = row[ header_map.get('default_owner', header_map['indicator_code']) ].value if header_map.get('default_owner') else None
        source = row[ header_map.get('default_source', header_map['indicator_code']) ].value if header_map.get('default_source') else None

        if sdg_no and sdg_no not in seen_goals:
            rows_goals.append([sdg_no, goal_title or ""])
            seen_goals.add(sdg_no)
        if sdg_no and target_code and (sdg_no, target_code) not in seen_targets:
            rows_targets.append([sdg_no, target_code, ""])
            seen_targets.add((sdg_no, target_code))
        rows_indicators.append([sdg_no, target_code, indicator_code, indicator_title or "", unit or "", freq or "Yıllık", topic or ""])
        if any([q1,q2,q3,owner,source]):
            rows_questions.append([indicator_code, q1 or "", q2 or "", q3 or "", unit or "", freq or "Yıllık", owner or "", source or ""]) 

    with open(goals_path,'w',newline='',encoding='utf-8') as f:
        wr = csv.writer(f)
        wr.writerow(["sdg_no","title_tr"])
        wr.writerows(rows_goals)
    with open(targets_path,'w',newline='',encoding='utf-8') as f:
        wr = csv.writer(f)
        wr.writerow(["sdg_no","target_code","title_tr"])
        wr.writerows(rows_targets)
    with open(indicators_path,'w',newline='',encoding='utf-8') as f:
        wr = csv.writer(f)
        wr.writerow(["sdg_no","target_code","indicator_code","title_tr","unit","frequency","topic"])
        wr.writerows(rows_indicators)
    if rows_questions:
        with open(questions_path,'w',newline='',encoding='utf-8') as f:
            wr = csv.writer(f)
            wr.writerow(["indicator_code","q1","q2","q3","default_unit","default_frequency","default_owner","default_source"])
            wr.writerows(rows_questions)
    logging.info("Excel’den CSV üretimi tamamlandı:", args.outdir)

if __name__ == "__main__":
    main()
